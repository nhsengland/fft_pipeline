from src.etl_functions import *
import pytest
import pandas as pd
import numpy as np
import os
import glob
from openpyxl import Workbook
from openpyxl.styles import NamedStyle
from unittest.mock import patch, MagicMock


# Mock file system for testing using monkeypath fixture to simulate presence of file directory.
# glob.glob method will return a predefined list of files that match the expected pattern.
@pytest.fixture
def mock_filesystem(monkeypatch):
    # Arrange: Mock glob.glob method to simulate finding files in a directory
    # https://www.geeksforgeeks.org/how-to-use-glob-function-to-find-files-recursively-in-python/
    mock_files = [
        '/mocked/path/IPFFT-Jan24.xlsx',
        '/mocked/path/IPFFT-Feb24.xlsx',
        '/mocked/path/IPFFT-Mar24.xlsx', 
    ]
    
    # Mock the result of glob.glob to return a controlled list of files
    # replace `glob.glob` function with a lambda to return `mock_files` simulating finding files based on mocked file pattern
    monkeypatch.setattr(glob, 'glob', lambda pattern: mock_files)
    
    # Mock os.path.join to avoid issues with different OS path formats
    # replace `os.path.join` with lambda function to avoid os path format issues ensuring test environment consistency
    monkeypatch.setattr(os.path, 'join', lambda *args: '/'.join(args))
    
    return mock_files

def test_list_excel_files_finds_matching_files(mock_filesystem):
    '''
    Test list_excel_files function returns all files of matching name format in descending order
    '''
    # Arrange: Define input parameters to pass to the function
    folder_path = '/mocked/path'
    file_pattern = 'IPFFT-*.xlsx'
    prefix_suffix_separator = '-'
    date_format = '%b%y'
    
    # Act: 
    result = list_excel_files(folder_path, file_pattern, prefix_suffix_separator, date_format)
    
    # Assert: Verify the returned file list is sorted by date in descending order
    expected_files = [
        '/mocked/path/IPFFT-Mar24.xlsx',
        '/mocked/path/IPFFT-Feb24.xlsx',
        '/mocked/path/IPFFT-Jan24.xlsx'
    ]
    assert result == expected_files

def test_list_excel_files_raises_error_when_no_files_found(monkeypatch):
    '''
    Test list_excel_files function rasies the expected error when no matching files are found
    '''
    # Arrange: Mock the glob.glob method to return an empty list (no files found)
    monkeypatch.setattr(glob, 'glob', lambda x: [])

    # Mock os.path.join 
    monkeypatch.setattr(os.path, 'join', lambda a, b: f'{a}/{b}')

    # Define input parameters to pass to the function
    folder_path = '/mocked/path'
    file_pattern = 'IPFFT-*.xlsx'
    prefix_suffix_separator = '-'
    date_format = '%b%y'

    # Act & Assert: Verify that calling the function raises a ValueError
    with pytest.raises(ValueError, match=r"No matching Excel files found in the specified folder."):
        list_excel_files(folder_path, file_pattern, prefix_suffix_separator, date_format)

def test_load_excel_sheet_loads_correctly(monkeypatch):
    '''
    Test _load_excel_sheet function correctly loads called sheet form file
    '''
    # Arrange: Create mock DataFrame and mock pandas.read_excel function
    mock_df = pd.DataFrame({"Site Code": ["ABC01", "DEF02"], "1 Very Good": [100, 50]})
    
    # Mocking pd.read_excel to return the mock DataFrame when called
    monkeypatch.setattr(pd, 'read_excel', lambda *args, **kwargs: mock_df)
    
    # Define the input parameters
    file_path = "mock_file.xlsx"
    sheet_name = "Parent_Self_Trusts_Site_Lev"
    
    # Act: Call the function with the mocked environment
    result_df = load_excel_sheet(file_path, sheet_name)
    
    # Assert: Check if the returned DataFrame matches the mock DataFrame unsing assert_frame_equal testing function
    pd.testing.assert_frame_equal(result_df, mock_df)

def test_load_excel_sheet_raises_error_when_sheet_not_found(monkeypatch):
    '''
    Test load_excel_sheet function rasies expected error when invalid sheet called
    '''
    # Arrange: Mock pd.read_excel to raise a ValueError when called
    def mock_read_excel(*args, **kwargs):
        raise ValueError("Sheet 'InvalidSheet' not found")
    
    monkeypatch.setattr(pd, 'read_excel', mock_read_excel)
    
    # Define input parameters
    file_path = "mock_file.xlsx"
    sheet_name = "InvalidSheet"
    
    # Act & Assert: Ensure that the ValueError is raised when calling the function
    with pytest.raises(ValueError, match=r"Sheet 'InvalidSheet' not found in the file."):
        load_excel_sheet(file_path, sheet_name)

def test_validate_column_length_on_single_column():
    '''
    Test validate_column_length function functions correctly on a single column
    '''
    # Arrange: test DataFrame 
    df = pd.DataFrame({'Org Code': ['BKR', 'GR01E', 'HER', 'DF34F'], 
                       'Org Name': ['Hospital 1', 'IS Hospital 1', 'Hospital 2', 'IS Hospital 2']})

    # Act & Assert that no error should be raised
    validate_column_length(df, 'Org Code', [3, 5])

def test_validate_column_length_multiple_columns_valid():
    '''
    Test validate_column_length function functions correctly on multiple column
    '''
    # Arrange: test DataFrame 
    df = pd.DataFrame({'Org Code': ['BKR', 'GR01E', 'HER', 'DF34F'],
                       'Site Code': ['BKR01', 'GRE', 'RED', 'DFF34'],
                       'Org Name': ['Hospital 1', 'IS Hospital 1', 'Hospital 2', 'IS Hospital 2']})

    # Act & Assert that no error should be raised
    validate_column_length(df, ['Org Code', 'Site Code'], [3, 5])

def test_validate_column_length_single_column_error():
    '''
    Test validate_column_length function generates error correctly on single column
    '''
    # Arrange: test DataFrame 
    df = pd.DataFrame({'Org Code': ['BKR', 'GR01E', 'HE', 'DF34F'], 
                       'Org Name': ['Hospital 1', 'IS Hospital 1', 'Hospital 2', 'IS Hospital 2']})

    # Act & Assert that ValueError should be raised for incorrect lengths
    with pytest.raises(ValueError, match=r"Row 2 in column 'Org Code' contains a value with invalid length."):
        validate_column_length(df, 'Org Code', [3, 5])

def test_validate_column_length_multiple_column_error():
    '''
    Test validate_column_length function generates error correctly on multiple columns
    '''
    # Arrange: test DataFrame 
    df = pd.DataFrame({'Org Code': ['BKR', 'GR01E', 'HER', 'DF34F'],
                       'Site Code': ['BKR0', 'GRE', 'RED', 'DFF34'],
                       'Org Name': ['Hospital 1', 'IS Hospital 1', 'Hospital 2', 'IS Hospital 2']})

    # Act & Assert that ValueError should be raised for incorrect lengths
    with pytest.raises(ValueError, match=r"Row 0 in column 'Site Code' contains a value with invalid length."):
        validate_column_length(df, ['Org Code', 'Site Code'], [3, 5])

def test_validate_column_length_single_length_multiple_columns_valid():
    '''
    Test validate_column_length function functions correctly for a single value on multiple columns
    '''
    # Arrange: test DataFrame 
    df = pd.DataFrame({'Org Code': ['BKR', 'GR0', 'HER', 'DF3'],
                       'Site Code': ['BK0', 'GRE', 'RED', 'DFF'],
                       'Org Name': ['Hospital 1', 'IS Hospital 1', 'Hospital 2', 'IS Hospital 2']})

    # Act & Assert that no error should be raised since all lengths are 3
    validate_column_length(df, ['Org Code', 'Site Code'], 3)

def test_validate_column_length_single_length_multiple_columns_error():
    '''
    Test validate_column_length function generates error correctly for a single value on multiple columns
    '''
    # Arrange: test DataFrame 
    df = pd.DataFrame({'Org Code': ['BKR', 'GR0', 'HER', 'DF3'],
                       'Site Code': ['BK0', 'GREF', 'RED', 'DFF'],
                       'Org Name': ['Hospital 1', 'IS Hospital 1', 'Hospital 2', 'IS Hospital 2']})

    # Act & Assert that ValueError should be raised since not all lengths are 3
    with pytest.raises(ValueError, match=r"Row 1 in column 'Site Code' contains a value with invalid length."):
        validate_column_length(df, ['Org Code', 'Site Code'], 3)

def test_validate_column_length_column_not_found():
    '''
    Test validate_column_length function generates correct error when column not found
    '''
    # Arrange: test DataFrame 
    df = pd.DataFrame({'Org Code': ['BKR', 'GRE', 'HER', 'DF3'],
                       'Site Code': ['BKR0', 'GRE', 'RED', 'DFF34'],
                       'Org Name': ['Hospital 1', 'IS Hospital 1', 'Hospital 2', 'IS Hospital 2']})

    # Act & Assert KeyError should be raised for column not found)
    with pytest.raises(KeyError, match=r"Column 'ICB Code' not found in DataFrame"):
        validate_column_length(df, ['Org Code', 'ICB Code'], 3)

def test_validate_numeric_columns_valid_single_column_int():
    '''
    Test validate_numeric_columns function functions correctly for a single integer column
    '''
    # Arrange: test DataFrame 
    df = pd.DataFrame({'1 Very Good': [255, 930, 459], 
                       '2 Good': [125, 300, 445]})

    # Act & Assert that no error should be raised for valid integers
    validate_numeric_columns(df, '2 Good', 'int')

def test_validate_numeric_columns_valid_single_column_float():
    '''
    Test validate_numeric_columns function functions correctly for a single float column
    '''
    # Arrange: test DataFrame 
    df = pd.DataFrame({'Prop_Pos': [98.5, 95.0, 96.7], 
                       'Prop_Neg': [12.5, 15.0, 13.7]})

    # Act & Assert that no error should be raised for valid floats
    validate_numeric_columns(df, 'Prop_Pos', 'float')

def test_validate_numeric_columns_valid_multiple_columns():
    '''
    Test validate_numeric_columns function functions correctly for multiple columns
    '''
    # Arrange: test DataFrame 
    df = pd.DataFrame({'Prop_Pos': [98.5, 95.0, 96.7], 
                       'Prop_Neg': [12.5, 15.0, 13.7]})

    # Act & Assert that no error should be raised for valid data across multiple columns
    validate_numeric_columns(df, ['Prop_Pos', 'Prop_Neg'], 'float')

def test_validate_numeric_columns_invalid_single_column():
    '''
    Test validate_numeric_columns function generates error correctly for a single column containing incorrect type
    '''
    # Arrange: test DataFrame 
    df = pd.DataFrame({'1 Very Good': [255, 'nine hundred', 459], 
                       '2 Good': [125, 300, 445]})

    # Act & Assert that a TypeError should be raised for non-integer value in 1 Very Good column
    with pytest.raises(TypeError, match=r"Row 1 in column '1 Very Good' contains a non-integer value."):
        validate_numeric_columns(df, '1 Very Good', 'int')

def test_validate_numeric_columns_invalid_multiple_columns():
    '''
    Test validate_numeric_columns function generates error correctly for multiple columns when containing incorrect type
    '''
    # Arrange: test DataFrame 
    df = pd.DataFrame({'Prop_Pos': [98.5, 95.0, 96.7], 
                       'Prop_Neg': [12.5, 'Ten', 13.7]})

    # Act & Assert that a TypeError should be raised for non-float value in 'Prop_Neg' column
    with pytest.raises(TypeError, match=r"Row 1 in column 'Prop_Neg' contains a non-float value."):
        validate_numeric_columns(df, ['Prop_Pos', 'Prop_Neg'], 'float')

# Test for invalid column name
def test_validate_numeric_columns_invalid_column_name():
    '''
    Test validate_numeric_columns function generates correct error when column not found
    '''
    # Arrange: test DataFrame 
    df = pd.DataFrame({'Prop_Pos': [98.5, 95.0, 96.7], 
                       'Prop_Neg': [12.5, 15.0, 13.7]})

    # Act & Assert that a KeyError should be raised for a missing column
    with pytest.raises(KeyError, match=r"Column 'Prop_Eligible' not found in DataFrame."):
        validate_numeric_columns(df, ['Prop_Pos', 'Prop_Neg', 'Prop_Eligible'], 'float')

def test_validate_numeric_columns_invalid_expected_type():
    '''
    Test validate_numeric_columns function generates correct error when presented with incorrect type
    '''
    # Arrange: test DataFrame 
    df = pd.DataFrame({'Prop_Pos': [98.5, 95.0, 96.7], 
                       'Prop_Neg': [12.5, 15.0, 13.7]})

    # Act & Assert that a TypeError should be raised  for invalid expected_type
    with pytest.raises(TypeError, match=r"Invalid expected_type 'string'. Must be 'int' or 'float'."):
        validate_numeric_columns(df, 'Prop_Pos', 'string')

def test_get_cell_content_as_string_valid():
    '''
    Test get_cell_content_as_string function functions correctly when presented with appropriate parameters/data
    '''
    # Arrange: test DataFrame 
    df = pd.DataFrame({"Periodname": ["FEBRUARY", "MARCH", "APRIL"], 
                     "Yearnumber": ["2022-23", "2023-24", "2024-25"]})
    source_row = 1 # Row index
    source_col = "Periodname" # Column name

    # Act: Call the function with valid inputs
    result = get_cell_content_as_string(df, source_row, source_col)

    # Assert: The result should be "MARCH" from the specified cell
    assert result == "MARCH"

def test_get_cell_content_as_string_column_not_found():
    '''
    Test get_cell_content_as_string generates KeyError when presented with column that doesn't exist
    '''
    # Arrange: test DataFrame 
    df = pd.DataFrame({"Periodname": ["FEBRUARY", "MARCH", "APRIL"],
                       "Yearnumber": ["2022-23", "2023-24", "2024-25"]})
    source_row = 1
    source_col = "NonExistentColumn" # column that does not exist in the DataFrame

    # Act & Assert: Expect a KeyError to be raised due to the missing column
    with pytest.raises(KeyError, match=r"Column 'NonExistentColumn' not found in the DataFrame"):
        get_cell_content_as_string(df, source_row, source_col)

def test_get_cell_content_as_string_row_out_of_range():
    '''
    Test get_cell_content_as_string generates IndexError when presented with row that doesn't exist
    '''
    # Arrange: test DataFrame 
    df = pd.DataFrame({
        "Periodname": ["FEBRUARY", "MARCH", "APRIL"],
        "Yearnumber": ["2022-23", "2023-24", "2024-25"]})
    source_row = 5 # This row index is out of bounds (only 3 rows in the DataFrame)
    source_col = "Yearnumber"

    # Act & Assert: Expect an IndexError due to the out-of-range row index
    with pytest.raises(IndexError, match=r"Row index 5 is out of range"):
        get_cell_content_as_string(df, source_row, source_col)

def test_map_fft_period_valid_july():
    '''
    Test map_fft_period generates expected output for months in first half of financial year    
    '''
    # Arrange: Set valid inputs for a period in the first half of the financial year
    periodname = "JULY"
    yearnumber = "2024-25"

    # Act: Call the function with valid inputs
    result = map_fft_period(periodname, yearnumber)

    # Assert: Check if the returned fft_period is correct ("Jul-24")
    assert result == "Jul-24"

def test_map_fft_period_valid_january():
    '''
    Test map_fft_period generates expected output for months in second half of financial year
    '''
    # Arrange: Set valid inputs for a period in the second half of the financial year
    periodname = "JANUARY"
    yearnumber = "2024-25"

    # Act: Call the function with valid inputs
    result = map_fft_period(periodname, yearnumber)

    # Assert: Check if the returned fft_period is correct ("Jan-25")
    assert result == "Jan-25"

def test_map_fft_period_invalid_periodname():
    '''
    Test map_fft_period generates ValueError when presented with invalid periodname
    '''
    # Arrange: Set an invalid period name
    periodname = "INVALIDMONTH"
    yearnumber = "2024-25"

    # Act & Assert: Expect a ValueError due to the invalid period name
    with pytest.raises(ValueError, match=r"Invalid period name 'INVALIDMONTH'."):
        map_fft_period(periodname, yearnumber)

def test_map_fft_period_invalid_yearnumber_length():
    '''
    Test map_fft_period generates ValueError when presented with invalid yearnumber
    '''
    # Arrange: Set an invalid yearnumber (incorrect length)
    periodname = "JANUARY"
    yearnumber = "2024-256" # Invalid length

    # Act & Assert: Expect a ValueError due to the incorrect length
    with pytest.raises(ValueError, match=r"Yearnumber mot int the correct format."):
        map_fft_period(periodname, yearnumber)

def test_remove_columns():
    '''
    Test remove_columns function removes the required columns and expected columns remain
    '''
    # Arrange: expected_columns_ppv_partners list at top of file contains all raw columns
    df = pd.DataFrame(columns=['Trust Code', 'Trust Name', 'Very Good', 'Good', 'Percentage Positive', 'Percentage Negative'])
    
    # Act:
    df = remove_columns(df, ['Very Good', 'Good'])
    
    # Assert: following df pass through function results will match below listed columns
    assert list(df.columns) == ['Trust Code', 'Trust Name', 'Percentage Positive', 'Percentage Negative']

def test_remove_non_existent_columns():
    '''
    Test remove_columns function generates KeyError when presented with non existent column names to remove
    '''   
    # Arrange: Create a DataFrame and define multiple columns, some of which do not exist
    df = pd.DataFrame(columns=['Trust Code', 'Trust Name', 'Very Good', 'Good', 'Percentage Positive', 'Percentage Negative'])
    columns_to_remove = ['Poor', 'Very Poor']
    
    # Act & Assert: Expect a KeyError for the missing columns
    with pytest.raises(KeyError, match=r"The following columns are not in the DataFrame: \['Poor', 'Very Poor'\]"):
        remove_columns(df, columns_to_remove)

def test_rename_columns():
    '''
    Test rename_columns function correctly renames all columns. 
    '''
    # Arrange: test DataFrame 
    df = pd.DataFrame(columns=['Org code', 'Org name', 'Prop Pos', 'Prop Neg'])
    
    # Define column renaming to apply 
    org_columns_to_rename = {
            'Org code': 'Trust Code', 
            'Org name': 'Trust Name',
            'Prop Pos': 'Percentage Positive',
            'Prop Neg': 'Percentage Negative'
        }
    
    # Act:
    df = rename_columns(df, org_columns_to_rename)
    
    # Assert:
    assert list(df.columns) == ['Trust Code', 'Trust Name', 'Percentage Positive', 'Percentage Negative']
    
def test_rename_non_existent_columns():
    '''
    Test rename_columns function generates KeyError when presented with non existent column names to rename
    '''       
    # Arrange: test DataFrame 
    df = pd.DataFrame(columns=['Org code', 'Org name', 'Prop Pos', 'Prop Neg'])
    
    # Define column renaimgn to apply
    org_columns_to_rename = {
            'Org code': 'Trust Code', 
            'Org name': 'Trust Name',
            'Prop Pos': 'Percentage Positive',
            'Prop Neg': 'Percentage Negative',
            '2 Good': 'Good',
            '4 Poor': 'Poor'
        }
    
    # Act & Assert: Expect a KeyError for the missing columns
    with pytest.raises(KeyError, match=r"The following columns to be renamed do not exist in the DataFrame: \['2 Good', '4 Poor'\]"):
        rename_columns(df, org_columns_to_rename)

def test_replace_non_matching_values():
    '''
    Test replace_no_matching_values function correctly replaces target values with replacement value
    '''
    # Arrange: test DataFrame with some non-matching values
    df = pd.DataFrame({'ICB Code': ['IS1', 'EDC', 'WSD', 'IS1', 'PLK']})
    column_name = 'ICB Code'
    target_value = 'IS1'
    replacement_value = 'NHS'

    # Act: Call the function to replace non-matching values
    result_df = replace_non_matching_values(df, column_name, target_value, replacement_value)

    # Assert: non-matching values ('NHS') are replaced with 'NHS', and 'IS1' values remain unchanged
    expected_data = {'ICB Code': ['IS1', 'NHS', 'NHS', 'IS1', 'NHS']}
    expected_df = pd.DataFrame(expected_data)
    assert result_df.equals(expected_df)

def test_replace_non_matching_values_presented_non_existent_column():
    '''
    Test replace_non_matching_values function gives KeyError when presented with non existent column
    '''
    # Arrange: test DataFrame that does not contain the specified column
    df = pd.DataFrame({'ICB Code': ['IS1', 'EDC', 'WSD', 'IS1', 'PLK']})
    column_name = 'NonExistentColumn'
    target_value = 'IS1'
    replacement_value = 'NHS'

    # Act & Assert: Expect a KeyError to be raised
    with pytest.raises(KeyError, match=r"Column 'NonExistentColumn' not found in the DataFrame"):
        replace_non_matching_values(df, column_name, target_value, replacement_value)

def test_sum_grouped_response_fields_valid():
    '''
    Test sum_grouped_response_fields function correctly aggregates values 
    '''
    # Arrange: test DataFrame with numerical columns to be grouped and summed
    df = pd.DataFrame({
        'Submitter Type': ['NHS', 'NHS', 'IS1', 'IS1'],
        'Total Response': [100, 200, 150, 250],
        'Total Eligible': [500, 500, 600, 600]
    })

    columns_to_group_by = ['Submitter Type']

    # Act: Call the function to group by 'Submitter Type' and sum the numerical fields
    result_df = sum_grouped_response_fields(df, columns_to_group_by)

    # Assert: Values are grouped/summed correctly by 'Submitter Type' matching xpected_df
    expected_df = pd.DataFrame({
        'Submitter Type': ['IS1', 'NHS'],
        'Total Response': [400, 300],
        'Total Eligible': [1200, 1000]
    })

    pd.testing.assert_frame_equal(result_df, expected_df)

def test_sum_grouped_response_fields_missing_column():
    '''
    Test test_sum_grouped_response_fields function gives KeyError when presented with non existent column
    '''
    # Arrange: test DataFrame without the 'NonExistentColumn' for grouping
    data = {
        'Total Response': [100, 200, 150, 250],
        'Total Eligible': [500, 500, 600, 600]
    }
    df = pd.DataFrame(data)
    # Present column to group by that does not exist in the DataFrame
    columns_to_group_by = ['Submitter Type'] 

    # Act & Assert: Expect a KeyError to be raised
    with pytest.raises(KeyError, match=r"The following columns are missing in the DataFrame: \['Submitter Type'\]"):
        sum_grouped_response_fields(df, columns_to_group_by)

def test_create_data_totals_success():
    '''
    Test create_data_totals function correctly sums the specified columns and adds a "Total" row
    '''
    # Arrange: test DataFrame with numerical data to sum
    df = pd.DataFrame({
        'Submitter Type': ['NHS', 'IS1', 'NHS'],
        'Period': ['Jan-24', 'Jan-24', 'Jan-24'],
        'Very Good': [10, 20, 30],
        'Good': [40, 50, 60]
    })
    # Set parameters
    current_fft_period = 'Jan-24'
    total_column_name = 'Submitter Type'
    columns_to_sum = ['Very Good', 'Good']

    # Act:
    result_df = create_data_totals(df, current_fft_period, total_column_name, columns_to_sum)

    # Assert: result is a DataFrame of one row with correctly summed totals with result matching expected
    expected_df = pd.DataFrame({
        'Submitter Type': ['Total'],
        'Period': ['Jan-24'],
        'Very Good': [60],
        'Good': [150]
    })
    pd.testing.assert_frame_equal(result_df, expected_df)

def test_create_data_totals_missing_column():
    '''
    Test that the function raises a KeyError if a column is missing from the DataFrame
    '''
    # Arrange: test DataFrame missing the 'Good' column used in summing
    df = pd.DataFrame({
        'Submitter Type': ['NHS', 'IS1', 'NHS'],
        'Period': ['Jan-24', 'Jan-24', 'Jan-24'],
        'Very Good': [10, 20, 30]
    })
    # Set parameters
    current_fft_period = 'Jan-24'
    total_column_name = 'Submitter Type'
    columns_to_sum = ['Very Good', 'Good']
    
    # Act & Assert: KeyError will be raised when the function is called with a missing column
    with pytest.raises(KeyError, match=r"The following columns are missing in the DataFrame: \['Good'\]"):
        create_data_totals(df, current_fft_period, total_column_name, columns_to_sum)

def test_append_dataframes():
    '''
    Test that the function successfully appends one DataFrame to another.
    '''
    # Arrange: test DataFrames that need to be appended
    df1 = pd.DataFrame({
        'Submitter Type': ['NHS', 'IS1'],
        'Period': ['Jan-24', 'Jan-24'],
        'Very Good': [20, 40],
        'Good': [60, 90]
    })

    df2 = pd.DataFrame({
        'Submitter Type': ['Total'],
        'Period': ['Jan-24'],
        'Very Good': [60],
        'Good': [150]
    })
    
    # Act:
    result_df = append_dataframes(df1, df2)
    
    #Assert: results_df is as expected_df
    expected_df = pd.DataFrame({
        'Submitter Type': ['NHS', 'IS1', 'Total'],
        'Period': ['Jan-24', 'Jan-24', 'Jan-24'],
        'Very Good': [20, 40, 60], 
        'Good': [60, 90, 150]
    })
    pd.testing.assert_frame_equal(result_df, expected_df)
    
    
def test_create_percentage_field():
    '''
    Test create_percentage_field function correctly creates and populates percentage field
    '''
    # Arrange: test DataFrame requiring addition of percentage field
    df = pd.DataFrame({
        'Submitter Type': ['NHS', 'IS1', 'Total'],
        'Period': ['Jan-24', 'Jan-24', 'Jan-24'],
        'Very Good': [20, 40, 60], 
        'Good': [60, 90, 150], 
        'Total Responses': [100, 200, 500]
    })
    # Act:
    result_df = create_percentage_field(df, 'Percentage Positive', 'Very Good', 'Good', 'Total Responses')
        
    # Assert: results_df is as expected_df
    expected_df = pd.DataFrame({
        'Submitter Type': ['NHS', 'IS1', 'Total'],
        'Period': ['Jan-24', 'Jan-24', 'Jan-24'],
        'Very Good': [20, 40, 60], 
        'Good': [60, 90, 150], 
        'Total Responses': [100, 200, 500],
        'Percentage Positive': [0.8, 0.65, 0.42]
    })
    pd.testing.assert_frame_equal(result_df, expected_df)
    
def test_create_percentage_field_missing_column():
    '''
    Test create_percentage_field function correctly generates KeyError when missing required column
    '''
    # Arrange: test DataFrame requiring addition of percentage field with missing 'Good' column
    df = pd.DataFrame({
        'Submitter Type': ['NHS', 'IS1', 'Total'],
        'Period': ['Jan-24', 'Jan-24', 'Jan-24'],
        'Very Good': [20, 40, 60], 
        'Total Responses': [100, 200, 500]
    })
    
    # Act & Assert: KeyError will be raised when the function is called with a missing column
    with pytest.raises(KeyError, match=r"The following columns are missing in the DataFrame: \['Good'\]"):
        create_percentage_field(df, 'Percentage Positive', 'Very Good', 'Good', 'Total Responses')

def test_remove_rows_by_cell_content():
    '''
    Test remove_rows_by_cell_content function correctly removes rows where column contains specified value
    '''
    # Arrange: test DataFrame 
    df = pd.DataFrame({
        'Submitter Type': ['NHS', 'IS1', 'Total'],
        'Period': ['Jan-24', 'Jan-24', 'Jan-24'],
        'Very Good': [20, 40, 60], 
        'Good': [60, 90, 150]
    })
    # Act: reset index to avoid index errors when compared with expected_df
    result_df = remove_rows_by_cell_content(df, 'Submitter Type', 'IS1').reset_index(drop=True)
    
    # Assert: results_df is as expected_df
    expected_df = pd.DataFrame({
        'Submitter Type': ['NHS', 'Total'],
        'Period': ['Jan-24', 'Jan-24'],
        'Very Good': [20, 60], 
        'Good': [60, 150]
    })   
    pd.testing.assert_frame_equal(result_df, expected_df)

def test_remove_rows_by_cell_content_missing_column():
    '''
    Test remove_rows_by_cell_content function correctly generates KeyError when missing required column
    '''
    # Arrange: test DataFrame requiring row ewmoval with missing 'Sumbitter Type' column
    df = pd.DataFrame({
        'Period': ['Jan-24', 'Jan-24', 'Jan-24'],
        'Very Good': [20, 40, 60], 
        'Total Responses': [100, 200, 500]
    })
    
    # Act & Assert: KeyError will be raised when the function is called with a missing column
    with pytest.raises(KeyError, match=r"Column 'Submitter Type' not found in the DataFrame"):
        remove_rows_by_cell_content(df, 'Submitter Type', 'IS1')

def test_reorder_columns():
    '''
    Test reorder_columns function correctly order columns by specified order
    '''
    # Arrange: test DataFrame 
    df = pd.DataFrame({
        'Submitter Type': ['NHS', 'IS1', 'Total'],
        'Very Good': [20, 40, 60], 
        'Good': [60, 90, 150], 
        'Total Responses': [100, 200, 500],
        'Percentage Positive': [0.8, 0.65, 0.42]
    }) 
    # Desired order
    totals_output_column_order = ['Total Responses', 'Percentage Positive', 'Very Good', 'Good', 'Submitter Type']

    # Act:
    result_df = reorder_columns(df, totals_output_column_order)
    
    # Assert: results_df is as expected_df
    expected_df = pd.DataFrame({
        'Total Responses': [100, 200, 500],
        'Percentage Positive': [0.8, 0.65, 0.42],
        'Very Good': [20, 40, 60], 
        'Good': [60, 90, 150],  
        'Submitter Type': ['NHS', 'IS1', 'Total']
    })
    pd.testing.assert_frame_equal(result_df, expected_df)
    
def test_reorder_columns_column_mismatch():
    '''
    Test reorder_columns function correctly generates ValueError when column reorder doesn't match DataFrame columns
    '''
    # Arrange: test DataFrame requiring row rewmoval with missing 'Sumbitter Type' column
    df = pd.DataFrame({
        'Submitter Type': ['NHS', 'IS1', 'Total'],
        'Very Good': [20, 40, 60], 
        'Good': [60, 90, 150], 
        'Total Responses': [100, 200, 500],
        'Percentage Positive': [0.8, 0.65, 0.42]
    }) 
    
    # Act & Assert: ValueError will be raised when the function is called when column reorder doesn't match DataFrame columns 
    with pytest.raises(ValueError, match=r"The columns list must contain exactly the same columns as the DataFrame."):
        reorder_columns(df, ['Total Responses', 'Percentage Positive', 'Very Good', 'Submitter Type'])

def test_convert_fields_to_object_type():
    '''
    Test successful conversion of specified numeric columns to object type.
    '''
    # Arrange: test DataFrame and define columns to convert
    df = pd.DataFrame({
        'Submitter Type': ['NHS', 'IS1', 'Total'],
        'Very Good': [20, 40, 60], 
        'Percentage Positive': [0.8, 0.65, 0.42]
    }) 
    # Numeric columns to convert
    fields_to_convert = ['Very Good', 'Percentage Positive']

    # Act:
    df_converted = convert_fields_to_object_type(df, fields_to_convert)

    # Assert: Verify that the specified columns are now of object dtype
    assert df_converted['Submitter Type'].dtype == 'object' # Should remain object dtype (already a string)"
    assert df_converted['Very Good'].dtype == 'object' # Should be converted to object dtype"
    assert df_converted['Percentage Positive'].dtype == 'object' # Should be converted to object dtype"

def test_replace_missing_values():
    '''
    Test clean_dataframe function successfully replaces null values with specified value
    '''
    # Arrange: test DataFrame 
    df = pd.DataFrame({
        'Submitter Type': ['NHS', 'IS1', 'Total'],
        'Very Good': [20, np.nan, 60], 
        'Percentage Positive': [0.8, np.nan, 0.42]
    }).astype(object)

    # Act:
    result_df = replace_missing_values(df, 'NA')
    
    # Assert: results_df is as expected_df
    expected_df = pd.DataFrame({
        'Submitter Type': ['NHS', 'IS1', 'Total'],
        'Very Good': [20, 'NA', 60], 
        'Percentage Positive': [0.8, 'NA', 0.42]        
    }).astype(object)

    pd.testing.assert_frame_equal(result_df, expected_df)

def test_replace_missing_values_no_nan():
    '''
    Test clean_dataframe function functions correctly when no null values are encountered
    '''
    # Arrange: test DataFrame 
    df = pd.DataFrame({
        'Submitter Type': ['NHS', 'IS1', 'Total'],
        'Very Good': [20, 40, 60], 
        'Percentage Positive': [0.8, 0.65, 0.42]
    }) 
    
    # Expected output is the same as the input DataFrame
    expected_df = df.copy()

    # Act: Call the function
    result_df = replace_missing_values(df, 0)

    # Assert: results match the original DataFrame
    pd.testing.assert_frame_equal(result_df, expected_df)

def test_count_nhs_is1_totals_counted_successfully():
    '''
    Test counting of 'IS1' and 'NHS' values in the specified column.
    '''
    # Arrange: test DataFrame 
    df = pd.DataFrame({
        'ICB_Code': ['IS1', 'NHS', 'IS1', 'Other', 'NHS', 'IS1']
    })
    
    # Act:
    result = count_nhs_is1_totals(df, 'ICB_Code', 'is1_total', 'nhs_total')
    
    # Assert: Check if the counts are correct
    assert result['is1_total'] == 3
    assert result['nhs_total'] == 2

def test_count_nhs_is1_totals_column_not_found():
    '''
    Test that a KeyError is raised when the specified column does not exist.
    '''
    # Arrange: test DataFrame 
    df = pd.DataFrame({
        'Good': [20, 30, 40]
    })
    
    # Act & Assert: Ensure a KeyError is raised when non-existent column raised
    with pytest.raises(KeyError, match=r"Column 'ICB_Code' not found in the DataFrame"):
        count_nhs_is1_totals(df, 'ICB_Code', 'is1_total', 'nhs_total')

def test_count_nhs_is1_totals_case_sensitivity():
    '''
    Test that the function is case-sensitive and only counts exact matches.
    '''
    # Arrange: test DataFrame with mixed-case values
    df = pd.DataFrame({
        'ICB_Code': ['IS1', 'is1', 'NHS', 'nhs', 'IS1']
    })
    
    # Act:
    result = count_nhs_is1_totals(df, 'ICB_Code', 'is1_total', 'nhs_total')
    
    # Assert: Ensure only exact matches ('IS1' and 'NHS') are counted
    assert result['is1_total'] == 2
    assert result['nhs_total'] == 1

def test_add_dataframe_column():
    '''
    Test adding new column to DataFrame.
    '''
    # Arrange: test DataFrame 
    df = pd.DataFrame({
        'Submitter Type': ['NHS', 'IS1', 'Total'],
        'Very Good': [20, 40, 60]
    })
    column_name = 'Percentage_Positive'
    column_value = 10
    
    # Act:
    result_df = add_dataframe_column(df, column_name, column_value)
    
    # Assert: results_df is as expected_df
    expected_df = pd.DataFrame({
        'Submitter Type': ['NHS', 'IS1', 'Total'],
        'Very Good': [20, 40, 60], 
        'Percentage_Positive': [10, 10, 10]  
    })
    pd.testing.assert_frame_equal(result_df, expected_df)

def test_add_dataframe_column_invalid_column_name():
    '''
    Test that a TypeError is raised when column_name is not a string or list.
    '''
    # Arrange: test DataFrame 
    df = pd.DataFrame({
        'Submitter Type': ['NHS', 'IS1', 'Total'],
        'Very Good': [20, 40, 60]
    })
    invalid_column_name = 123
    
    # Act & Assert: Ensure a TypeError is raised for invalid column name type
    with pytest.raises(TypeError, match="New column name should be a list or a string."):
        add_dataframe_column(df, invalid_column_name, 10)

def test_add_submission_counts_success():
    '''
    Test that submission counts are correctly added for 'IS1' and 'NHS'.
    '''
    # Arrange: test DataFrame 
    df = pd.DataFrame({
        'Code': ['IS1', 'NHS', 'Other'],
        'SubmissionCount': [0, 0, 0]
    }).astype(object)
    code_column_name = 'Code'
    target_column = 'SubmissionCount'
    is1_count = 10
    nhs_count = 20
    
    # Act: Call the function to add counts
    result_df = add_submission_counts_to_df(df, code_column_name, is1_count, nhs_count, target_column)
    
    # Assert: results_df is as expected_df
    expected_df = pd.DataFrame({
        'Code': ['IS1', 'NHS', 'Other'],
        'SubmissionCount': [10, 20, 0]       
    }).astype(object)

    pd.testing.assert_frame_equal(result_df, expected_df)

def test_add_submission_counts_to_df_code_column_missing():
    '''
    Test that KeyError is raised when code_column_name is missing.
    '''
    # Arrange: test DataFrame without the required 'Code' column
    df = pd.DataFrame({
        'WrongCode': ['IS1', 'NHS', 'Other'],
        'Very Poor': [0, 0, 0]
    })
    # Non-existing column
    code_column_name = 'Code'
    target_column = 'Very Poor'
    
    # Act & Assert: Ensure KeyError is raised for missing code_column_name
    with pytest.raises(KeyError, match="One or both of columns 'Code' or 'Very Poor' are missing from the DataFrame."):
        add_submission_counts_to_df(df, code_column_name, 10, 20, target_column)

def test_add_submission_counts_to_df_target_column_missing():
    '''
    Test that KeyError is raised when target_column is missing.
    '''
    # Arrange: test DataFrame without the target column
    df = pd.DataFrame({
        'Code': ['IS1', 'NHS', 'Other'],
        'Very Poor': [0, 0, 0]
    })
    code_column_name = 'Code'
    # Non-existing column
    target_column = 'SubmissionCount' 
    
    # Act & Assert: Ensure KeyError is raised for missing target_column
    with pytest.raises(KeyError, match="One or both of columns 'Code' or 'SubmissionCount' are missing from the DataFrame."):
        add_submission_counts_to_df(df, code_column_name, 10, 20, target_column)

def test_update_monthly_rolling_totals_successful():
    '''
    Test that new FFT period data is added to df2 correctly.
    '''
    # Arrange: test DataFrames
    df1 = pd.DataFrame({
        'Submitter Type': ['Total', 'NHS', 'IS1'],
        'Number of organisations submitting': [10, 7, 3],
        'Total Responses': [1000, 700, 300],
        'Percentage Positive': [80, 75, 85],
        'Percentage Negative': [10, 15, 5]
    })
    
    df2 = pd.DataFrame({
        'FFT Period': ['2024-04', '2024-05'],
        'Total submitters': [12, 15],
        'Number of NHS submitters': [8, 9],
        'Number of Independent submitters': [4, 6],
        'Total responses to date': [1200, 1500],
        'Total NHS responses to date': [800, 900],
        'Total independent responses to date': [400, 600],
        'Monthly total responses': [1000, 1500],
        'Monthly NHS responses': [700, 900],
        'Monthly independent responses': [300, 600],
        'Monthly total percentage positive': [80, 85],
        'Monthly NHS percentage positive': [75, 80],
        'Monthly independent percentage positive': [85, 90],
        'Monthly total percentage negative': [10, 12],
        'Monthly NHS percentage negative': [15, 10],
        'Monthly independent percentage negative': [5, 4]
    })

    current_fft_period = '2024-06'

    # Act:
    result_df = update_monthly_rolling_totals(df1, df2, current_fft_period)

    # Assert: results_df is as expected_df
    expected_df = pd.DataFrame({
        'FFT Period': ['2024-04', '2024-05', '2024-06'],
        'Total submitters': [12, 15, 10],
        'Number of NHS submitters': [8, 9, 7],
        'Number of Independent submitters': [4, 6, 3],
        'Total responses to date': [1200, 1500, 0],
        'Total NHS responses to date': [800, 900, 0],
        'Total independent responses to date': [400, 600, 0],
        'Monthly total responses': [1000, 1500, 1000],
        'Monthly NHS responses': [700, 900, 700],
        'Monthly independent responses': [300, 600, 300],
        'Monthly total percentage positive': [80, 85, 80],
        'Monthly NHS percentage positive': [75, 80, 75],
        'Monthly independent percentage positive': [85, 90, 85],
        'Monthly total percentage negative': [10, 12, 10],
        'Monthly NHS percentage negative': [15, 10, 15],
        'Monthly independent percentage negative': [5, 4, 5]      
    })

    pd.testing.assert_frame_equal(result_df, expected_df)
        
def test_update_monthly_rolling_totals_overwrites_existing():
    '''
    Test that existing FFT period data in df2 is correctly overwritten.
    '''
    # Arrange: test DataFrames 
    df1 = pd.DataFrame({
        'Submitter Type': ['Total', 'NHS', 'IS1'],
        'Number of organisations submitting': [10, 7, 3],
        'Total Responses': [1000, 700, 300],
        'Percentage Positive': [80, 75, 85],
        'Percentage Negative': [10, 15, 5]
    })
    
    df2 = pd.DataFrame({
        'FFT Period': ['2024-04', '2024-05', '2024-06'],
        'Total submitters': [12, 15, 10],
        'Number of NHS submitters': [8, 9, 7],
        'Number of Independent submitters': [4, 6, 3],
        'Total responses to date': [1200, 1500, 0],
        'Total NHS responses to date': [800, 900, 0],
        'Total independent responses to date': [400, 600, 0],
        'Monthly total responses': [1000, 1500, 1000],
        'Monthly NHS responses': [700, 900, 700],
        'Monthly independent responses': [300, 600, 300],
        'Monthly total percentage positive': [80, 85, 80],
        'Monthly NHS percentage positive': [75, 80, 75],
        'Monthly independent percentage positive': [85, 90, 85],
        'Monthly total percentage negative': [10, 12, 10],
        'Monthly NHS percentage negative': [15, 10, 15],
        'Monthly independent percentage negative': [5, 4, 5]      
    })

    current_fft_period = '2024-06'

    # Act: Call the function overwriting existing data
    result_df = update_monthly_rolling_totals(df1, df2, current_fft_period)

    # Assert: results_df is as expected_df
    expected_df = pd.DataFrame({
        'FFT Period': ['2024-04', '2024-05', '2024-06'],
        'Total submitters': [12, 15, 10],
        'Number of NHS submitters': [8, 9, 7],
        'Number of Independent submitters': [4, 6, 3],
        'Total responses to date': [1200, 1500, 0],
        'Total NHS responses to date': [800, 900, 0],
        'Total independent responses to date': [400, 600, 0],
        'Monthly total responses': [1000, 1500, 1000],
        'Monthly NHS responses': [700, 900, 700],
        'Monthly independent responses': [300, 600, 300],
        'Monthly total percentage positive': [80, 85, 80],
        'Monthly NHS percentage positive': [75, 80, 75],
        'Monthly independent percentage positive': [85, 90, 85],
        'Monthly total percentage negative': [10, 12, 10],
        'Monthly NHS percentage negative': [15, 10, 15],
        'Monthly independent percentage negative': [5, 4, 5]      
    })

    pd.testing.assert_frame_equal(result_df, expected_df)        

def test_update_monthly_rolling_totals_missing_columns_errors():
    '''
    Test that KeyError is raised when df1 is missing required columns.
    '''
    # Arrange: test DataFrames with df1 missig required columns
    df1 = pd.DataFrame({
        'Submitter Type': ['Total', 'NHS', 'IS1'],
        'Some other column': [10, 7, 3]
    })
    
    df2 = pd.DataFrame({'FFT Period': ['2023-01', '2023-02']})
    current_fft_period = '2023-03'

    # Act & Assert: Ensure KeyError is raised for missing columns in df1
    with pytest.raises(KeyError, match="Missing columns in df1"):
        update_monthly_rolling_totals(df1, df2, current_fft_period)

def test_missing_fft_period_column_in_df2():
    '''
    Test that KeyError is raised when df2 is missing the 'FFT Period' column.
    '''
    # Arrange: test DataFrames with df2 without 'FFT Period' column
    df1 = pd.DataFrame({
        'Submitter Type': ['Total', 'NHS', 'IS1'],
        'Number of organisations submitting': [10, 7, 3],
        'Total Responses': [1000, 700, 300],
        'Percentage Positive': [80, 75, 85],
        'Percentage Negative': [10, 15, 5]
    })
    
    df2 = pd.DataFrame({
        'Total Eligible': [2000, 1700]
    })
    current_fft_period = '2024-06'

    # Act & Assert: Ensure KeyError is raised for missing 'FFT Period' column in df2
    with pytest.raises(KeyError, match="'FFT Period' column is missing in df2."):
        update_monthly_rolling_totals(df1, df2, current_fft_period)

def test_update_cumulative_value_success():
    '''
    Test the successful update of cumulative values.
    '''
    # Arrange: test DataFrame with sufficient rows and columns
    df = pd.DataFrame({
        'Monthly total responses': [10, 20, 30],
        'Total responses to date': [5, 15, 25]
    })
    
    # Act: Call the function to update the DataFrame
    updated_df = update_cumulative_value(df, 'Monthly total responses', 'Total responses to date')
    
    # Assert: Check that the Total responses to date's last row is correctly updated
    assert updated_df.at[2, 'Total responses to date'] == 30 + 15

def test_update_cumulative_value_key_error():
    '''
    Test that a KeyError is raised when columns are missing.
    '''
    # Arrange: test DataFrame missing the second column
    df = pd.DataFrame({
        'Monthly total responses': [10, 20, 30],
        'third_column': [5, 15, 25]
    })
    
    # Act & Assert: Check for KeyError when calling the function
    with pytest.raises(KeyError, match="One or both columns"):
        update_cumulative_value(df, 'Monthly total responses', 'Total responses to date')

def test_update_cumulative_value_value_error():
    '''
    Test that a ValueError is raised when there are fewer than two rows.
    '''
    # Arrange: test DataFrame with only one row
    df = pd.DataFrame({
        'Monthly total responses': [10],
        'Total responses to date': [5]
    })
    
    # Act & Assert: Check for ValueError when calling the function
    with pytest.raises(ValueError, match="The DataFrame must have at least two rows"):
        update_cumulative_value(df, 'Monthly total responses', 'Total responses to date')

def test_update_existing_excel_sheet_success():
    '''
    Test that the function successfully updates the Excel sheet when the file exists,
    using a mock file to avoid actual file system interactions.
    '''
    # Arrange: mock file to act as an Excel file
    file_path = r'mock_files\Monthly Rolling Totals.xlsx'
    sheet_name = 'OP'
    updated_df = pd.DataFrame({'A': [1, 2], 'B': [3, 4]})  # Sample DataFrame
 
    # Act:
    update_existing_excel_sheet(file_path, sheet_name, updated_df)
 
    # Assert
    # No exceptions should be raised, meaning the function works as expected

def test_update_existing_excel_sheet_file_not_found(mocker):
    '''
    Test that FileNotFoundError is raised when the Excel file does not exist.
    '''
    # Arrange: mocked file path
    file_path = 'non_existent_file.xlsx'
    sheet_name = 'Sheet1'
    updated_df = pd.DataFrame({'A': [1, 2], 'B': [3, 4]})  # Sample DataFrame

    # Mock os.path.exists to return False (simulate file not existing)
    mocker.patch('os.path.exists', return_value=False)

    # Act & Assert: FileNotFoundError is raised when the file does not exist
    with pytest.raises(FileNotFoundError, match=f"The file '{file_path}' does not exist."):
        update_existing_excel_sheet(file_path, sheet_name, updated_df)

def test_copy_value_between_dataframes_success():
    '''
    Test successful copying of value from source DataFrame to target DataFrame.
    '''
    # Arrange: Test source and target DataFrames
    df_source = pd.DataFrame({
        'Total responses to date': [10, 20, 30],
        'Monthly total responses': [5, 7, 15]
    })

    df_target = pd.DataFrame({
        'Total Responses to Date': [0, 0, 0],
        'Previous Months Responses': ['3', '5', '7']
    })

    source_column = 'Total responses to date'
    source_row = 1
    target_column = 'Total Responses to Date'
    target_row = 2

    # Act:
    updated_df_target = copy_value_between_dataframes(df_source, df_target, source_column, source_row, target_column, target_row)

    # Assert: Check that the value was copied correctly
    assert updated_df_target.at[target_row, target_column] == df_source.at[source_row, source_column]

def test_copy_value_between_dataframes_missing_source_column():
    '''
    Test that the function raises a ValueError if the source column does not exist.
    '''
    # Arrange: Create source and target DataFrames
    df_source = pd.DataFrame({
        'Total responses to date': [10, 20, 30],
        'Monthly total responses': [5, 7, 15]
    })

    df_target = pd.DataFrame({
        'Total Responses to Date': [0, 0, 0],
        'Previous Months Responses': ['3', '5', '7']
    })

    source_column = 'Previous Months Percentage Positive' # non existent source column
    source_row = 1
    target_column = 'Total Responses to Date'
    target_row = 2

    # Act & Assert: Expect a KeyError due to the missing source column
    with pytest.raises(KeyError, match=f"Source column '{source_column}' does not exist in df_source."):
        copy_value_between_dataframes(df_source, df_target, source_column, source_row, target_column, target_row)

def test_copy_value_between_dataframes_missing_target_column():
    '''
    Test that the function raises a KeyError if the target column does not exist.
    '''
    # Arrange: Test source and target DataFrames
    df_source = pd.DataFrame({
        'Total responses to date': [10, 20, 30],
        'Monthly total responses': [5, 7, 15]
    })

    df_target = pd.DataFrame({
        'Total Responses to Date': [0, 0, 0],
        'Previous Months Responses': ['3', '5', '7']
    })

    source_column = 'Total responses to date'
    source_row = 1
    target_column = 'Previous Months Percentage Positive' # non existent target column
    target_row = 2

    # Act & Assert: Expect a KeyError due to the missing target column
    with pytest.raises(KeyError, match=f"Target column '{target_column}' does not exist in df_target."):
        copy_value_between_dataframes(df_source, df_target, source_column, source_row, target_column, target_row)

def test_copy_value_between_dataframes_missing_source_row():
    '''
    Test that the function raises an IndexError if the source row does not exist.
    '''
    # Arrange: Test source and target DataFrames
    df_source = pd.DataFrame({
        'Total responses to date': [10, 20, 30],
        'Monthly total responses': [5, 7, 15]
    })

    df_target = pd.DataFrame({
        'Total Responses to Date': [0, 0, 0],
        'Previous Months Responses': ['3', '5', '7']
    })

    source_column = 'Total responses to date'
    source_row = 5 # non existent row
    target_column = 'Total Responses to Date'
    target_row = 2

    # Act & Assert: Expect an IndexError due to the missing source row
    with pytest.raises(IndexError, match=f"Source row '{source_row}' does not exist in df_source."):
        copy_value_between_dataframes(df_source, df_target, source_column, source_row, target_column, target_row)

# Test for IndexError when target row does not exist
def test_copy_value_between_dataframes_missing_target_row():
    '''
    Test that the function raises an IndexError if the target row does not exist.
    '''
    # Arrange: Test source and target DataFrames
    df_source = pd.DataFrame({
        'Column1': [10, 20, 30],
        'Column2': ['A', 'B', 'C']
    })

    df_target = pd.DataFrame({
        'Column3': [0, 0, 0],
        'Column4': ['X', 'Y', 'Z']
    })

    source_column = 'Column1'
    source_row = 1
    target_column = 'Column3'
    target_row = 5  # This row does not exist

    # Act & Assert: Expect an IndexError due to the missing target row
    with pytest.raises(IndexError, match=f"Target row '{target_row}' does not exist in df_target."):
        copy_value_between_dataframes(df_source, df_target, source_column, source_row, target_column, target_row)

def test_new_column_name_with_period_prefix_success():
    '''
    Test that the function correctly concatenates period_prefix and new_column_suffix to create the new column name.
    '''
    # Arrange: Create valid inputs for the function
    period_prefix = "2024_08"
    new_column_suffix = "Responses"

    # Act:
    result = new_column_name_with_period_prefix(period_prefix, new_column_suffix)

    # Assert: Verify that the result is the concatenated string as expected
    assert result == "2024_08_Responses"

def test_new_column_name_with_period_prefix_non_string_period_prefix():
    '''
    Test that the function raises a TypeError when period_prefix is not a string.
    '''
    # Arrange: Set up an invalid period_prefix (integer instead of string)
    period_prefix = 202408 # Non-string type
    new_column_suffix = "Responses"

    # Act & Assert: Expect the function to raise a TypeError due to non-string period_prefix
    with pytest.raises(TypeError, match="period_prefix must be a string"):
        new_column_name_with_period_prefix(period_prefix, new_column_suffix)

def test_new_column_name_with_period_prefix_non_string_column_suffix():
    '''
    Test that the function raises a TypeError when new_column_suffix is not a string.
    '''
    # Arrange: Create invalid new_column_suffix (list instead of string)
    period_prefix = "2023_08"
    new_column_suffix = ["Responses"] # Non-string type

    # Act & Assert: Expect the function to raise a TypeError due to non-string new_column_suffix
    with pytest.raises(TypeError, match="new_column_suffix must be a string"):
        new_column_name_with_period_prefix(period_prefix, new_column_suffix)

def test_sort_dataframe_single_field_ascending():
    '''
    Test sorting a DataFrame by a single field in ascending order.
    '''
    # Arrange: Test DataFrame and sorting parameters
    df = pd.DataFrame({
        'A': [3, 1, 2],
        'B': ['x', 'y', 'z']
    })
    df_fields = 'A' # Sort by
    directions = True # In ascending order

    # Act:
    result = sort_dataframe(df, df_fields, directions)

    # Assert: Check if the DataFrame is sorted correctly by column 'A'
    expected = pd.DataFrame({
        'A': [1, 2, 3],
        'B': ['y', 'z', 'x']
    }, index=[0, 1, 2])
    pd.testing.assert_frame_equal(result, expected)

def test_sort_dataframe_multiple_fields_mixed_directions():
    '''
    Test sorting a DataFrame by multiple fields with mixed directions.
    '''
    # Arrange: Test DataFrame and sorting parameters
    df = pd.DataFrame({
        'A': [3, 1, 2, 1],
        'B': [2, 1, 3, 2]
    })
    df_fields = ['A', 'B'] # Sort by
    directions = [True, False] # Sort A ascending and B descending

    # Act:
    result = sort_dataframe(df, df_fields, directions)

    # Assert: Check if the DataFrame is sorted correctly by 'A' and then 'B'
    expected = pd.DataFrame({
        'A': [1, 1, 2, 3],
        'B': [2, 1, 3, 2]
    }, index=[0, 1, 2, 3])
    pd.testing.assert_frame_equal(result, expected)

def test_sort_dataframe_invalid_field():
    '''
    Test that the function raises a KeyError when the field to sort by does not exist in the DataFrame.
    '''
    # Arrange: Test DataFrame with invalid sort field
    df = pd.DataFrame({
        'A': [3, 1, 2],
        'B': ['x', 'y', 'z']
    })
    df_fields = 'C' # non existent field
    directions = True

    # Act & Assert: Expect a KeyError
    with pytest.raises(KeyError, match="One or more fields to sort by do not exist in the DataFrame."):
        sort_dataframe(df, df_fields, directions)

def test_create_first_level_suppression_success():
    '''
    Test that the suppression column is created correctly when the responses_field exists.
    '''
    # Arrange: Test sample DataFrame with a valid 'Responses' field
    df = pd.DataFrame({
        'Responses': [0, 3, 5, 7, 2]
    })
    responses_field = 'Responses'
    first_level_suppression = 'Suppression'

    # Act: 
    result = create_first_level_suppression(df, first_level_suppression, responses_field)

    # Assert: Check that the suppression column was created correctly
    expected = pd.DataFrame({
        'Responses': [0, 3, 5, 7, 2],
        'Suppression': [0, 1, 0, 0, 1]
    })
    pd.testing.assert_frame_equal(result, expected)

# Test: Raise KeyError if responses_field does not exist
def test_create_first_level_suppression_missing_field():
    '''
    Test that the function raises a KeyError when the responses_field does not exist in the DataFrame.
    '''
    # Arrange: Test DataFrame without the 'Responses' field
    df = pd.DataFrame({
        'Field_Supression': [1, 2, 3]
    })
    responses_field = 'Responses' # non existent field in df
    first_level_suppression = 'Suppression'

    # Act & Assert: Expect a KeyError
    with pytest.raises(KeyError, match="'Responses' does not exist in the DataFrame."):
        create_first_level_suppression(df, first_level_suppression, responses_field)

def test_create_icb_second_level_suppression_success():
    '''
    Test that the second level suppression column is created correctly when the first level suppression field is present.
    '''
    # Arrange: Test DataFrame with the 'first_level_suppression' column
    df = pd.DataFrame({
        'first_level_suppression': [0, 1, 0, 1, 0],
        'Total_Responses': [10, 20, 30, 40, 50]
    })
    first_level_suppression = 'first_level_suppression'
    second_level_suppression = 'second_level_suppression'

    # Act:
    result = create_icb_second_level_suppression(df, first_level_suppression, second_level_suppression)

    # Assert: Check that the second level suppression column was created correctly
    expected = pd.DataFrame({
        'first_level_suppression': [0, 1, 0, 1, 0],
        'Total_Responses': [10, 20, 30, 40, 50],
        'second_level_suppression': [0, 0, 1, 0, 1]
    })
    pd.testing.assert_frame_equal(result, expected)

def test_create_icb_second_level_suppression_no_first_level():
    '''
    Test that no second level suppression is applied when there are no first level suppression flags set to 1.
    '''
    # Arrange: Test DataFrame with no first level suppression
    df = pd.DataFrame({
        'first_level_suppression': [0, 0, 0, 0, 0],
        'Total_Responses': [10, 20, 30, 40, 50]
    })
    first_level_suppression = 'first_level_suppression'
    second_level_suppression = 'second_level_suppression'

    # Act:
    result = create_icb_second_level_suppression(df, first_level_suppression, second_level_suppression)

    # Assert: Check that no second level suppression occurs
    expected = pd.DataFrame({
        'first_level_suppression': [0, 0, 0, 0, 0],
        'Total_Responses': [10, 20, 30, 40, 50],
        'second_level_suppression': [0, 0, 0, 0, 0]
    })
    pd.testing.assert_frame_equal(result, expected)

def test_confirm_row_level_suppression_success():
    '''
    Test suppression field created correctly when suppression columns are present.
    '''
    # Arrange: Test DataFrame with multiple suppression columns
    df = pd.DataFrame({
        'field1': [0, 1, 0, 1, 0],
        'field2': [1, 0, 0, 1, 0],
        'OtherField': [10, 20, 30, 40, 50] # Irrelevant data column
    }).astype(int)
    suppression_field = 'suppress'
    suppression_columns = ['field1', 'field2']

    # Act:
    result = confirm_row_level_suppression(df, suppression_field, *suppression_columns)

    # Assert: Check that the suppression field is created correctly
    expected = pd.DataFrame({
        'field1': [0, 1, 0, 1, 0],
        'field2': [1, 0, 0, 1, 0],
        'OtherField': [10, 20, 30, 40, 50],
        'suppress': [1, 1, 0, 1, 0] # Suppression applied (1) if either field1 or field2 is 1
    }).astype(int)
    pd.testing.assert_frame_equal(result, expected)

def test_confirm_row_level_suppression_missing_column():
    '''
    Test the function raises a KeyError when one of the suppression columns does not exist.
    '''
    # Arrange: Test DataFrame missing one of the suppression columns
    df = pd.DataFrame({
        'field1': [0, 1, 0, 1, 0],
        # Missing 'field2'
        'OtherField': [10, 20, 30, 40, 50] 
    })
    suppression_field = 'suppress'
    suppression_columns = ['field1', 'field2']

    # Act & Assert: Expect a KeyError with a specific error message
    with pytest.raises(KeyError, match="The following suppression columns are missing from the DataFrame: field2"):
        confirm_row_level_suppression(df, suppression_field, *suppression_columns)

def test_confirm_row_level_suppression_no_conditions_met():
    '''
    Test no suppression occurs when no suppression columns contain 1.
    '''
    # Arrange: Test DataFrame where no suppression is needed
    df = pd.DataFrame({
        'field1': [0, 0, 0, 0, 0],
        'field2': [0, 0, 0, 0, 0],
        'OtherField': [10, 20, 30, 40, 50]
    }).astype(int)
    suppression_field = 'suppress'
    suppression_columns = ['field1', 'field2']

    # Act:
    result = confirm_row_level_suppression(df, suppression_field, *suppression_columns)

    # Assert: The suppression field should contain only 0s
    expected = pd.DataFrame({
        'field1': [0, 0, 0, 0, 0],
        'field2': [0, 0, 0, 0, 0],
        'OtherField': [10, 20, 30, 40, 50],
        'suppress': [0, 0, 0, 0, 0]
    }).astype(int)
    pd.testing.assert_frame_equal(result, expected)

def test_suppress_data_success():
    '''
    Test data suppression is correctly applied to response and percentage columns
    based on 'overall_suppression_field' and 'first_level_suppression_field'.
    '''
    # Arrange: Test DataFrame with relevant columns and suppression conditions
    df = pd.DataFrame({
        'Very Good': [3, 5, 10, 20],
        'Good': [5, 10, 15, 25],
        'Neither Good nor Poor': [15, 20, 30, 35],
        'Poor': [5, 10, 15, 20],
        'Very Poor': [3, 5, 15, 20],
        'Dont Know': [2, 4, 6, 8],
        'Percentage Positive': [40, 60, 70, 75],
        'Percentage Negative': [60, 40, 30, 25],
        'overall_suppression': [1, 1, 1, 0],
        'first_level_suppression': [1, 1, 0, 0]
    }).astype(object)
    overall_suppression_field = 'overall_suppression'
    first_level_suppression_field = 'first_level_suppression'

    # Act:
    result = suppress_data(df, overall_suppression_field, first_level_suppression_field)

    # Assert: Verify that only the necessary fields were suppressed with '*'
    expected = pd.DataFrame({
        'Very Good': ['*', '*', '*', 20],
        'Good': ['*', '*', '*', 25],
        'Neither Good nor Poor': ['*', '*', '*', 35],
        'Poor': ['*', '*', '*', 20],
        'Very Poor': ['*', '*', '*', 20],
        'Dont Know': ['*', '*', '*', 8],
        'Percentage Positive': ['*', '*', 70, 75],
        'Percentage Negative': ['*', '*', 30, 25],
        'overall_suppression': [1, 1, 1, 0],
        'first_level_suppression': [1, 1, 0, 0]
    }).astype(object)
    
    # Ensure the resulting DataFrame matches the expected output
    pd.testing.assert_frame_equal(result, expected)

def test_suppress_data_no_suppression():
    '''
    Test that no suppression is applied when neither 'overall_suppression' nor 'first_level_suppression' is 1.
    '''
    # Arrange: Test DataFrame where no suppression is required
    df = pd.DataFrame({
        'Very Good': [10, 20],
        'Good': [15, 25],
        'Neither Good nor Poor': [30, 35],
        'Poor': [40, 45],
        'Very Poor': [50, 55],
        'Dont Know': [60, 65],
        'Percentage Positive': [70, 75],
        'Percentage Negative': [80, 85],
        'overall_suppression': [0, 0], # No suppression
        'first_level_suppression': [0, 0]
    })
    overall_suppression_field = 'overall_suppression'
    first_level_suppression_field = 'first_level_suppression'

    # Act:
    result = suppress_data(df, overall_suppression_field, first_level_suppression_field)

    # Assert: The DataFrame should remain unchanged
    pd.testing.assert_frame_equal(result, df)

def test_suppress_data_missing_columns():
    '''
    Test that a KeyError is raised if any of the required columns for suppression are missing.
    '''
    # Arrange: Test DataFrame missing required suppression columns
    df = pd.DataFrame({
        'Very Good': [10, 20],
        'Good': [15, 25],
        # Missing other required columns
        'Dont Know': [60, 65],
        'Percentage Positive': [70, 75],
        'Percentage Negative': [80, 85],        
        'overall_suppression': [1, 0], 
        'first_level_suppression': [1, 0]
    })
    overall_suppression_field = 'overall_suppression'
    first_level_suppression_field = 'first_level_suppression'

    # Act & Assert: Expect a KeyError due to missing columns
    with pytest.raises(KeyError, match="The following columns are missing in the DataFrame: Neither Good nor Poor, Poor, Very Poor"):
        suppress_data(df, overall_suppression_field, first_level_suppression_field)

def test_move_independent_provider_rows_to_bottom_success():
    '''
    Test function successfully moves IS1 row to the bottom of the DataFrame
    '''
    # Arrange: Test DataFrame with IS1 rows
    df = pd.DataFrame({
        'ICB Code': ['IS1', 'NHS', 'IS1', 'NHS'],
        'Total Response': [100, 200, 150, 250],
        'Total Eligible': [500, 500, 600, 600]
    })
    
    # Act:
    result = move_independent_provider_rows_to_bottom(df)

    # Assert: result matches expected
    expected = pd.DataFrame({
        'ICB Code': ['NHS', 'NHS', 'IS1', 'IS1'],
        'Total Response': [200, 250, 100, 150],
        'Total Eligible': [500, 600, 500, 600]
    })
    
    pd.testing.assert_frame_equal(result, expected) 

def test_adjust_percentage_field_success():
    '''
    Test function correctly converts number format to decimal percentage format expected in outputs
    '''
    # Arrange: Test DataFrame with IS1 rows
    df = pd.DataFrame({
        'ICB Code': ['ABC', 'DEF', 'GHI', 'JKL'],
        'Percentage Positive': [90.476190, 100.000000, 97.368421, 95.777778],
        'Percentage Negative': [9.548795, 0.00000, 2.6923546, 4.233333]
    })    
    
    # Act:
    result = adjust_percentage_field(df, 'Percentage Positive')

    # Assert: result matches expected
    expected = pd.DataFrame({
        'ICB Code': ['ABC', 'DEF', 'GHI', 'JKL'],
        'Percentage Positive': [0.9, 1.0, 0.97, 0.96],
        'Percentage Negative': [9.548795, 0.00000, 2.6923546, 4.233333]
    })
    
    pd.testing.assert_frame_equal(result, expected) 
    
def test_adjust_percentage_field_missing_column():
    '''
    Test that a KeyError is raised if required column missing form DataFrame.
    '''
    # Arrange: Test DataFrame missing required column
    df = pd.DataFrame({
        'ICB Code': ['ABC', 'DEF', 'GHI', 'JKL'],
        'Percentage Positive': [90.476190, 100.000000, 97.368421, 95.777778]
        # Missing 'Percentage Negative' column
    })  

    # Act & Assert: Expect a KeyError due to missing columns
    with pytest.raises(KeyError, match="Column 'Percentage Negative' not found in the DataFrame."):
        adjust_percentage_field(df, 'Percentage Negative')

def test_rank_organisation_results_success():
    '''
    Test the function correctly ranks responses within each organisation group.
    '''
    # Arrange: Test DataFrame with organisation and response data
    df = pd.DataFrame({
        'Site_Code': ['A', 'A', 'A', 'B', 'B'],
        'Responses': [10, 0, 5, 20, 15]
    })
    org_field = 'Site_Code'
    responses_field = 'Responses'
    rank_field = 'Rank'

    # Act: 
    result = rank_organisation_results(df, org_field, responses_field, rank_field)

    # Assert: Verify that rankings are correct, with 0 responses not ranked
    expected = pd.DataFrame({
        'Site_Code': ['A', 'A', 'A', 'B', 'B'],
        'Responses': [10, 0, 5, 20, 15],
        'Rank': [2, 0, 1, 2, 1] # Non-zero responses are ranked, 0 responses are left as 0
    })
    
    pd.testing.assert_frame_equal(result, expected)

def test_rank_organisation_results_missing_columns():
    '''
    Test the function raises a KeyError if either the org_field or responses_field is missing.
    '''
    # Arrange: Test DataFrame without the necessary columns
    df = pd.DataFrame({
        'Site_Code': ['A', 'A', 'B'],
        # 'Responses' field is missing
        'Data': [10, 5, 20]
    })
    org_field = 'Site_Code'
    responses_field = 'Responses' # This column is missing in df
    rank_field = 'Rank'

    # Act & Assert: Expect a KeyError due to the missing column
    with pytest.raises(KeyError, match="'Site_Code' or 'Responses' does not exist in the DataFrame."):
        rank_organisation_results(df, org_field, responses_field, rank_field)

def test_create_second_level_suppression_success():
    '''
    Test second-level suppression applied correctly when the first-level suppression
    is 1 and the next submission has a rank of 2.
    '''
    # Arrange: Test DataFrame where second-level suppression should be applied
    df = pd.DataFrame({
        'Trust Code': ['A', 'A', 'A', 'B', 'B', 'C'],
        'Rank': [1, 2, 3, 1, 2, 0], # Second submission for 'A' and 'B' has rank 2
        'First_Level_Suppression': [1, 0, 0, 1, 0, 0]
    })
    first_level_suppression = 'First_Level_Suppression'
    rank_field = 'Rank'
    second_level_suppression = 'Second_Level_Suppression'

    # Act:
    result = create_second_level_suppression(df, first_level_suppression, rank_field, second_level_suppression)

    # Assert: second-level suppression is applied only to the second submission for 'A' and 'B'
    expected = pd.DataFrame({
        'Trust Code': ['A', 'A', 'A', 'B', 'B', 'C'],
        'Rank': [1, 2, 3, 1, 2, 0],
        'First_Level_Suppression': [1, 0, 0, 1, 0, 0],
        'Second_Level_Suppression': [0, 1, 0, 0, 1, 0]
    })
    
    pd.testing.assert_frame_equal(result, expected)

def test_create_second_level_suppression_missing_columns():
    '''
    Test function raises a KeyError if either the first_level_suppression or rank_field 
    columns are missing from the DataFrame.
    '''
    # Arrange: Test DataFrame without the necessary columns
    df = pd.DataFrame({
        'Trust Code': ['A', 'A', 'B', 'B'],
        # 'Rank' field is missing
        'First_Level_Suppression': [1, 0, 1, 0]
    })
    first_level_suppression = 'First_Level_Suppression'
    rank_field = 'Rank' # This column is missing in df
    second_level_suppression = 'Second_Level_Suppression'

    # Act & Assert: Expect a KeyError due to the missing 'Rank' column
    with pytest.raises(KeyError, match="'First_Level_Suppression' or 'Rank' does not exist in the DataFrame."):
        create_second_level_suppression(df, first_level_suppression, rank_field, second_level_suppression)

def test_add_suppression_required_from_upper_level_column_success():
    '''
    Test suppression is correctly applied from upper-level DataFrame to lower-level DataFrame 
    based on the suppression requirement from the upper level.
    '''
    # Arrange: Test Dataframes for upper-level and lower-level
    upper_level_df = pd.DataFrame({
        'Org_Code': ['ICB1', 'ICB2'],
        'Suppression': [1, 0] # Suppression required for ICB1, not for ICB2
    })

    lower_level_df = pd.DataFrame({
        'Org_Code': ['ICB1', 'ICB1', 'ICB2', 'ICB2'],
        'Rank': [1, 2, 1, 2] # Two ranks for both ICB1 and ICB2
    })

    upper_level_suppression_column = 'Suppression_Required'
    code_lookup_field = 'Org_Code'
    suppression_lookup_field = 'Suppression'

    # Act: 
    result = add_suppression_required_from_upper_level_column(
        upper_level_df, lower_level_df, 
        upper_level_suppression_column, code_lookup_field, 
        suppression_lookup_field
    )

    # Assert: Verify that suppression is correctly applied for Org_Code 'ICB1' where rank is 1 or 2
    expected = pd.DataFrame({
        'Org_Code': ['ICB1', 'ICB1', 'ICB2', 'ICB2'],
        'Rank': [1, 2, 1, 2],
        'Suppression_Required': [1, 1, 0, 0] # Suppression applied only to ICB1
    })
    pd.testing.assert_frame_equal(result, expected)

def test_add_suppression_required_from_upper_level_column_missing_columns_upper_level():
    '''
    Test the function raises a KeyError if either the code_lookup_field or
    suppression_lookup_field are missing in the upper-level DataFrame.
    '''
    # Arrange: Test DataFrames with upper_level_df missing suppression_lookup_field
    upper_level_df = pd.DataFrame({
        'Org_Code': ['ICB1', 'ICB2'],
        # Missing 'Suppression' field
    })

    lower_level_df = pd.DataFrame({
        'Org_Code': ['ICB1', 'ICB2'],
        'Rank': [1, 1]
    })

    upper_level_suppression_column = 'Suppression_Required'
    code_lookup_field = 'Org_Code'
    suppression_lookup_field = 'Suppression'

    # Act & Assert: Expect KeyError due to missing column in upper_level_df
    with pytest.raises(KeyError, match="Missing required columns 'Org_Code' or 'Suppression' in upper_level_df."):
        add_suppression_required_from_upper_level_column(
            upper_level_df, lower_level_df, 
            upper_level_suppression_column, code_lookup_field, 
            suppression_lookup_field
        )

def test_add_suppression_required_from_upper_level_column_missing_columns_lower_level():
    '''
    Test function raises a KeyError if the code_lookup_field or 'Rank' column are missing from the lower-level DataFrame.
    '''
    # Arrange: Test DataFrames with lower-level DataFrame missing the 'Rank' column
    upper_level_df = pd.DataFrame({
        'Org_Code': ['ICB1', 'ICB2'],
        'Suppression': [1, 0]
    })

    lower_level_df = pd.DataFrame({
        'Org_Code': ['ICB1', 'ICB1', 'ICB2', 'ICB2'],
        # Missing 'Rank' field
    })

    upper_level_suppression_column = 'Suppression_Required'
    code_lookup_field = 'Org_Code'
    suppression_lookup_field = 'Suppression'

    # Act & Assert: Expect KeyError due to missing 'Rank' column in lower_level_df
    with pytest.raises(KeyError, match="Missing required columns 'Org_Code' or 'Rank' in lower_level_df"):
        add_suppression_required_from_upper_level_column(
            upper_level_df, lower_level_df, 
            upper_level_suppression_column, code_lookup_field, 
            suppression_lookup_field
        )

def test_join_dataframes_success():
    '''
    Test the function successfully joins two DataFrames on a common column.
    '''
    # Arrange: Two test DataFrames with a common column to join on
    df1 = pd.DataFrame({
        'column_to_join_on': [1, 2, 3],
        'df1_data': ['A', 'B', 'C']
    })
    
    df2 = pd.DataFrame({
        'column_to_join_on': [1, 2, 3],
        'df2_data': ['X', 'Y', 'Z']
    })
    
    # Act:
    result_df = join_dataframes(df1, df2, on='column_to_join_on', how='left')

    # Assert: Check if the result matches the expected DataFrame
    expected_df = pd.DataFrame({
        'column_to_join_on': [1, 2, 3],
        'df1_data': ['A', 'B', 'C'],
        'df2_data': ['X', 'Y', 'Z']
    })

    pd.testing.assert_frame_equal(result_df, expected_df)

def test_join_dataframes_missing_column_df1():
    '''
    Test the function raises a KeyError if the join column is missing in df1.
    '''
    # Arrange: Two test DataFrames, df1 without the join column
    df1 = pd.DataFrame({
        'wrong_column': [1, 2, 3],
        'df1_data': ['A', 'B', 'C']
    })
    
    df2 = pd.DataFrame({
        'column_to_join_on': [1, 2, 3],
        'df2_data': ['X', 'Y', 'Z']
    })
    
    # Act & Assert: Expect KeyError due to missing join column in df1
    with pytest.raises(KeyError, match="Join column 'column_to_join_on' not found in one of the DataFrames."):
        join_dataframes(df1, df2, on='column_to_join_on', how='left')

def test_join_dataframes_missing_column_df2():
    '''
    Test the function raises a KeyError if the join column is missing in df2.
    '''
    # Arrange: Two test DataFrames, df2 without the join column
    df1 = pd.DataFrame({
        'column_to_join_on': [1, 2, 3],
        'df1_data': ['A', 'B', 'C']
    })
    
    df2 = pd.DataFrame({
        'wrong_column': [1, 2, 3],
        'df2_data': ['X', 'Y', 'Z']
    })
    
    # Act & Assert: Expect KeyError due to missing join column in df2
    with pytest.raises(KeyError, match="Join column 'column_to_join_on' not found in one of the DataFrames."):
        join_dataframes(df1, df2, on='column_to_join_on', how='left')

def test_join_dataframes_invalid_how():
    '''
    Test the function raises a ValueError if an invalid join type is specified.
    '''
    # Arrange: 
    df1 = pd.DataFrame({
        'column_to_join_on': [1, 2, 3],
        'df1_data': ['A', 'B', 'C']
    })
    
    df2 = pd.DataFrame({
        'column_to_join_on': [1, 2, 3],
        'df2_data': ['X', 'Y', 'Z']
    })
    
    # Act & Assert: Expect ValueError due to invalid 'how' argument
    with pytest.raises(ValueError, match="Invalid join type 'invalid_type'."):
        join_dataframes(df1, df2, on='column_to_join_on', how='invalid_type')

def test_replace_character_in_columns_success():
    '''
    Test function successfully repalces all specified characters in specified columns with alternative character.
    '''
    # Arrange: Test DataFrame
    df = pd.DataFrame({
        'Ward Name': ['Ward@1', 'Ward2!a', 'Ward=4', 'Ward 5', '1+Ward'],
        'Rank': [1, 2, 3, 4, 5]
    })
    target_column = 'Ward Name'
    target_chars = ['>', '<', '/', '+', '*', '!', '£', '$', '"', '%', '^', '=', '#', '@']
    replacement_char = '-'
    
    # Act:
    result = replace_character_in_columns(df, target_column, target_chars, replacement_char)

    # Assert: second-level suppression is applied only to the second submission for 'A' and 'B'
    expected = pd.DataFrame({
        'Ward Name': ['Ward-1', 'Ward2-a', 'Ward-4', 'Ward 5', '1-Ward'],
        'Rank': [1, 2, 3, 4, 5]
    })
    pd.testing.assert_frame_equal(result, expected)
    
def test_replace_character_in_columns_column_not_found():
    '''
    Test function generates correct error when column not found
    '''
    # Arrange: test DataFrame 
    df = pd.DataFrame({ # Ward Name field missing
        'Org Code': ['BKR', 'GRE', 'HER', 'DF3'],
        'Site Code': ['BKR0', 'GRE', 'RED', 'DFF34'],
        'Org Name': ['Hospital 1', 'IS Hospital 1', 'Hospital 2', 'IS Hospital 2']
        })

    target_column = 'Ward Name'
    target_chars = ['>', '<', '/', '+', '*', '!', '£', '$', '"', '%', '^', '=', '#', '@']
    replacement_char = '-'

    # Act & Assert KeyError should be raised for column not found)
    with pytest.raises(KeyError, match=r"Column 'Ward Name' not in DataFrame"):
        replace_character_in_columns(df, target_column, target_chars, replacement_char)    
    
def test_remove_duplicate_rows_success():
    '''
    Test function correctly removes duplicate rows.
    '''
       # Arrange: Test DataFrame
    df = pd.DataFrame({
        'Organisation Code': [1, 2, 2, 3],
        'Ward Name': ['A', 'B', 'B', 'C']
    })

    # Act:
    result = remove_duplicate_rows(df)

    # Assert: result matches expected
    expected = pd.DataFrame({
        'Organisation Code': [1, 2, 3],
        'Ward Name': ['A', 'B', 'C']
    }).reset_index(drop=True)
    pd.testing.assert_frame_equal(result.reset_index(drop=True), expected)

def test_limit_retained_columns_success():
    '''
    Test function retains only the specified columns.
    '''
    # Arrange: Create a DataFrame with multiple columns
    df = pd.DataFrame({
        'col1': [1, 2, 3],
        'col2': ['A', 'B', 'C'],
        'col3': [10, 20, 30]
    })
    
    # Define the columns to retain
    columns_to_retain = ['col1', 'col3']
    
    expected_df = pd.DataFrame({
        'col1': [1, 2, 3],
        'col3': [10, 20, 30]
    })

    # Act:
    result_df = limit_retained_columns(df, columns_to_retain)

    # Assert: The result should only contain the specified columns
    pd.testing.assert_frame_equal(result_df, expected_df)

def test_limit_retained_columns_single_column_string():
    '''
    Test the function works when a single column is passed as a string.
    '''
    # Arrange: Create a DataFrame with multiple columns
    df = pd.DataFrame({
        'col1': [1, 2, 3],
        'col2': ['A', 'B', 'C'],
        'col3': [10, 20, 30]
    })
    
    # Define the single column to retain (as a string)
    column_to_retain = 'col2'
    
    expected_df = pd.DataFrame({
        'col2': ['A', 'B', 'C']
    })
 
    # Act:
    result_df = limit_retained_columns(df, column_to_retain)
 
    # Assert: The result should contain only the single specified column
    pd.testing.assert_frame_equal(result_df, expected_df)

def test_limit_retained_columns_type_error():
    '''
    Test function raises a TypeError when 'columns' is not a string or list.
    '''
    # Arrange: Create a DataFrame
    df = pd.DataFrame({
        'col1': [1, 2, 3],
        'col2': ['A', 'B', 'C']
    })
    
    # Act & Assert: Expect a TypeError due to invalid type (integer passed instead of string/list)
    with pytest.raises(TypeError, match="columns should be a string or list of strings"):
        limit_retained_columns(df, 123)

def test_limit_retained_columns_key_error():
    '''
    Test function raises a KeyError when a non-existent column is specified.
    '''
    # Arrange: Create a DataFrame with some columns
    df = pd.DataFrame({
        'col1': [1, 2, 3],
        'col2': ['A', 'B', 'C']
    })
    
    # Act & Assert: Expect a KeyError when trying to retain a column that doesn't exist in the DataFrame
    with pytest.raises(KeyError, match="Required columns are missing form the DataFrame."):
        limit_retained_columns(df, ['col1', 'col3'])

def test_open_macro_excel_file_success(mocker):
    '''
    Test the function successfully opens an existing macro-enabled Excel file.
    '''
    # Arrange: Mock the load_workbook function to simulate opening a file directly from functions.etl_functions module
    mock_workbook = Workbook()
    mocker.patch('src.etl_functions.load_workbook', return_value=mock_workbook)

    file_path = r'mock_files\FFT-inpatient-data-template.xlsm'
    # Act: Call the function with a valid file path
    result_workbook = open_macro_excel_file(file_path)

    # Assert: Check if the workbook was opened correctly
    assert result_workbook is mock_workbook

def test_write_dataframes_to_sheets_success():
    '''
    Test that the function writes a DataFrame to an existing sheet correctly.
    '''
    # Arrange: Create a workbook and a sheet
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Sheet1'
    
    # Create a DataFrame to write
    df = pd.DataFrame({
        'col1': [1, 2],
        'col2': ['A', 'B']
    })
    
    dfs_info = [(df, 'Sheet1', 1, 1)]  # Write to Sheet1 starting at row 1, column 1

    # Act: Write the DataFrame to the sheet
    write_dataframes_to_sheets(workbook, dfs_info)

    # Assert: Check the content of the sheet
    assert sheet.cell(row=1, column=1).value == 1
    assert sheet.cell(row=1, column=2).value == 'A'
    assert sheet.cell(row=2, column=1).value == 2
    assert sheet.cell(row=2, column=2).value == 'B'

def test_write_dataframes_to_sheets_non_existent_sheet():
    '''
    Test that the function raises a ValueError when trying to write to a non-existent sheet.
    '''
    # Arrange: Create a workbook without a specific sheet
    workbook = Workbook()

    # Create a DataFrame to write
    df = pd.DataFrame({
        'col1': [1, 2],
        'col2': ['A', 'B']
    })

    dfs_info = [(df, 'NonExistentSheet', 1, 1)]  # Attempt to write to a non-existent sheet

    # Act & Assert: Expect a ValueError when trying to access a non-existent sheet
    with pytest.raises(ValueError, match="Sheet NonExistentSheet does not exist in the workbook."):
        write_dataframes_to_sheets(workbook, dfs_info)

def test_update_cell_with_formatting_success():
    '''
    Test function successfully updates multiple cells with applied formatting
    '''
    # Arrange: create dummy workbook and pass in data to update with 
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "TestSheet"
    data = [["A1", "B1"], ["A2", "B2"]]
    
    # Act
    update_cell_with_formatting(workbook, "TestSheet", 1, 1, data)

    # Assert: values in cells are correctly set
    assert sheet.cell(row=1, column=1).value == "A1"
    assert sheet.cell(row=1, column=2).value == "B1"
    assert sheet.cell(row=2, column=1).value == "A2"
    assert sheet.cell(row=2, column=2).value == "B2"

def test_update_cell_with_formatting_sheet_not_found():
    '''
    Test function raises a ValueError when trying to access a non-existent sheet.
    '''
    # Arrange:
    workbook = Workbook()
    
    # Act and Assert: If sheet not found ValueError is made
    with pytest.raises(ValueError, match="Sheet NonExistentSheet does not exist in the workbook."):
        update_cell_with_formatting(workbook, "NonExistentSheet", 1, 1, "Data")

def test_update_cell_with_formatting_single_cell_success():
    '''
    Test function successfully updates a single cell with applied formatting
    '''
    # Arrange: create dummy workbook
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "TestSheet"
    
    # Act:
    update_cell_with_formatting(workbook, "TestSheet", 1, 1, "Test Data", font_size=12, bg_color="0000FF00")
 
    # Assert: updates are made to single cell
    cell = sheet.cell(row=1, column=1)
    assert cell.value == "Test Data"
    assert cell.font.size == 12
    assert cell.fill.start_color.rgb == "0000FF00"
    assert cell.alignment.horizontal == "center"
    assert cell.alignment.vertical == "center"

def test_create_percentage_style_new():
    '''
    Test function successfully creates new style and correctly registers it in the workbook.
    '''
    # Arrange: create new workbook
    workbook = Workbook()

    # Act: 
    percentage_style = create_percentage_style(workbook)

    # Assert: 
    # Check if the percentage style has the expected properties
    assert percentage_style.name == "percentage_style"  # Name of the style
    assert percentage_style.number_format == '0%'  # Percentage format with 0 decimal places
    # Check if the style is registered in the workbook
    assert any(style == "percentage_style" for style in workbook.named_styles)

def test_create_percentage_style_invalid_workbook():
    '''
    Test function raises a TypeError when not presented with a workbook object to access.
    '''
    # Act and Assert:
    with pytest.raises(TypeError, match="The 'workbook' must be an openpyxl Workbook object."):
        create_percentage_style("not_a_workbook")

def test_format_column_as_percentage_success():
    '''
    Test function successfully applies percentage style within Excel workbook sheet.
    '''
    # Arrange: create dummy workbook with data and simulate percentage stlye to apply 
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "TestSheet"
    # Add dummy data to the sheet
    for row in range(1, 6):
        sheet.append([row * 0.01, row * 0.02])

    # Simulate percentage style with 0 decimal places
    percentage_style = NamedStyle(name="percentage_style")
    percentage_style.number_format = '0%'  

    # Act:
    format_column_as_percentage(workbook, "TestSheet", 1, [1, 2], percentage_style)

    # Assert: cells in the specified columns have percentage style applied
    for row in range(1, 6):  # From row 1 to 5
        cell1 = sheet.cell(row=row, column=1)
        cell2 = sheet.cell(row=row, column=2)
        assert cell1.style == "percentage_style"
        assert cell2.style == "percentage_style"

def test_format_column_as_percentage_invalid_sheet():
    '''
    Test function raises a ValueError when presented with a sheet not within the workbook.
    '''
    # Arrange: create dummy workbook with data and simulate percentage stlye to apply 
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "TestSheet"
    # Simulate percentage style with 0 decimal places
    percentage_style = NamedStyle(name="percentage_style")
    percentage_style.number_format = '0%'         

    # Act & Assert:
    with pytest.raises(ValueError, match="Sheet NonExistentSheet does not exist in this workbook."):
        format_column_as_percentage(workbook, "NonExistentSheet", 1, [1], percentage_style)

def test_combine_text_and_dataframe_cells_success():
    '''
    Test function correctly combines content of a DataFrame cell with a string value to create a new value.
    '''
    # Arrange: Create a DataFrame
    df = pd.DataFrame({
        'periodname': ['April-24', 'May-24', 'June-24'],
    })
    # get DataFrame value
    current_period = df.iat[2,0]
    # specify string to combine
    combo_prefix = 'Current month is'
        
    # Act:
    result = combine_text_and_dataframe_cells(combo_prefix, current_period)
    
    # Assert: result matches expected
    expected = 'Current month is June-24'
    assert result == expected

def test_save_macro_excel_file_existing(mocker):
    '''
    Test function can successfully save over existing workbook.
    '''
    # Arrange: mock workbook and file path
    workbook = MagicMock()  
    source_file_path = "original_file_path.xlsm"
    new_folder_path = "new_folder_path"  # Mocked new folder path (redundant here as overwriting existing)

    # Act:
    with patch("logging.info") as mock_logging:
        save_macro_excel_file(workbook, source_file_path, new_folder_path, new=False)

    # Assert:
    workbook.save.assert_called_once_with(source_file_path)  # Check if 'save' was called with the source file path
    mock_logging.assert_called_once_with(f"Workbook saved as {source_file_path}")  # Check the log message

def test_save_macro_excel_file_new_file(mocker):
    '''
    Test function can successfully save a new workbook.
    '''
    # Arrange: mock workbook and file path as well as file name details
    workbook = MagicMock()
    source_file_path = "original_file_path.xlsm"
    new_folder_path = "new_folder_path"
    prefix = "report"
    fft_period_suffix = "2024-01"

    # Act:
    with patch("os.path.join", return_value="new_folder_path/report-2024-01.xlsm") as mock_join, \
         patch("logging.info") as mock_logging:
        save_macro_excel_file(workbook, source_file_path, new_folder_path, new=True, prefix=prefix, fft_period_suffix=fft_period_suffix)

    # Assert:
    mock_join.assert_called_once_with(new_folder_path, "report-2024-01.xlsm")  # Check if os.path.join was called correctly
    workbook.save.assert_called_once_with("new_folder_path/report-2024-01.xlsm")  # Check if 'save' was called with the new file path
    mock_logging.assert_called_once_with(f"Workbook saved as new_folder_path/report-2024-01.xlsm")  # Check the log message
