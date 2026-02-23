#!/usr/bin/env python3
"""Formula Extraction Tool.

Extracts ALL formulas from ANY suppression workbook (.xlsm, .xls, etc.) and saves them
to a comprehensive markdown documentation file. This tool is CRUCIAL for understanding
and documenting the complex suppression logic in FFT workbooks.

Usage:
    python extract_formulas.py <input_file> [--output <output_dir>] [--verbose]

Example:
    python extract_formulas.py data/inputs/suppression_files/AE_Suppression_V3.5.xlsm \
        --output data/inputs/suppression_files/AE_FORMULAS \
        --verbose

"""

import argparse
import sys
from datetime import datetime
from pathlib import Path

import openpyxl


def extract_all_formulas(input_file, output_dir, verbose=False):
    """Extract ALL formulas from a suppression workbook and create.

    comprehensive documentation.

    Args:
        input_file (str): Path to the Excel workbook
        output_dir (str): Directory to save documentation
        verbose (bool): Whether to print detailed output

    Returns:
        bool: True if successful, False otherwise

    """
    try:
        # Convert to Path objects
        input_path = Path(input_file)
        output_path = Path(output_dir)

        if not input_path.exists():
            print(f"❌ Error: Input file '{input_file}' not found.")
            return False

        # Create output directory if it doesn't exist
        output_path.mkdir(parents=True, exist_ok=True)

        if verbose:
            print(f"🔍 Loading workbook: {input_file}")

        # Load the workbook
        wb = openpyxl.load_workbook(input_path, data_only=False)

        if verbose:
            print(f"📊 Found {len(wb.sheetnames)} sheets: {wb.sheetnames}")

        # Create markdown documentation
        doc_content = f"# {input_path.stem} Formulas\n\n"
        doc_content += (
            f"**Generated**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"
        )
        doc_content += f"**Source File**: {input_file}\n\n"
        doc_content += f"**Sheets**: {len(wb.sheetnames)}\n\n"

        # Analyze each sheet
        for sheet_name in wb.sheetnames:
            if verbose:
                print(f"  📄 Analyzing sheet: {sheet_name}")

            sheet = wb[sheet_name]
            sheet_formulas = {}

            # Find all formulas in the sheet
            for row in range(1, sheet.max_row + 1):
                for col in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=row, column=col)
                    if cell.data_type == "f" and cell.value:
                        cell_ref = f"{openpyxl.utils.get_column_letter(col)}{row}"
                        sheet_formulas[cell_ref] = cell.value

            if sheet_formulas:
                doc_content += f"## {sheet_name}\n\n"
                doc_content += (
                    f"**Dimensions**: {sheet.max_row} rows × "
                    f"{sheet.max_column} columns\n\n"
                )
                doc_content += f"**Formulas Found**: {len(sheet_formulas)}\n\n"

                # Group formulas by type for better organization
                ranking_formulas = {}
                suppression_formulas = {}
                aggregation_formulas = {}
                other_formulas = {}

                for cell_ref, formula in sheet_formulas.items():
                    formula_lower = formula.lower()

                    if "rank" in formula_lower or "if(a2=a1" in formula_lower:
                        ranking_formulas[cell_ref] = formula
                    elif (
                        "suppress" in formula_lower
                        or "sum(f2:h2" in formula_lower
                        or "vlookup" in formula_lower
                    ):
                        suppression_formulas[cell_ref] = formula
                    elif "sumif" in formula_lower or "sum(" in formula_lower:
                        aggregation_formulas[cell_ref] = formula
                    else:
                        other_formulas[cell_ref] = formula

                # Add formulas by category
                if ranking_formulas:
                    doc_content += "### Ranking Formulas\n\n"
                    for cell_ref, formula in sorted(ranking_formulas.items()):
                        doc_content += f"- **{cell_ref}**: `{formula}`\n"
                    doc_content += "\n"

                if suppression_formulas:
                    doc_content += "### Suppression Formulas\n\n"
                    for cell_ref, formula in sorted(suppression_formulas.items()):
                        doc_content += f"- **{cell_ref}**: `{formula}`\n"
                    doc_content += "\n"

                if aggregation_formulas:
                    doc_content += "### Aggregation Formulas\n\n"
                    for cell_ref, formula in sorted(aggregation_formulas.items()):
                        doc_content += f"- **{cell_ref}**: `{formula}`\n"
                    doc_content += "\n"

                if other_formulas:
                    doc_content += "### Other Formulas\n\n"
                    for cell_ref, formula in sorted(other_formulas.items()):
                        doc_content += f"- **{cell_ref}**: `{formula}`\n"
                    doc_content += "\n"

        # Save the documentation
        output_file = output_path / "FORMULAS.md"
        with open(output_file, "w", encoding="utf-8") as f:
            f.write(doc_content)

        if verbose:
            print(f"✅ Documentation saved to: {output_file}")

        print(f"🎉 Successfully extracted formulas from {len(wb.sheetnames)} sheets")
        print(f"📄 Documentation saved to: {output_file}")

        return True

    except Exception as e:
        print(f"❌ Error extracting formulas: {e}")
        return False


def main():
    """Extract formulas from suppression workbooks."""
    parser = argparse.ArgumentParser(
        description="Extract ALL formulas from ANY suppression workbook",
        epilog=(
            "Example: python extract_formulas.py "
            "data/inputs/suppression_files/AE_Suppression_V3.5.xlsm "
            "--output AE_FORMULAS --verbose"
        ),
    )

    parser.add_argument(
        "input_file", help="Path to the Excel workbook containing formulas"
    )

    parser.add_argument(
        "--output",
        "-o",
        default="FORMULAS",
        help="Output directory for documentation (default: FORMULAS)",
    )

    parser.add_argument(
        "--verbose",
        "-v",
        action="store_true",
        help="Print detailed output during extraction",
    )

    args = parser.parse_args()

    # Extract formulas
    success = extract_all_formulas(args.input_file, args.output, args.verbose)

    if not success:
        sys.exit(1)


if __name__ == "__main__":
    main()
