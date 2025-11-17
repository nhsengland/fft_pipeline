#!/usr/bin/env python
import openpyxl

def main():
    # Load files
    output = openpyxl.load_workbook('data/outputs/friends-and-family-test-inpatient-data-august-2025.xlsm', keep_vba=True)

    # Check each sheet
    for sheet_name in output.sheetnames:
        sheet = output[sheet_name]
        print(f"\nSheet: {sheet_name}")
        print(f"Dimensions: {sheet.dimensions}")

        # Print the first few rows to understand structure
        for i in range(1, min(10, sheet.max_row + 1)):
            row_values = []
            for j in range(1, min(6, sheet.max_column + 1)):
                row_values.append(sheet.cell(row=i, column=j).value)
            print(f"Row {i}: {row_values}")

if __name__ == "__main__":
    main()