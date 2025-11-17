#!/usr/bin/env python
import openpyxl
import sys
from pathlib import Path
import pandas as pd
import numpy as np

def check_cell(output_sheet, truth_sheet, row, col, description):
    """Check if a cell value matches between output and ground truth sheets."""
    output_val = output_sheet.cell(row=row, column=col).value
    truth_val = truth_sheet.cell(row=row, column=col).value

    # Handle different datetime formats
    if isinstance(output_val, str) and isinstance(truth_val, pd.Timestamp):
        output_val_str = output_val
        truth_val_str = truth_val.strftime('%b-%y')
        match = output_val_str == truth_val_str
    elif isinstance(truth_val, str) and isinstance(output_val, pd.Timestamp):
        truth_val_str = truth_val
        output_val_str = output_val.strftime('%b-%y')
        match = output_val_str == truth_val_str
    elif isinstance(output_val, (int, float)) and isinstance(truth_val, (int, float)):
        # For numeric values, allow small differences
        match = abs(output_val - truth_val) < 0.0001
    else:
        match = output_val == truth_val

    status = "✅ MATCH" if match else "❌ MISMATCH"
    print(f"{status} - {description}: Output={output_val}, Truth={truth_val}")
    return match

def compare_summary_sheet(output, truth):
    """Compare key metrics in the Summary sheet."""
    print("\n=== SUMMARY SHEET VALIDATION ===")

    summary_output = output['Summary']
    summary_truth = truth['Summary']

    # Check periods comparison
    periods_match = True
    current_period_match = check_cell(summary_output, summary_truth, 7, 6, "Current period")
    previous_period_match = check_cell(summary_output, summary_truth, 7, 8, "Previous period")
    periods_match = current_period_match and previous_period_match

    # Check key metrics
    total_responses_match = check_cell(summary_output, summary_truth, 8, 5, "Total responses")
    nhs_responses_match = check_cell(summary_output, summary_truth, 9, 5, "NHS responses")

    # Overall result for this sheet
    all_match = periods_match and total_responses_match and nhs_responses_match
    print(f"\nSummary sheet validation: {'✅ PASSED' if all_match else '❌ FAILED'}")
    return all_match

def compare_icb_sheet(output, truth):
    """Compare key metrics in the ICB sheet."""
    print("\n=== ICB SHEET VALIDATION ===")

    icb_output = output['ICB']
    icb_truth = truth['ICB']

    # Find row with actual data (can vary)
    data_start_row = None
    for i in range(10, 20):
        if icb_output.cell(row=i, column=1).value is not None:
            data_start_row = i
            break

    if not data_start_row:
        print("❌ FAILED - Could not find ICB data")
        return False

    # Check ICB count matches
    data_end_row = 50  # Safe assumption based on typical data
    icb_count_output = sum(1 for i in range(data_start_row, data_end_row)
                          if icb_output.cell(row=i, column=1).value is not None)

    # Find row with actual data in truth file
    truth_data_start_row = None
    for i in range(10, 20):
        if icb_truth.cell(row=i, column=1).value is not None:
            truth_data_start_row = i
            break

    if not truth_data_start_row:
        print("❌ FAILED - Could not find ICB data in ground truth")
        return False

    icb_count_truth = sum(1 for i in range(truth_data_start_row, data_end_row)
                         if icb_truth.cell(row=i, column=1).value is not None)

    print(f"{'✅ MATCH' if icb_count_output == icb_count_truth else '❌ MISMATCH'} - "
          f"ICB count: Output={icb_count_output}, Truth={icb_count_truth}")

    return icb_count_output == icb_count_truth

def compare_trust_counts(output, truth):
    """Compare trust counts."""
    print("\n=== TRUST COUNT VALIDATION ===")

    trusts_output = output['Trusts']
    trusts_truth = truth['Trusts']

    # Count non-empty rows in trusts sheet
    output_count = sum(1 for i in range(1, trusts_output.max_row)
                      if trusts_output.cell(row=i, column=1).value is not None)
    truth_count = sum(1 for i in range(1, trusts_truth.max_row)
                     if trusts_truth.cell(row=i, column=1).value is not None)

    print(f"{'✅ MATCH' if output_count == truth_count else '❌ MISMATCH'} - "
          f"Trust count: Output={output_count}, Truth={truth_count}")

    return output_count == truth_count

def run_validation(output_path, truth_path):
    """Run comprehensive validation between output and ground truth files."""
    try:
        # Load files
        output = openpyxl.load_workbook(output_path, keep_vba=True, data_only=True)
        truth = openpyxl.load_workbook(truth_path, keep_vba=True, data_only=True)

        print(f"Validating:\n - Output: {output_path}\n - Ground Truth: {truth_path}")

        # Ensure both files have the same sheets
        output_sheets = set(output.sheetnames)
        truth_sheets = set(truth.sheetnames)

        if output_sheets != truth_sheets:
            missing = truth_sheets - output_sheets
            extra = output_sheets - truth_sheets
            print(f"❌ SHEET MISMATCH - Missing: {missing}, Extra: {extra}")
            return False

        # Validate key sheets
        summary_valid = compare_summary_sheet(output, truth)
        icb_valid = compare_icb_sheet(output, truth)
        trust_valid = compare_trust_counts(output, truth)

        # Overall validation result
        all_valid = summary_valid and icb_valid and trust_valid

        print("\n=== VALIDATION SUMMARY ===")
        print(f"Summary sheet: {'✅ PASSED' if summary_valid else '❌ FAILED'}")
        print(f"ICB sheet: {'✅ PASSED' if icb_valid else '❌ FAILED'}")
        print(f"Trust counts: {'✅ PASSED' if trust_valid else '❌ FAILED'}")
        print(f"\nOverall validation: {'✅ PASSED' if all_valid else '❌ FAILED'}")

        return all_valid

    except Exception as e:
        print(f"❌ ERROR during validation: {str(e)}")
        return False

def main():
    # Default paths
    output_path = Path("data/outputs/friends-and-family-test-inpatient-data-august-2025.xlsm")
    truth_path = Path("data/outputs/ground_truth/friends-and-family-test-inpatient-data-august-2025.xlsm")

    # Allow custom paths via command line
    if len(sys.argv) > 1:
        output_path = Path(sys.argv[1])
    if len(sys.argv) > 2:
        truth_path = Path(sys.argv[2])

    # Run validation
    success = run_validation(output_path, truth_path)

    # Exit with appropriate code
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()