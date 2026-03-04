"""Excel output functions."""

import logging
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.workbook import Workbook

from fft.config import (
    BS_SHEET_CONFIG,
    ENGLAND_ROWS_SKIP_COLUMNS,
    ENGLAND_TOTALS_DATA_SOURCE,
    IS1_CODE,
    OUTPUT_COLUMNS,
    OUTPUTS_DIR,
    PERCENTAGE_COLUMN_CONFIG,
    PERCENTAGE_NUMBER_FORMAT,
    PERIOD_LABEL_CONFIG,
    SPECIALITY_COLS,
    SUMMARY_SHEET_CONFIG,
    TEMPLATE_CONFIG,
    TEMPLATES_DIR,
    get_count_columns_for_service,
)
from fft.loaders import load_collections_overview
from fft.processors import extract_summary_data

logger = logging.getLogger(__name__)


def load_template(service_type: str) -> Workbook:
    """Load Excel template for the specified service type, preserving VBA macros.

    Args:
        service_type: 'inpatient', 'ae', or 'ambulance'

    Returns:
        Openpyxl Workbook object with VBA preserved

    Raises:
        KeyError: If service_type is not configured
        FileNotFoundError: If template file doesn't exist

    >>> from src.fft.writers import load_template
    >>> wb = load_template('inpatient')
    >>> type(wb).__name__
    'Workbook'
    >>> 'ICB' in wb.sheetnames
    True

    # Edge case: Unknown service type
    >>> load_template('unknown_service')
    Traceback (most recent call last):
        ...
    KeyError: "Unknown service type: 'unknown_service'"

    # Edge case: Missing template file
    >>> from fft.config import TEMPLATE_CONFIG
    >>> TEMPLATE_CONFIG['test_missing'] = {'template_file': 'nonexistent.xlsm'}
    >>> load_template('test_missing')  # doctest: +ELLIPSIS
    Traceback (most recent call last):
        ...
    FileNotFoundError: Template not found: nonexistent.xlsm
    >>> del TEMPLATE_CONFIG['test_missing']

    """
    if service_type not in TEMPLATE_CONFIG:
        raise KeyError(f"Unknown service type: '{service_type}'")

    template_file = TEMPLATE_CONFIG[service_type]["template_file"]
    template_path = TEMPLATES_DIR / template_file

    if not template_path.exists():
        raise FileNotFoundError(f"Template not found: {template_file}")

    return load_workbook(template_path, keep_vba=True)


# %%
def _has_formula(cell) -> bool:
    """Check if an openpyxl cell contains a formula.

    Args:
        cell: openpyxl cell object

    Returns:
        bool: True if cell contains a formula, False otherwise

    """
    if cell.value is None:
        return False
    return isinstance(cell.value, str) and cell.value.startswith("=")


def write_dataframe_to_sheet(
    workbook: Workbook,
    df: pd.DataFrame,
    sheet_name: str,
    start_row: int,
    start_col: int = 1,
    service_type: str = None,
) -> None:
    """Write DataFrame contents to a specific sheet location.

    Writes data without headers - assumes template already has headers in place.
    Sets proper Excel data types: numbers, percentages, and general text.

    Args:
        workbook: Openpyxl Workbook object
        df: DataFrame to write
        sheet_name: Name of target sheet
        start_row: Row number to start writing (1-indexed)
        start_col: Column number to start writing (1-indexed, default 1)
        service_type: Service type ('inpatient', 'ae', 'ambulance') for percentage detection

    Returns:
        None (modifies workbook in place)

    Raises:
        KeyError: If sheet_name doesn't exist in workbook

    >>> from src.fft.writers import load_template, write_dataframe_to_sheet
    >>> import pandas as pd
    >>> wb = load_template('inpatient')
    >>> df = pd.DataFrame({
    ...     'ICB_Code': ['ABC', 'DEF'],
    ...     'ICB_Name': ['Test ICB 1', 'Test ICB 2'],
    ...     'Total Responses': [100, 200]
    ... })
    >>> write_dataframe_to_sheet(wb, df, 'ICB', start_row=15, start_col=1, service_type='inpatient')
    >>> wb['ICB'].cell(row=15, column=1).value
    'ABC'
    >>> wb['ICB'].cell(row=16, column=2).value
    'Test ICB 2'

    # Edge case: Sheet doesn't exist
    >>> write_dataframe_to_sheet(wb, df, 'NonExistent', start_row=15, service_type='inpatient')
    Traceback (most recent call last):
        ...
    KeyError: "Sheet 'NonExistent' not found in workbook"

    # Edge case: Empty DataFrame
    >>> df_empty = pd.DataFrame(columns=['A', 'B'])
    >>> write_dataframe_to_sheet(wb, df_empty, 'ICB', start_row=20, service_type='inpatient')
    >>> wb['ICB'].cell(row=20, column=1).value is None
    True

    """
    if sheet_name not in workbook.sheetnames:
        raise KeyError(f"Sheet '{sheet_name}' not found in workbook")

    sheet = workbook[sheet_name]

    # Get percentage columns for this sheet if service_type provided
    percentage_columns = set()
    if service_type and service_type in PERCENTAGE_COLUMN_CONFIG:
        if sheet_name in PERCENTAGE_COLUMN_CONFIG[service_type]:
            percentage_columns = set(PERCENTAGE_COLUMN_CONFIG[service_type][sheet_name])

    for row_idx, row in enumerate(df.itertuples(index=False), start=start_row):
        for col_idx, cell_value in enumerate(row, start=start_col):
            cell = sheet.cell(row=row_idx, column=col_idx)

            # Convert NaN values to dashes to match VBA behaviour
            if pd.isna(cell_value):
                cell.value = "-"
                cell.number_format = "General"
            else:
                cell.value = cell_value

                # Set appropriate number format based on data type and column
                if col_idx in percentage_columns and isinstance(cell_value, (int, float)) and cell_value != "*":
                    # Set percentage format
                    cell.number_format = PERCENTAGE_NUMBER_FORMAT
                elif isinstance(cell_value, (int, float)):
                    # Set number format with thousands separator for integers, general for floats
                    if isinstance(cell_value, int) or (isinstance(cell_value, float) and cell_value.is_integer()):
                        cell.number_format = "#,##0"  # Thousands separator for whole numbers
                    else:
                        cell.number_format = "General"  # Default for decimals
                else:
                    # Text values get general format
                    cell.number_format = "General"


def add_terminating_row(
    workbook: Workbook,
    sheet_name: str,
    terminating_row: int,
    num_columns: int,
) -> None:
    """Add a non-table row after data and trim sheet to prevent scrolling to row 999.

    Args:
        workbook: Openpyxl Workbook object
        sheet_name: Name of target sheet
        terminating_row: Row number to add terminating row (1-indexed)
        num_columns: Number of columns to fill with dashes

    """
    if sheet_name not in workbook.sheetnames:
        return

    sheet = workbook[sheet_name]

    # Add dashes to all data columns in the terminating row
    for col_idx in range(1, num_columns + 1):
        sheet.cell(row=terminating_row, column=col_idx).value = "-"

    # Delete excess rows and columns to prevent infinite scrolling
    if sheet.max_row > terminating_row:
        sheet.delete_rows(terminating_row + 1, sheet.max_row - terminating_row)
    if sheet.max_column > num_columns:
        sheet.delete_cols(num_columns + 1, sheet.max_column - num_columns)


# %%
def write_bs_lookup_data(
    workbook: Workbook, ward_df: pd.DataFrame, service_type: str
) -> None:
    """Populate BS sheet lookup tables for dropdown filtering.

    Writes reference lists and linked lists used by VBA macros for filtering.

    The BS sheet has three main sections:
    1. Reference Lists (U:Z) - Full hierarchy from ward data
    2. Region Reference (O:P, R:S) - ICB data for region-level dropdown support
    (if configured)
    3. Linked Lists (AA+) - Deduplicated pairs, each column sorted independently

    For A&E service type, ICBs are treated as "regions" for VBA compatibility.
    The VBA dropdown cascade works: Region(ICB) → Trust → Site

    The VBA logic for linked lists:
    - Columns are grouped into pairs (e.g., Trust Code + Trust Name)
    - Each pair is deduplicated together (keeping Code/Name aligned)
    - Then each column is sorted independently (breaking the pairing)

    Args:
        workbook: Openpyxl Workbook object
        ward_df: DataFrame containing full ward-level data with hierarchy columns
        service_type: 'inpatient', 'ae', or 'ambulance'

    Returns:
        None (modifies workbook in place)

    Raises:
        KeyError: If BS sheet doesn't exist or required columns missing

    >>> from fft.writers import load_template, write_bs_lookup_data
    >>> import pandas as pd
    >>> wb = load_template('inpatient')
    >>> df = pd.DataFrame({
    ...     'ICB_Code': ['QE1', 'QE1', 'QWO'],
    ...     'Trust_Code': ['RXA', 'RXA', 'RY6'],
    ...     'Trust_Name': ['Trust Alpha', 'Trust Alpha', 'Trust Beta'],
    ...     'Site_Code': ['RXA01', 'RXA02', 'RY601'],
    ...     'Site_Name': ['Site One', 'Site Two', 'Site Three'],
    ...     'Ward_Name': ['Ward A', 'Ward B', 'Ward C']
    ... })
    >>> write_bs_lookup_data(wb, df, 'inpatient')

    # Reference list (U:Z) starts at row 2
    >>> wb['BS'].cell(row=2, column=21).value  # U2 = ICB_Code
    'QE1'
    >>> wb['BS'].cell(row=2, column=26).value  # Z2 = Ward_Name
    'Ward A'

    # Trusts linked list (AE:AF) - dedupe pairs, sort independently
    >>> wb['BS'].cell(row=1, column=31).value  # AE1 = Template header (unchanged)
    'Code…'
    >>> wb['BS'].cell(row=2, column=31).value  # AE2 = First Trust Code (sorted)
    'RXA'
    >>> wb['BS'].cell(row=3, column=31).value  # AE3 = Second Trust Code (sorted)
    'RY6'
    >>> wb['BS'].cell(row=1, column=32).value  # AF1 = Template header (unchanged)
    'Name…'
    >>> wb['BS'].cell(row=2, column=32).value  # AF2 = First Trust Name (sorted)
    'Trust Alpha'

    # Edge case: Missing BS sheet
    >>> from openpyxl import Workbook as NewWorkbook
    >>> wb_no_bs = NewWorkbook()
    >>> write_bs_lookup_data(wb_no_bs, df, 'inpatient')
    Traceback (most recent call last):
        ...
    KeyError: "Sheet 'BS' not found in template workbook"

    # Edge case: Missing required columns
    >>> df_missing = pd.DataFrame({'ICB_Code': ['QE1']})
    >>> wb2 = load_template('inpatient')
    >>> write_bs_lookup_data(wb2, df_missing, 'inpatient')
    Traceback (most recent call last):
        ...
    KeyError: "Required column 'Trust_Code' not found in DataFrame"

    """
    if "BS" not in workbook.sheetnames:
        raise KeyError("Sheet 'BS' not found in template workbook")

    if service_type not in BS_SHEET_CONFIG:
        raise KeyError(f"Unknown service type: '{service_type}'")

    config = BS_SHEET_CONFIG[service_type]
    sheet = workbook["BS"]

    # Validate required columns exist
    ref_cols = config["reference_columns"]
    for col in ref_cols:
        if col not in ward_df.columns:
            raise KeyError(f"Required column '{col}' not found in DataFrame")

    # 1. Write Reference Lists (full hierarchy to U:Z)
    ref_start_col = config["reference_list_start_col"]
    ref_start_row = config["reference_list_start_row"]

    ref_data = ward_df[ref_cols].copy()
    for row_idx, row in enumerate(ref_data.itertuples(index=False), start=ref_start_row):
        for col_idx, value in enumerate(row, start=ref_start_col):
            sheet.cell(row=row_idx, column=col_idx).value = value

    # 2. Write Region Reference Data (if configured)
    if "region_reference" in config:
        region_config = config["region_reference"]
        region_start_col = region_config["start_col"]
        region_start_row = region_config["start_row"]
        region_pairs = region_config["pairs"]

        # For A&E, treat ICBs as regions by using ICB data for region reference
        col_offset = 0
        for pair in region_pairs:
            # Use unique ICB pairs from the data
            unique_pair = ward_df[pair].drop_duplicates()

            # Sort pairs together to maintain code-to-name relationship
            if len(pair) > 1:
                # Sort by first column (typically the code) to maintain pairing
                sorted_pair = unique_pair.astype(str).sort_values(by=pair[0]).reset_index(drop=True)
            else:
                sorted_pair = unique_pair.astype(str).sort_values(by=pair[0]).reset_index(drop=True)

            # Write each column from the sorted pairs
            for pair_col_idx, col_name in enumerate(pair):
                sorted_values = sorted_pair[col_name]

                for row_idx, value in enumerate(sorted_values, start=region_start_row):
                    sheet.cell(
                        row=row_idx, column=region_start_col + col_offset + pair_col_idx
                    ).value = value

            col_offset += len(pair)

    # 3. Write Linked Lists (dedupe pairs, sort pairs together maintaining relationships)
    for level, level_config in config["linked_lists"].items():
        start_col = level_config["start_col"]
        pairs = level_config["pairs"]

        col_offset = 0
        for pair in pairs:
            # Deduplicate the pair together
            unique_pair = ward_df[pair].drop_duplicates()

            # Sort pairs together to maintain relationships (e.g., code-to-name pairing)
            if len(pair) > 1:
                # Sort by first column (typically the code) to maintain pairing
                sorted_pair = unique_pair.astype(str).sort_values(by=pair[0]).reset_index(drop=True)
            else:
                sorted_pair = unique_pair.astype(str).sort_values(by=pair[0]).reset_index(drop=True)

            # Write each column from the sorted pairs
            for pair_col_idx, col_name in enumerate(pair):
                sorted_values = sorted_pair[col_name]

                for row_idx, value in enumerate(sorted_values, start=2):
                    sheet.cell(
                        row=row_idx, column=start_col + col_offset + pair_col_idx
                    ).value = value

            col_offset += len(pair)


# %%
def update_period_labels(workbook: Workbook, service_type: str, fft_period: str) -> None:
    """Update period-dependent labels in the workbook.

    Updates cells containing dynamic period text (e.g., "August 2024").

    Args:
        workbook: Openpyxl Workbook object
        service_type: 'inpatient', 'ae', or 'ambulance'
        fft_period: FFT period string (e.g., 'Aug-24')

    Returns:
        None (modifies workbook in place)

    Raises:
        KeyError: If service_type is not configured or sheet doesn't exist

    >>> from src.fft.writers import load_template, update_period_labels
    >>> wb = load_template('inpatient')
    >>> update_period_labels(wb, 'inpatient', 'Aug-24')
    >>> wb['Notes'].cell(row=2, column=1).value
    'Inpatient Friends and Family Test (FFT) Data - Aug-24'

    # Edge case: Unknown service type
    >>> update_period_labels(wb, 'unknown', 'Aug-24')
    Traceback (most recent call last):
        ...
    KeyError: "Unknown service type: 'unknown'"

    # Edge case: Missing sheet
    >>> from openpyxl import Workbook as NewWorkbook
    >>> wb_no_notes = NewWorkbook()
    >>> update_period_labels(wb_no_notes, 'inpatient', 'Aug-24')
    Traceback (most recent call last):
        ...
    KeyError: "Sheet 'Notes' not found in workbook"

    # Edge case: Missing configuration for service type - test graceful handling
    >>> from fft.config import PERIOD_LABEL_CONFIG
    >>> PERIOD_LABEL_CONFIG['test_missing'] = {
    ...     'notes_title': {
    ...         'sheet': 'Notes',
    ...         'cell': 'A2',
    ...         'template': 'Test FFT Data - {period}',
    ...     }
    ... }
    >>> update_period_labels(wb, 'test_missing', 'Aug-24') # Should work with valid config
    >>> del PERIOD_LABEL_CONFIG['test_missing']

    # Edge case: Empty configuration for service type - no labels to update
    >>> PERIOD_LABEL_CONFIG['test_empty'] = {}
    >>> update_period_labels(wb, 'test_empty', 'Aug-24')  # Handle empty config gracefully
    >>> del PERIOD_LABEL_CONFIG['test_empty']

    """
    if service_type not in PERIOD_LABEL_CONFIG:
        raise KeyError(f"Unknown service type: '{service_type}'")

    config = PERIOD_LABEL_CONFIG[service_type]

    for label_name, label_config in config.items():
        sheet_name = label_config["sheet"]
        cell = label_config["cell"]
        template = label_config["template"]

        if sheet_name not in workbook.sheetnames:
            raise KeyError(f"Sheet '{sheet_name}' not found in workbook")

        sheet = workbook[sheet_name]
        sheet[cell].value = template.format(period=fft_period)

        # FIXME: This is a temporary fix
        # Fix A&E Notes sheet rows 39-40 alignment
        if service_type == "ae" and sheet_name == "Notes":
            for row in [39, 40]:
                for col in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=row, column=col)
                    cell.alignment = Alignment(horizontal="left", wrap_text=True)


# %%
def write_england_totals(
    workbook: Workbook,
    service_type: str,
    national_df: pd.DataFrame,
    org_counts: dict,
    data_options: dict = None,
) -> None:
    """Write England-level totals to rows 12-14 of output sheets.

    Writes three rows:
    - Row 12: England (including Independent Sector Providers)
    - Row 13: England (excluding Independent Sector Providers)
    - Row 14: Selection (excluding suppressed data) - placeholder

    Args:
        workbook: Openpyxl Workbook object
        service_type: 'inpatient', 'ae', or 'ambulance'
        national_df: DataFrame from aggregate_to_national() with Total/NHS/IS1 rows
        org_counts: Dict with 'total_count', 'nhs_count', 'is1_count'
        data_options: Optional dict with 'suppressed_data' and 'all_level_data' keys

    Returns:
        None (modifies workbook in place)

    Raises:
        KeyError: If service_type not configured or required data missing

    >>> from src.fft.writers import load_template, write_england_totals
    >>> import pandas as pd
    >>> wb = load_template('inpatient')
    >>> nat_df = pd.DataFrame({
    ...     'Submitter_Type': ['Total', 'NHS'],
    ...     'Total Responses': [1000, 800],
    ...     'Total Eligible': [5000, 4000],
    ...     'Percentage_Positive': [0.95, 0.94],
    ...     'Percentage_Negative': [0.02, 0.03],
    ...     'Very Good': [800, 650],
    ...     'Good': [150, 120],
    ...     'Neither Good nor Poor': [30, 20],
    ...     'Poor': [15, 8],
    ...     'Very Poor': [5, 2],
    ...     "Don't Know": [0, 0]
    ... })
    >>> counts = {'total_count': 150, 'nhs_count': 130, 'is1_count': 20}
    >>> write_england_totals(wb, 'inpatient', nat_df, counts)
    >>> wb['ICB'].cell(row=12, column=3).value
    np.int64(1000)
    >>> wb['ICB'].cell(row=12, column=5).value
    np.float64(0.95)

    # Edge case: Missing Submitter_Type
    >>> bad_df = pd.DataFrame({'Total Responses': [1000]})
    >>> write_england_totals(wb, 'inpatient', bad_df, counts)
    Traceback (most recent call last):
        ...
    KeyError: "'Submitter_Type' column not found in national_df"

    """
    # Extract data options with defaults
    data_options = data_options or {}
    all_level_data = data_options.get("all_level_data")

    _validate_england_totals_inputs(service_type, national_df)

    config = TEMPLATE_CONFIG[service_type]
    england_rows = config["england_rows"]

    # Process each sheet
    for level, sheet_config in config["sheets"].items():
        sheet_name = sheet_config["sheet_name"]

        if sheet_name not in workbook.sheetnames:
            continue

        config = {
            "sheet_config": sheet_config,
            "service_type": service_type,
            "england_rows": england_rows,
        }
        options = {"all_level_data": all_level_data}
        _process_single_sheet(workbook, sheet_name, config, national_df, options)


def _validate_england_totals_inputs(service_type: str, national_df: pd.DataFrame) -> None:
    """Validate inputs for England totals writing."""
    if service_type not in TEMPLATE_CONFIG:
        raise KeyError(f"Unknown service type: '{service_type}'")

    if "Submitter_Type" not in national_df.columns:
        raise KeyError("'Submitter_Type' column not found in national_df")


def _process_single_sheet(
    workbook: Workbook,
    sheet_name: str,
    config: dict,
    national_df: pd.DataFrame,
    options: dict = None,
) -> None:
    """Process a single sheet for England totals."""
    # Extract parameters
    sheet_config = config["sheet_config"]
    service_type = config["service_type"]
    england_rows = config["england_rows"]

    options = options or {}
    all_level_data = options.get("all_level_data")

    sheet = workbook[sheet_name]

    # Get data for this sheet
    total_row, nhs_row = _get_sheet_data(
        sheet_name, national_df, all_level_data, service_type
    )

    # Get output configuration
    output_cols, data_columns, name_col_idx = _get_sheet_configuration(
        sheet_name, sheet_config, service_type
    )

    # Write England including IS row
    england_config = {
        "england_rows": england_rows,
        "name_col_idx": name_col_idx,
        "data_columns": data_columns,
        "output_cols": output_cols,
        "service_type": service_type,
    }
    _write_england_including_is_row(sheet, england_config, total_row)

    # Write England excluding IS row
    _write_england_excluding_is_row(sheet, england_config, nhs_row)

    # Write Selection row
    _write_selection_row(sheet, england_config, nhs_row)


def _get_sheet_data(
    sheet_name: str,
    national_df: pd.DataFrame,
    all_level_data: dict = None,
    service_type: str = "inpatient",
) -> tuple:
    """Get data for a specific sheet."""
    # Determine appropriate data source for this sheet
    data_source_level = ENGLAND_TOTALS_DATA_SOURCE.get(sheet_name)

    # Use sheet-appropriate data if available, fallback to national_df
    if all_level_data and data_source_level and data_source_level in all_level_data:
        return _get_data_from_level(
            all_level_data[data_source_level], all_level_data, service_type
        )
    else:
        return _get_data_from_national(national_df, service_type)


def _get_data_from_level(
    level_df: pd.DataFrame, all_level_data: dict = None, service_type: str = "inpatient"
) -> tuple:
    """Get data from level-specific DataFrame."""
    # Calculate totals excluding IS1 (NHS only)
    if "ICB_Code" in level_df.columns:
        nhs_df = level_df[level_df["ICB_Code"] != IS1_CODE]
    else:
        nhs_df = level_df

    # Select count columns
    count_cols = _get_count_columns(level_df, service_type)
    available_count_cols = [col for col in count_cols if col in level_df.columns]

    # Calculate totals
    nhs_totals = nhs_df[available_count_cols].sum()
    all_totals = level_df[available_count_cols].sum()

    # Create DataFrames
    total_row = pd.DataFrame([all_totals], index=[0])
    nhs_row = pd.DataFrame([nhs_totals], index=[0])

    # Recalculate percentages
    _recalculate_percentages(total_row)
    _recalculate_percentages(nhs_row)

    return total_row, nhs_row
def _get_count_columns(level_df: pd.DataFrame, service_type: str = "inpatient") -> list:
    """Get list of count columns for aggregation."""
    return get_count_columns_for_service(service_type)


def _recalculate_percentages(row_df: pd.DataFrame) -> None:
    """Recalculate percentage columns for a DataFrame row."""
    if all(col in row_df.columns for col in ["Very Good", "Good", "Total Responses"]):
        row_df["Percentage_Positive"] = (
            (row_df["Very Good"] + row_df["Good"]) / row_df["Total Responses"]
        ).round(4)

    if all(col in row_df.columns for col in ["Poor", "Very Poor", "Total Responses"]):
        row_df["Percentage_Negative"] = (
            (row_df["Poor"] + row_df["Very Poor"]) / row_df["Total Responses"]
        ).round(4)


def _get_data_from_national(
    national_df: pd.DataFrame, service_type: str = "inpatient"
) -> tuple:
    """Get data from national DataFrame."""
    if "Submitter_Type" not in national_df.columns:
        raise KeyError("'Submitter_Type' column not found in national_df")

    total_row = national_df[national_df["Submitter_Type"] == "Total"]
    nhs_row = national_df[national_df["Submitter_Type"] == "NHS"]

    if total_row.empty or nhs_row.empty:
        raise KeyError("national_df must contain 'Total' and 'NHS' rows")

    return total_row, nhs_row


def _get_sheet_configuration(
    sheet_name: str, sheet_config: dict, service_type: str
) -> tuple:
    """Get configuration for a specific sheet."""
    output_cols = OUTPUT_COLUMNS[service_type].get(sheet_name, [])
    skip_cols = ENGLAND_ROWS_SKIP_COLUMNS.get(sheet_name, 0)
    data_columns = output_cols[skip_cols:]
    name_col_idx = output_cols.index(sheet_config["england_label_column"]) + 1

    return output_cols, data_columns, name_col_idx


def _write_england_including_is_row(
    sheet,
    config: dict,
    total_row: pd.DataFrame,
) -> None:
    """Write England including IS row."""
    # Extract config parameters
    england_rows = config["england_rows"]
    name_col_idx = config["name_col_idx"]
    data_columns = config["data_columns"]
    output_cols = config["output_cols"]
    service_type = config["service_type"]

    # Write label
    england_label = (
        "England"
        if service_type == "ae"
        else "England (including Independent Sector Providers)"
    )
    sheet.cell(
        row=england_rows["including_is"], column=name_col_idx
    ).value = england_label

    # Write data
    for col_name in data_columns:
        if col_name in output_cols and col_name in total_row.columns:
            col_idx = output_cols.index(col_name) + 1
            value = total_row[col_name].values[0]
            if pd.isna(value):
                value = "-"
            sheet.cell(row=england_rows["including_is"], column=col_idx).value = value

    # Write dashes to speciality columns
    for col_name in SPECIALITY_COLS:
        if col_name in output_cols:
            col_idx = output_cols.index(col_name) + 1
            sheet.cell(row=england_rows["including_is"], column=col_idx).value = "-"


def _write_england_excluding_is_row(
    sheet,
    config: dict,
    nhs_row: pd.DataFrame,
) -> None:
    """Write England excluding IS row."""
    # Extract config parameters
    england_rows = config["england_rows"]
    name_col_idx = config["name_col_idx"]
    data_columns = config["data_columns"]
    output_cols = config["output_cols"]

    # Skip if same row as including IS (A&E case)
    if england_rows["excluding_is"] != england_rows["including_is"]:
        sheet.cell(
            row=england_rows["excluding_is"], column=name_col_idx
        ).value = "England (excluding Independent Sector Providers)"

        # Write data
        for col_name in data_columns:
            if col_name in output_cols and col_name in nhs_row.columns:
                col_idx = output_cols.index(col_name) + 1
                value = nhs_row[col_name].values[0]
                if pd.isna(value):
                    value = "-"
                sheet.cell(row=england_rows["excluding_is"], column=col_idx).value = value

        # Write dashes to speciality columns
        for col_name in SPECIALITY_COLS:
            if col_name in output_cols:
                col_idx = output_cols.index(col_name) + 1
                sheet.cell(row=england_rows["excluding_is"], column=col_idx).value = "-"


def _write_selection_row(
    sheet,
    config: dict,
    nhs_row: pd.DataFrame,
) -> None:
    """Write Selection row."""
    # Extract config parameters
    england_rows = config["england_rows"]
    name_col_idx = config["name_col_idx"]
    data_columns = config["data_columns"]
    output_cols = config["output_cols"]

    # Write label
    sheet.cell(
        row=england_rows["selection"], column=name_col_idx
    ).value = "Selection (excluding suppressed data)"

    # Write data, preserving formulas
    for col_name in data_columns:
        if col_name in output_cols and col_name in nhs_row.columns:
            col_idx = output_cols.index(col_name) + 1
            cell = sheet.cell(row=england_rows["selection"], column=col_idx)

            # Only write if cell doesn't already contain a formula
            if not _has_formula(cell):
                cell.value = nhs_row[col_name].values[0]

    # Cache formula results for validation
    # This ensures formulas work correctly when workbook is read with data_only=True
    _cache_all_formula_results(sheet.parent)


# %%
def get_cached_formula_results(sheet, row: int = None) -> dict:
    """Retrieve cached formula results from sheet metadata.

    Args:
        sheet: Openpyxl worksheet object
        row: Optional row number to filter results. If None, returns all cached results.

    Returns:
        Dict with cell coordinates as keys and calculated values as values.
        If row specified, returns only results for that row.
        If no cached results exist, returns empty dict.

    """
    if not hasattr(sheet, "_fft_cached_formulas"):
        return {}

    if row is None:
        # Return all cached results flattened
        all_results = {}
        for row_results in sheet._fft_cached_formulas.values():
            all_results.update(row_results)
        return all_results

    return sheet._fft_cached_formulas.get(row, {})


def _cache_all_formula_results(workbook: Workbook) -> None:
    """Cache all formula results in the workbook for validation.

    Calculates and caches results for all formulas across all sheets.
    Handles SUBTOTAL and IFERROR formula types commonly used in FFT templates.
    Processes formulas in dependency order: SUBTOTAL first, then IFERROR.

    Args:
        workbook: Openpyxl Workbook object

    Note:
        Results are stored in sheet._fft_cached_formulas[row][coordinate] format.

    """
    for sheet in workbook.worksheets:
        sheet._fft_cached_formulas = {}

        # Collect all formula cells, grouped by type and row
        formula_cells_by_row = {}
        for row in sheet.iter_rows():
            for cell in row:
                if cell.data_type == "f" and isinstance(cell.value, str):
                    row_num = cell.row
                    if row_num not in formula_cells_by_row:
                        formula_cells_by_row[row_num] = []
                    formula_cells_by_row[row_num].append(cell)

        # Process each row
        for row_num in sorted(formula_cells_by_row.keys()):
            cells = formula_cells_by_row[row_num]
            row_cache = {}

            # First pass: Calculate SUBTOTAL formulas (no dependencies)
            for cell in cells:
                if "SUBTOTAL(" in cell.value:
                    try:
                        result = _calculate_formula_result(sheet, cell)
                        if result is not None:
                            row_cache[cell.coordinate] = result
                    except (IndexError, ValueError, AttributeError, ZeroDivisionError):
                        pass

            # Update sheet cache after first pass so IFERROR formulas can use results
            if row_cache:
                sheet._fft_cached_formulas[row_num] = row_cache

            # Second pass: Calculate IFERROR formulas (may depend on SUBTOTAL results)
            for cell in cells:
                if "IFERROR(" in cell.value and cell.coordinate not in row_cache:
                    try:
                        result = _calculate_formula_result(sheet, cell)
                        if result is not None:
                            row_cache[cell.coordinate] = result
                    except (IndexError, ValueError, AttributeError, ZeroDivisionError):
                        pass

            # Update final cache
            if row_cache:
                sheet._fft_cached_formulas[row_num] = row_cache


def _calculate_formula_result(sheet, cell):
    """Calculate result for a single formula cell.

    Args:
        sheet: Openpyxl worksheet object
        cell: Cell containing formula

    Returns:
        Calculated result or None if calculation fails

    """
    formula = cell.value

    # Handle SUBTOTAL(9,range) formulas
    if "SUBTOTAL(9," in formula:
        return _calculate_subtotal_formula(sheet, formula)

    # Handle IFERROR formulas
    if formula.startswith("=IFERROR("):
        return _calculate_iferror_formula(sheet, cell, formula)

    return None


def _calculate_subtotal_formula(sheet, formula: str):
    """Calculate SUBTOTAL(9,range) formula result."""
    range_part = formula.split("SUBTOTAL(9,")[1].split(")")[0]
    col_letter = range_part.split(":")[0][0]
    start_row = int(range_part.split(":")[0][1:])
    end_row = int(range_part.split(":")[1][1:])

    col_idx = ord(col_letter) - ord("A") + 1

    total = 0
    for row in range(start_row, end_row + 1):
        data_cell = sheet.cell(row=row, column=col_idx)
        if data_cell.value and isinstance(data_cell.value, (int, float)):
            if not sheet.row_dimensions[row].hidden:
                total += data_cell.value

    return total if total > 0 else 0


def _calculate_iferror_formula(sheet, cell, formula: str):
    """Calculate IFERROR formula result."""
    # Extract the main expression from IFERROR(expression, fallback)
    inner_expr = formula[9:-1]  # Remove "=IFERROR(" and ")"

    # Find the comma that separates expression from fallback
    # Need to handle nested parentheses and functions
    paren_count = 0
    split_pos = -1
    for i, char in enumerate(inner_expr):
        if char == "(":
            paren_count += 1
        elif char == ")":
            paren_count -= 1
        elif char == "," and paren_count == 0:
            split_pos = i
            break

    if split_pos == -1:
        return None

    main_expr = inner_expr[:split_pos].strip()
    fallback = inner_expr[split_pos + 1 :].strip().replace('"', "")

    try:
        # Handle division expressions like (G14+H14)/SUM(G14:L14)
        if "/" in main_expr:
            div_pos = main_expr.rfind("/")  # Find last division operator
            numerator_part = main_expr[:div_pos].strip()
            denominator_part = main_expr[div_pos + 1 :].strip()

            # Calculate numerator and denominator
            numerator = _evaluate_expression(sheet, numerator_part)
            denominator = _evaluate_expression(sheet, denominator_part)

            if denominator == 0 or denominator is None:
                return fallback

            result = numerator / denominator
            return round(result, 4)  # Match Excel precision

    except (ValueError, ZeroDivisionError, TypeError):
        return fallback

    return fallback


def _evaluate_expression(sheet, expression: str):
    """Evaluate complex expressions incl. cell references, additions, and functions."""
    expression = expression.strip()

    # Handle parentheses by removing outer ones if they enclose the entire expression
    if expression.startswith("(") and expression.endswith(")"):
        paren_count = 0
        for i, char in enumerate(expression):
            if char == "(":
                paren_count += 1
            elif char == ")":
                paren_count -= 1
                if paren_count == 0 and i == len(expression) - 1:
                    # Outer parentheses enclose entire expression
                    expression = expression[1:-1]
                    break

    # Handle SUM function calls like SUM(G14:L14)
    if expression.startswith("SUM(") and expression.endswith(")"):
        range_part = expression[4:-1]  # Extract G14:L14
        return _evaluate_sum_range(sheet, range_part)

    # Handle addition expressions like G14+H14
    if "+" in expression:
        parts = [part.strip() for part in expression.split("+")]
        total = 0
        for part in parts:
            value = _get_cell_value(sheet, part)
            if isinstance(value, (int, float)):
                total += value
        return total

    # Single cell reference
    return _get_cell_value(sheet, expression)


def _get_cell_value(sheet, cell_ref: str):
    """Get the actual value from a cell, handling formulas by using cached results."""
    cell_ref = cell_ref.strip()
    cell = sheet[cell_ref]

    # If cell contains a formula, try to get cached result first
    if cell.data_type == "f" and hasattr(sheet, "_fft_cached_formulas"):
        for row_cache in sheet._fft_cached_formulas.values():
            if cell_ref in row_cache:
                return row_cache[cell_ref]

    # Otherwise return the raw value
    value = cell.value
    return value if isinstance(value, (int, float)) else 0


def _evaluate_sum_range(sheet, range_expr: str):
    """Evaluate SUM range like G14:L14."""
    if ":" not in range_expr:
        # Single cell
        return _get_cell_value(sheet, range_expr)

    start_ref, end_ref = range_expr.split(":")
    start_cell = sheet[start_ref]
    end_cell = sheet[end_ref]

    total = 0
    for row in range(start_cell.row, end_cell.row + 1):
        for col in range(start_cell.column, end_cell.column + 1):
            cell = sheet.cell(row=row, column=col)
            value = _get_cell_value(sheet, cell.coordinate)
            if isinstance(value, (int, float)):
                total += value

    return total


# %%
def format_percentage_columns(workbook: Workbook, service_type: str) -> None:
    """Apply percentage formatting (0%) to percentage columns.

    Args:
        workbook: Openpyxl Workbook object
        service_type: 'inpatient', 'ae', or 'ambulance'

    Returns:
        None (modifies workbook in place)

    Raises:
        KeyError: If service_type is not configured

    >>> from src.fft.writers import load_template, format_percentage_columns
    >>> wb = load_template('inpatient')
    >>> wb['ICB'].cell(row=15, column=5).value = 0.95
    >>> format_percentage_columns(wb, 'inpatient')
    >>> wb['ICB'].cell(row=15, column=5).number_format
    '0%'

    # Edge case: Missing sheet in workbook (should skip gracefully)
    >>> from src.fft.config import PERCENTAGE_COLUMN_CONFIG
    >>> original_config = PERCENTAGE_COLUMN_CONFIG['inpatient'].copy()
    >>> PERCENTAGE_COLUMN_CONFIG['inpatient']['NonExistentSheet'] = [5]
    >>> format_percentage_columns(wb, 'inpatient')  # Should not raise error
    >>> PERCENTAGE_COLUMN_CONFIG['inpatient'] = original_config

    # Edge case: Cell contains non-numeric value
    >>> wb['ICB'].cell(row=16, column=5).value = "text"
    >>> format_percentage_columns(wb, 'inpatient')  # Should still format other cells
    >>> wb['ICB'].cell(row=15, column=5).number_format  # Verify previous cell formatting
    '0%'

    # Error case: Unknown service type
    >>> format_percentage_columns(wb, 'unknown')
    Traceback (most recent call last):
        ...
    KeyError: "Unknown service type: 'unknown'"

    """
    if service_type not in PERCENTAGE_COLUMN_CONFIG:
        raise KeyError(f"Unknown service type: '{service_type}'")

    config = PERCENTAGE_COLUMN_CONFIG[service_type]
    data_start_row = TEMPLATE_CONFIG[service_type]["data_start_row"]

    for sheet_name, columns in config.items():
        if sheet_name not in workbook.sheetnames:
            continue

        sheet = workbook[sheet_name]

        for col_idx in columns:
            for row in range(data_start_row, sheet.max_row + 1):
                cell = sheet.cell(row=row, column=col_idx)
                if cell.value is not None and cell.value != "*":
                    cell.number_format = PERCENTAGE_NUMBER_FORMAT


# %%
def save_output(workbook: Workbook, service_type: str, fft_period: str) -> Path:
    """Save workbook with correct filename pattern to outputs directory.

    Saves as macro-enabled workbook (.xlsm) with naming pattern:
    FFT-{service}-data-{period}.xlsm (e.g., FFT-inpatient-data-Aug-24.xlsm)

    Caches all formula results before saving for validation purposes.

    Args:
        workbook: Openpyxl Workbook object to save
        service_type: 'inpatient', 'ae', or 'ambulance'
        fft_period: FFT period string (e.g., 'Aug-24')

    Returns:
        Path to the saved file

    Raises:
        KeyError: If service_type is not configured

    >>> from fft.writers import load_template, save_output
    >>> import tempfile
    >>> from pathlib import Path
    >>> import fft.writers as writers_module
    >>> wb = load_template('inpatient')
    >>> with tempfile.TemporaryDirectory() as temp_dir:
    ...     original_outputs_dir = writers_module.OUTPUTS_DIR
    ...     writers_module.OUTPUTS_DIR = Path(temp_dir) / "outputs"
    ...     output_path = save_output(wb, 'inpatient', 'Aug-24')
    ...     filename = output_path.name
    ...     file_exists = output_path.exists()
    ...     writers_module.OUTPUTS_DIR = original_outputs_dir
    ...     (filename, file_exists)
    ('FFT-inpatient-data-Aug-24.xlsm', True)

    # Edge case: FFT period with complex format
    >>> with tempfile.TemporaryDirectory() as temp_dir:
    ...     original_outputs_dir = writers_module.OUTPUTS_DIR
    ...     writers_module.OUTPUTS_DIR = Path(temp_dir) / "outputs"
    ...     output_path = save_output(wb, 'inpatient', 'Dec-2024')
    ...     filename = output_path.name
    ...     file_exists = output_path.exists()
    ...     writers_module.OUTPUTS_DIR = original_outputs_dir
    ...     (filename, file_exists)
    ('FFT-inpatient-data-Dec-2024.xlsm', True)

    # Edge case: Outputs directory creation when missing
    >>> with tempfile.TemporaryDirectory() as temp_dir:
    ...     original_outputs_dir = writers_module.OUTPUTS_DIR
    ...     writers_module.OUTPUTS_DIR = Path(temp_dir) / "outputs"
    ...     output_path = save_output(wb, 'inpatient', 'Sep-24')
    ...     file_exists = output_path.exists()
    ...     writers_module.OUTPUTS_DIR = original_outputs_dir
    ...     file_exists
    True

    # Error case: Unknown service type
    >>> save_output(wb, 'unknown', 'Aug-24')
    Traceback (most recent call last):
        ...
    KeyError: "Unknown service type: 'unknown'"

    """
    if service_type not in TEMPLATE_CONFIG:
        raise KeyError(f"Unknown service type: '{service_type}'")

    config = TEMPLATE_CONFIG[service_type]
    output_prefix = config["output_prefix"]

    filename = f"{output_prefix}-{fft_period}.xlsm"
    output_path = OUTPUTS_DIR / filename

    # Ensure outputs directory exists
    OUTPUTS_DIR.mkdir(parents=True, exist_ok=True)

    # Cache all formula results for validation
    _cache_all_formula_results(workbook)

    workbook.save(output_path)

    return output_path


# %%
def write_summary_sheet(
    workbook: Workbook,
    summary_data: dict,
    current_period: str,
    previous_period: str,
    service_type: str = "inpatient",
) -> None:
    """Write summary data to the Summary sheet in the template.

    Populates the Summary sheet with organisations submitting, responses,
    and percentages for service-specific provider types.

    Args:
        workbook: Openpyxl Workbook object
        summary_data: Dict from extract_summary_data()
        current_period: Current FFT period (e.g., 'Jul-25')
        previous_period: Previous FFT period (e.g., 'Jun-25')
        service_type: Service type ('inpatient', 'ae', etc.)

    Returns:
        None (modifies workbook in place)

    Raises:
        KeyError: If Summary sheet doesn't exist or service type not configured

    >>> from src.fft.writers import load_template, write_summary_sheet
    >>> wb = load_template('inpatient')
    >>> summary_data = {
    ...     'orgs_submitting': {'total': 150, 'nhs': 134, 'is': 19},
    ...     'responses_to_date': {'total': 25769386, 'nhs': 24003672, 'is': 1769240},
    ...     'responses_current': {'total': 202745, 'nhs': 186977, 'is': 15883},
    ...     'responses_previous': {'total': 213043, 'nhs': 195590, 'is': 17606},
    ...     'pct_positive_current': {'total': 0.95, 'nhs': 0.95, 'is': 0.99},
    ...     'pct_positive_previous': {'total': 0.95, 'nhs': 0.95, 'is': 0.99},
    ...     'pct_negative_current': {'total': 0.02, 'nhs': 0.03, 'is': 0.0},
    ...     'pct_negative_previous': {'total': 0.02, 'nhs': 0.03, 'is': 0.0},
    ... }
    >>> write_summary_sheet(wb, summary_data, 'Jul 25', 'Jun 25', 'inpatient')
    >>> wb['Summary'].cell(row=8, column=3).value
    150
    >>> wb['Summary'].cell(row=9, column=3).value
    134

    # Test AE service type
    >>> wb_ae = load_template('ae')
    >>> summary_data_ae = {
    ...     'orgs_submitting': {'total': 200, 'acute': 150, 'wics': 50},
    ...     'responses_to_date': {'total': 3000000, 'acute': 2500000, 'wics': 500000},
    ...     'responses_current': {'total': 300000, 'acute': 250000, 'wics': 50000},
    ...     'responses_previous': {'total': 290000, 'acute': 240000, 'wics': 48000},
    ...     'pct_positive_current': {'total': 0.95, 'acute': 0.94, 'wics': 0.98},
    ...     'pct_positive_previous': {'total': 0.94, 'acute': 0.93, 'wics': 0.97},
    ...     'pct_negative_current': {'total': 0.03, 'acute': 0.04, 'wics': 0.02},
    ...     'pct_negative_previous': {'total': 0.04, 'acute': 0.05, 'wics': 0.03},
    ... }
    >>> write_summary_sheet(wb_ae, summary_data_ae, 'Jul 25', 'Jun 25', 'ae')
    >>> wb_ae['Summary'].cell(row=5, column=3).value
    200
    >>> wb_ae['Summary'].cell(row=6, column=3).value
    150

    """
    if "Summary" not in workbook.sheetnames:
        raise KeyError("Sheet 'Summary' not found in workbook")

    if service_type not in SUMMARY_SHEET_CONFIG:
        raise KeyError(
            f"No summary sheet configuration for service type: '{service_type}'"
        )

    sheet = workbook["Summary"]
    config = SUMMARY_SHEET_CONFIG[service_type]

    # Get provider types from the summary data
    provider_types = list(summary_data["orgs_submitting"].keys())

    # Helper function to safely write to potentially merged cells
    def safe_write_cell(sheet, row, col, value):
        """Write to cell, handling merged cells by writing to top-left cell."""
        try:
            cell = sheet.cell(row=row, column=col)
            if hasattr(cell, 'coordinate') and str(type(cell)) == "<class 'openpyxl.worksheet.merge.MergedCell'>":
                # Find the top-left cell of the merged range
                for merged_range in sheet.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        top_left_cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                        top_left_cell.value = value
                        return
            cell.value = value
        except AttributeError:
            # If it's a MergedCell, find the top-left cell
            for merged_range in sheet.merged_cells.ranges:
                if (merged_range.min_row <= row <= merged_range.max_row and
                    merged_range.min_col <= col <= merged_range.max_col):
                    top_left_cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                    top_left_cell.value = value
                    return
            # If not found in any merged range, just try regular assignment
            sheet.cell(row=row, column=col).value = value
    # Write period headers
    period_row = config["period_row"]
    for col_idx in range(3, 11):  # Columns C to J
        if col_idx in [3, 4, 5, 7, 9]:  # Current period columns
            safe_write_cell(sheet, period_row, col_idx, current_period)
        elif col_idx in [6, 8, 10]:  # Previous period columns
            safe_write_cell(sheet, period_row, col_idx, previous_period)

    # Write data for each provider type
    for provider_key in provider_types:
        if provider_key not in config["rows"]:
            continue  # Skip if provider type not configured for this service

        row = config["rows"][provider_key]
        for data_key, col in config["cols"].items():
            if data_key in summary_data and provider_key in summary_data[data_key]:
                value = summary_data[data_key][provider_key]
                safe_write_cell(sheet, row, col, value)


# %%
def calculate_previous_period(current_period: str) -> str:
    """Calculate the previous FFT period from current period.

    Args:
        current_period: Current FFT period (e.g., 'Jul-25')

    Returns:
        Previous FFT period (e.g., 'Jun-25')

    >>> calculate_previous_period('Jul-25')
    'Jun-25'
    >>> calculate_previous_period('Jan-25')
    'Dec-24'
    >>> calculate_previous_period('Apr-24')
    'Mar-24'

    """
    # Create reverse mapping from abbreviation to month number
    abbrev_to_num = {
        abbrev: num
        for num, abbrev in enumerate(
            [
                "Jan",
                "Feb",
                "Mar",
                "Apr",
                "May",
                "Jun",
                "Jul",
                "Aug",
                "Sep",
                "Oct",
                "Nov",
                "Dec",
            ],
            1,
        )
    }
    num_to_abbrev = {num: abbrev for abbrev, num in abbrev_to_num.items()}

    # Parse current period
    month_abbrev, year = current_period.split("-")
    month_num = abbrev_to_num[month_abbrev]
    year_num = int(year)

    # Calculate previous month
    if month_num == 1:  # January -> December of previous year
        prev_month_num = 12
        prev_year_num = year_num - 1
    else:
        prev_month_num = month_num - 1
        prev_year_num = year_num

    # Convert back to period format
    prev_month_abbrev = num_to_abbrev[prev_month_num]
    prev_year_str = f"{prev_year_num:02d}"  # Format as 2-digit year

    return f"{prev_month_abbrev}-{prev_year_str}"


# %%
def populate_summary_sheet(
    workbook: Workbook, service_type: str, current_period: str
) -> None:
    """Populate Summary sheet with Collections Overview time series data.

    Loads Collections Overview data, extracts summary metrics,
    and writes to template Summary sheet.

    Args:
        workbook: Loaded template workbook
        service_type: 'inpatient', 'ae', or 'ambulance'
        current_period: Current FFT period (e.g., 'Jul-25')

    Returns:
        None (modifies workbook in place)

    Raises:
        FileNotFoundError: If Collections Overview file not found
        KeyError: If service type not supported or data missing
        ValueError: If periods not found in Collections Overview data

    """
    try:
        # Calculate previous period
        previous_period = calculate_previous_period(current_period)

        # Load Collections Overview time series data
        time_series_df = load_collections_overview()

        # Extract summary data for this service type and periods
        summary_data = extract_summary_data(
            time_series_df, service_type, current_period, previous_period
        )

        # Write to Summary sheet with service type
        write_summary_sheet(
            workbook, summary_data, current_period, previous_period, service_type
        )

    except FileNotFoundError as e:
        logger.warning(f"Collections Overview file not found: {e}")
        logger.warning("Skipping Summary sheet population")
    except (KeyError, ValueError) as e:
        logger.warning(f"Unable to populate Summary sheet: {e}")
        logger.warning("Skipping Summary sheet population")
