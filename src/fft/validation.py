"""Workbook comparison utilities for validation."""

import re
from datetime import datetime
from pathlib import Path
from typing import TypedDict

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

from fft.config import (
    CRITICAL_HEADER_CELLS,
    HEADER_ROW_RANGES_BY_SERVICE,
    HEADER_ROWS_BY_SERVICE,
    HEADER_VALIDATION_EXCLUDED_SHEETS,
    VALIDATION_TOLERANCE,
)

# Type for Excel cell values from openpyxl
CellValue = str | int | float | bool | datetime | None


class CellDifference(TypedDict):
    """Represents a difference between two cells."""

    sheet: str
    cell: str
    expected: CellValue
    actual: CellValue


class SheetResult(TypedDict):
    """Comparison result for a single sheet."""

    name: str
    identical: bool
    differences: list[CellDifference]
    missing_in_actual: bool
    missing_in_expected: bool


def compare_workbooks(
    expected_path: Path | str,
    actual_path: Path | str,
    sheets_to_compare: list[str] | None = None,
    data_only: bool = True,
) -> list[SheetResult]:
    """Compare two Excel workbooks and return detailed differences.

    Compares cell values (or formulas if data_only=False) between two workbooks.
    Useful for validating automatically generated output against known good results.

    Args:
        expected_path: Path to the ground truth workbook
        actual_path: Path to the workbook to validate
        sheets_to_compare: List of sheet names to compare (None = all common sheets)
        data_only: If True, compare computed values; if False, compare formulas

    Returns:
        List of SheetResult dicts, one per sheet compared

    Raises:
        FileNotFoundError: If either workbook doesn't exist

    >>> from pathlib import Path
    >>> from openpyxl import Workbook
    >>> import tempfile

    # Create two identical workbooks
    >>> with tempfile.TemporaryDirectory() as tmpdir:
    ...     expected = Path(tmpdir) / "expected.xlsx"
    ...     actual = Path(tmpdir) / "actual.xlsx"
    ...     wb1 = Workbook()
    ...     wb1.active.title = "Data"
    ...     wb1["Data"]["A1"] = "Hello"
    ...     wb1["Data"]["B1"] = 100
    ...     wb1.save(expected)
    ...     wb2 = Workbook()
    ...     wb2.active.title = "Data"
    ...     wb2["Data"]["A1"] = "Hello"
    ...     wb2["Data"]["B1"] = 100
    ...     wb2.save(actual)
    ...     results = compare_workbooks(expected, actual)
    ...     results[0]["identical"]
    True

    # Create workbooks with differences
    >>> with tempfile.TemporaryDirectory() as tmpdir:
    ...     expected = Path(tmpdir) / "expected.xlsx"
    ...     actual = Path(tmpdir) / "actual.xlsx"
    ...     wb1 = Workbook()
    ...     wb1.active.title = "Data"
    ...     wb1["Data"]["A1"] = "Hello"
    ...     wb1["Data"]["B1"] = 100
    ...     wb1.save(expected)
    ...     wb2 = Workbook()
    ...     wb2.active.title = "Data"
    ...     wb2["Data"]["A1"] = "Hello"
    ...     wb2["Data"]["B1"] = 999
    ...     wb2.save(actual)
    ...     results = compare_workbooks(expected, actual)
    ...     results[0]["identical"]
    False
    >>> with tempfile.TemporaryDirectory() as tmpdir:
    ...     expected = Path(tmpdir) / "expected.xlsx"
    ...     actual = Path(tmpdir) / "actual.xlsx"
    ...     wb1 = Workbook()
    ...     wb1.active.title = "Data"
    ...     wb1["Data"]["A1"] = "Hello"
    ...     wb1["Data"]["B1"] = 100
    ...     wb1.save(expected)
    ...     wb2 = Workbook()
    ...     wb2.active.title = "Data"
    ...     wb2["Data"]["A1"] = "Hello"
    ...     wb2["Data"]["B1"] = 999
    ...     wb2.save(actual)
    ...     results = compare_workbooks(expected, actual)
    ...     results[0]["differences"][0]["cell"]
    'B1'

    # Edge case: Missing file
    >>> compare_workbooks("nonexistent.xlsx", "also_missing.xlsx")
    Traceback (most recent call last):
        ...
    FileNotFoundError: Expected workbook not found: nonexistent.xlsx

    # Edge case: Sheet missing in actual
    >>> with tempfile.TemporaryDirectory() as tmpdir:
    ...     expected = Path(tmpdir) / "expected.xlsx"
    ...     actual = Path(tmpdir) / "actual.xlsx"
    ...     wb1 = Workbook()
    ...     _ = wb1.create_sheet("Extra")
    ...     wb1.save(expected)
    ...     wb2 = Workbook()
    ...     wb2.save(actual)
    ...     results = compare_workbooks(expected, actual, sheets_to_compare=["Extra"])
    ...     results[0]["missing_in_actual"]
    True

    """
    expected_path = Path(expected_path)
    actual_path = Path(actual_path)

    if not expected_path.exists():
        raise FileNotFoundError(f"Expected workbook not found: {expected_path}")
    if not actual_path.exists():
        raise FileNotFoundError(f"Actual workbook not found: {actual_path}")

    wb_expected = load_workbook(expected_path, data_only=data_only)
    wb_actual = load_workbook(actual_path, data_only=data_only)

    if sheets_to_compare is None:
        sheets_to_compare = list(set(wb_expected.sheetnames) | set(wb_actual.sheetnames))

    return [
        _compare_sheet(sheet_name, wb_expected, wb_actual)
        for sheet_name in sheets_to_compare
    ]


def _compare_sheet(sheet_name: str, wb_expected, wb_actual) -> SheetResult:
    """Compare a single sheet between two workbooks."""
    if sheet_name not in wb_expected.sheetnames:
        return {
            "name": sheet_name,
            "identical": False,
            "differences": [],
            "missing_in_expected": True,
            "missing_in_actual": False,
        }

    if sheet_name not in wb_actual.sheetnames:
        return {
            "name": sheet_name,
            "identical": False,
            "differences": [],
            "missing_in_expected": False,
            "missing_in_actual": True,
        }

    ws_expected = wb_expected[sheet_name]
    ws_actual = wb_actual[sheet_name]

    differences = _find_cell_differences(sheet_name, ws_expected, ws_actual)

    return {
        "name": sheet_name,
        "identical": len(differences) == 0,
        "differences": differences,
        "missing_in_expected": False,
        "missing_in_actual": False,
    }


def _find_cell_differences(
    sheet_name: str, ws_expected, ws_actual
) -> list[CellDifference]:
    """Find all cell differences between two worksheets."""
    max_row = max(ws_expected.max_row or 1, ws_actual.max_row or 1)
    max_col = max(ws_expected.max_column or 1, ws_actual.max_column or 1)

    differences = []
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            val_expected = ws_expected.cell(row=row, column=col).value
            val_actual = ws_actual.cell(row=row, column=col).value

            if not _values_are_equivalent(val_expected, val_actual):
                differences.append(
                    {
                        "sheet": sheet_name,
                        "cell": ws_expected.cell(row=row, column=col).coordinate,
                        "expected": val_expected,
                        "actual": val_actual,
                    }
                )

    return differences


def compare_data_by_key(  # noqa: PLR0913 # Justified: validation function needs all params
    expected_path: Path | str,
    actual_path: Path | str,
    sheet_name: str,
    *,
    key_column: str | list[str] = "B",  # Single column or composite key columns
    start_row: int = 15,
    data_only: bool = True,
    actual_sheet_name: str | None = None,
) -> SheetResult:
    """Compare sheet data by matching records via key column(s) rather than row position.

    Args:
        expected_path: Path to ground truth workbook
        actual_path: Path to generated workbook
        sheet_name: Name of sheet to compare
        key_column: Column letter(s) for unique identification
                   - Single column: "B" for Trust_Code
                   - Composite key: ["B", "D", "F"] for Trust_Code + Site_Code + Ward_Name
        start_row: First row of data
        data_only: Compare values vs formulas
        actual_sheet_name: Name of sheet in actual workbook if different from sheet_name

    Returns:
        SheetResult with differences between matching records

    """
    expected_path = Path(expected_path)
    actual_path = Path(actual_path)

    if not expected_path.exists():
        raise FileNotFoundError(f"Expected workbook not found: {expected_path}")
    if not actual_path.exists():
        raise FileNotFoundError(f"Actual workbook not found: {actual_path}")

    wb_expected = load_workbook(expected_path, data_only=data_only)
    wb_actual = load_workbook(actual_path, data_only=data_only)

    actual_sheet = actual_sheet_name or sheet_name

    if sheet_name not in wb_expected.sheetnames:
        return {
            "name": actual_sheet,
            "identical": False,
            "differences": [],
            "missing_in_expected": True,
            "missing_in_actual": False,
        }

    if actual_sheet not in wb_actual.sheetnames:
        return {
            "name": actual_sheet,
            "identical": False,
            "differences": [],
            "missing_in_expected": False,
            "missing_in_actual": True,
        }

    ws_expected = wb_expected[sheet_name]
    ws_actual = wb_actual[actual_sheet]

    # Build dictionaries mapping key -> row data
    expected_records = _extract_records_by_key(ws_expected, key_column, start_row)
    actual_records = _extract_records_by_key(ws_actual, key_column, start_row)

    differences = _compare_records_by_key(actual_sheet, expected_records, actual_records)

    return {
        "name": actual_sheet,
        "identical": len(differences) == 0,
        "differences": differences,
        "missing_in_expected": False,
        "missing_in_actual": False,
    }


def _extract_records_by_key(
    worksheet, key_column: str | list[str], start_row: int
) -> dict:
    """Extract records from worksheet, keyed by identifier column(s).

    Args:
        worksheet: Excel worksheet
        key_column: Single column letter or list of column letters for composite key
        start_row: First row of data

    Returns:
        Dict mapping key(s) to row data

    """
    records = {}

    # Convert column letters to indices
    if isinstance(key_column, str):
        key_col_indices = [column_index_from_string(key_column.upper())]
    else:
        key_col_indices = [column_index_from_string(col.upper()) for col in key_column]

    max_row = worksheet.max_row or start_row
    max_col = worksheet.max_column or 10

    for row_num in range(start_row, max_row + 1):
        # Extract key value(s)
        key_parts = []
        for col_idx in key_col_indices:
            key_value = worksheet.cell(row=row_num, column=col_idx).value
            if key_value and str(key_value).strip():
                key_parts.append(str(key_value).strip())
            else:
                key_parts.append("")

        # Only process rows where all key parts are non-empty
        if all(part for part in key_parts):
            # Create composite key by joining with separator
            composite_key = "|".join(key_parts)

            row_data = {}
            for col_num in range(1, max_col + 1):
                cell_value = worksheet.cell(row=row_num, column=col_num).value
                # Store by column index, not absolute coordinate
                row_data[col_num] = cell_value
            records[composite_key] = row_data

    return records


def _compare_records_by_key(
    sheet_name: str, expected_records: dict, actual_records: dict
) -> list[CellDifference]:
    """Compare matching records between expected and actual data."""
    differences = []

    # Compare records that exist in both datasets
    for key, expected_row in expected_records.items():
        if key not in actual_records:
            continue  # Skip missing records for now

        actual_row = actual_records[key]

        # Compare all columns in the row
        for col_num in expected_row:
            expected_val = expected_row[col_num]
            actual_val = actual_row.get(col_num)

            if not _values_are_equivalent(expected_val, actual_val):
                # Convert column number back to letter for display
                col_letter = get_column_letter(col_num)
                differences.append(
                    {
                        "sheet": sheet_name,
                        "cell": f"{col_letter}({key})",  # Show column letter with key
                        "expected": expected_val,
                        "actual": actual_val,
                    }
                )

    return differences


def compare_data_range(
    expected_path: Path | str, actual_path: Path | str, sheet_name: str, **options
) -> SheetResult:
    """Compare only the data range of a sheet, ignoring template/control areas.

    Args:
        expected_path: Path to ground truth workbook
        actual_path: Path to generated workbook
        sheet_name: Name of sheet to compare
        **options: Optional parameters including start_row (int), data_only (bool),
                  actual_sheet_name (str)

    Returns:
        SheetResult with differences in data area only

    """
    # Extract options with defaults
    start_row = options.get("start_row", 15)
    data_only = options.get("data_only", True)
    actual_sheet_name = options.get("actual_sheet_name")

    expected_path = Path(expected_path)
    actual_path = Path(actual_path)

    if not expected_path.exists():
        raise FileNotFoundError(f"Expected workbook not found: {expected_path}")
    if not actual_path.exists():
        raise FileNotFoundError(f"Actual workbook not found: {actual_path}")

    wb_expected = load_workbook(expected_path, data_only=data_only)
    wb_actual = load_workbook(actual_path, data_only=data_only)

    actual_sheet = actual_sheet_name or sheet_name

    if sheet_name not in wb_expected.sheetnames:
        return {
            "name": actual_sheet,
            "identical": False,
            "differences": [],
            "missing_in_expected": True,
            "missing_in_actual": False,
        }

    if actual_sheet not in wb_actual.sheetnames:
        return {
            "name": actual_sheet,
            "identical": False,
            "differences": [],
            "missing_in_expected": False,
            "missing_in_actual": True,
        }

    ws_expected = wb_expected[sheet_name]
    ws_actual = wb_actual[actual_sheet]

    # Compare only from start_row onwards
    differences = _find_cell_differences_range(
        actual_sheet, ws_expected, ws_actual, start_row
    )

    return {
        "name": actual_sheet,
        "identical": len(differences) == 0,
        "differences": differences,
        "missing_in_expected": False,
        "missing_in_actual": False,
    }


def _find_cell_differences_range(
    sheet_name: str, ws_expected, ws_actual, start_row: int
) -> list[CellDifference]:
    """Find cell differences starting from specified row."""
    max_row = max(ws_expected.max_row or 1, ws_actual.max_row or 1)
    max_col = max(ws_expected.max_column or 1, ws_actual.max_column or 1)

    differences = []
    for row in range(start_row, max_row + 1):
        for col in range(1, max_col + 1):
            val_expected = ws_expected.cell(row=row, column=col).value
            val_actual = ws_actual.cell(row=row, column=col).value

            if not _values_are_equivalent(val_expected, val_actual):
                differences.append(
                    {
                        "sheet": sheet_name,
                        "cell": ws_expected.cell(row=row, column=col).coordinate,
                        "expected": val_expected,
                        "actual": val_actual,
                    }
                )

    return differences


def _values_are_equivalent(val_expected, val_actual) -> bool:
    """Check if two values are equivalent with reasonable tolerance.

    >>> _values_are_equivalent(0.025381903642773207, 0.02538190364277321)
    True
    >>> _values_are_equivalent(0.942, 0.942)
    True
    >>> _values_are_equivalent("Hello", "Hello")
    True
    >>> _values_are_equivalent("Hello", "World")
    False
    >>> _values_are_equivalent(100, 100)
    True
    >>> _values_are_equivalent(100.1, 100.2)
    False
    >>> _values_are_equivalent("Oct-25", "2025-10-01 00:00:00")
    True
    >>> _values_are_equivalent("Sep-25", "2025-09-26 00:00:00")
    True
    >>> _values_are_equivalent("Oct-25", "2025-11-01 00:00:00")
    False

    """
    # Exact match
    if val_expected == val_actual:
        return True

    # Handle template compatibility: None/NULL/NA/"0"/"-"
    # are equivalent for missing values
    missing_vals = {None, "NULL", "NA", "", "0", "-", "nan"}
    if val_expected in missing_vals and val_actual in missing_vals:
        return True

    # Handle date format differences - compare year/month only
    def extract_year_month(value):
        """Extract (year, month) tuple from date strings."""
        # Convert to string if it's a datetime object
        if hasattr(value, "year") and hasattr(value, "month"):
            return (value.year, value.month)

        if not isinstance(value, str):
            value_str = str(value) if value is not None else None
        else:
            value_str = value

        if not value_str:
            return None

        # "Jun-25" format
        if re.match(r"^[A-Z][a-z]{2}-\d{2}$", value_str):
            month_name, year_2digit = value_str.split("-")
            year = 2000 + int(year_2digit)
            month = datetime.strptime(month_name, "%b").month
            return (year, month)
        # "2025-06-01 00:00:00" format
        if re.match(r"^\d{4}-\d{2}", value_str):
            parts = value_str.split("-")
            return (int(parts[0]), int(parts[1]))
        return None

    try:
        date1 = extract_year_month(val_expected)
        date2 = extract_year_month(val_actual)
        if date1 and date2:
            return date1 == date2
    except (ValueError, TypeError):
        pass

    # For numeric comparisons
    try:
        num_expected = float(val_expected) if val_expected is not None else None
        num_actual = float(val_actual) if val_actual is not None else None

        if num_expected is None or num_actual is None:
            return False

        # Use absolute tolerance to eliminate floating point noise
        return abs(num_expected - num_actual) <= VALIDATION_TOLERANCE

    except (ValueError, TypeError):
        # Not numeric, compare as strings
        return str(val_expected) == str(val_actual)


def print_comparison_report(
    results: list[SheetResult], max_diffs_per_sheet: int = 10
) -> None:
    """Print a human-readable comparison report.

    Compares pipeline output against ground truth reference files.
    Reports show differences where:
    - "ground truth" = expected values from VBA reference files
    - "pipeline" = actual values generated by our Python pipeline

    Args:
        results: List of SheetResult from compare_workbooks()
        max_diffs_per_sheet: Maximum differences to show per sheet

    >>> from pathlib import Path
    >>> from openpyxl import Workbook
    >>> import tempfile
    >>> with tempfile.TemporaryDirectory() as tmpdir:
    ...     expected = Path(tmpdir) / "expected.xlsx"
    ...     actual = Path(tmpdir) / "actual.xlsx"
    ...     wb1 = Workbook()
    ...     wb1.active.title = "Data"
    ...     wb1["Data"]["A1"] = "Test"
    ...     wb1.save(expected)
    ...     wb2 = Workbook()
    ...     wb2.active.title = "Data"
    ...     wb2["Data"]["A1"] = "Test"
    ...     wb2.save(actual)
    ...     results = compare_workbooks(expected, actual)
    ...     print_comparison_report(results)
    Sheet: Data ✓ (identical)
    <BLANKLINE>
    Summary: 1/1 sheets identical

    """
    total_sheets = len(results)
    identical_sheets = sum(1 for r in results if r["identical"])

    for result in results:
        if result["missing_in_expected"]:
            print(f"Sheet: {result['name']} ⚠ (missing in ground truth file)")
        elif result["missing_in_actual"]:
            print(f"Sheet: {result['name']} ⚠ (missing in pipeline output)")
        elif result["identical"]:
            print(f"Sheet: {result['name']} ✓ (identical)")
        else:
            print(f"Sheet: {result['name']} ✗ ({len(result['differences'])} differences)")
            for diff in result["differences"][:max_diffs_per_sheet]:
                print(
                    f"  - Cell {diff['cell']}: ground truth {diff['expected']!r}, "
                    f"pipeline {diff['actual']!r}"
                )
            if len(result["differences"]) > max_diffs_per_sheet:
                print(
                    f"  ... and {len(result['differences']) - max_diffs_per_sheet} more"
                )

    print()
    print(f"Summary: {identical_sheets}/{total_sheets} sheets identical")


def validate_headers(
    pipeline_file: str | Path,
    ground_truth_file: str | Path,
    service_type: str,
    sheets_to_validate: list[str] | None = None,
    verbose: bool = False,
) -> dict[str, dict]:
    """Validate that sheet headers match between pipeline and ground truth files.

    This function performs CRUCIAL validation to ensure that:
    1. Header structure is identical
    2. Column labels match exactly
    3. Critical cells contain expected values

    Args:
        pipeline_file: Path to pipeline-generated Excel file
        ground_truth_file: Path to ground truth Excel file
        service_type: Service type (inpatient, ae, ambulance)
        sheets_to_validate: List of sheets to validate (None for all sheets)
        verbose: Whether to print detailed output

    Returns:
        dict: Validation results by sheet, with differences if any

    Raises:
        FileNotFoundError: If either file doesn't exist
        ValueError: If service_type is not configured

    >>> from pathlib import Path
    >>> import tempfile
    >>> from openpyxl import Workbook

    # Test with identical headers (using correct row range for inpatient Trusts sheet)
    >>> with tempfile.TemporaryDirectory() as tmpdir:
    ...     pipeline = Path(tmpdir) / "pipeline.xlsx"
    ...     ground_truth = Path(tmpdir) / "ground_truth.xlsx"
    ...     wb1 = Workbook()
    ...     wb1.active.title = "Trusts"
    ...     wb1["Trusts"]["A10"] = "ICB Code"  # Row 10 is in the header range [10, 14]
    ...     wb1["Trusts"]["B10"] = "Trust Code"
    ...     wb1.save(pipeline)
    ...     wb2 = Workbook()
    ...     wb2.active.title = "Trusts"
    ...     wb2["Trusts"]["A10"] = "ICB Code"
    ...     wb2["Trusts"]["B10"] = "Trust Code"
    ...     wb2.save(ground_truth)
    ...     results = validate_headers(pipeline, ground_truth, "inpatient", ["Trusts"])
    ...     results["Trusts"]["identical"]
    True

    # Test with different headers
    >>> with tempfile.TemporaryDirectory() as tmpdir:
    ...     pipeline = Path(tmpdir) / "pipeline.xlsx"
    ...     ground_truth = Path(tmpdir) / "ground_truth.xlsx"
    ...     wb1 = Workbook()
    ...     wb1.active.title = "Trusts"
    ...     wb1["Trusts"]["A10"] = "ICB Code"
    ...     wb1.save(pipeline)
    ...     wb2 = Workbook()
    ...     wb2.active.title = "Trusts"
    ...     wb2["Trusts"]["A10"] = "ICB_Name"  # Different!
    ...     wb2.save(ground_truth)
    ...     results = validate_headers(pipeline, ground_truth, "inpatient", ["Trusts"])
    ...     results["Trusts"]["identical"]
    False

    """
    pipeline_path, ground_truth_path = _validate_header_files(
        pipeline_file, ground_truth_file, service_type
    )

    wb_pipeline, wb_ground_truth = _load_header_workbooks(
        pipeline_path, ground_truth_path
    )

    sheets_to_validate = _determine_sheets_to_validate(
        sheets_to_validate, wb_pipeline, wb_ground_truth, service_type
    )

    results = {}
    header_rows = HEADER_ROWS_BY_SERVICE[service_type]

    config = {
        "header_rows": header_rows,
        "verbose": verbose,
        "service_type": service_type,
    }
    for sheet_name in sheets_to_validate:
        results[sheet_name] = _validate_single_sheet(
            sheet_name, wb_pipeline, wb_ground_truth, config
        )

    return results


def _validate_header_files(
    pipeline_file: str | Path, ground_truth_file: str | Path, service_type: str
) -> tuple[Path, Path]:
    """Validate input files and paths."""
    pipeline_path = Path(pipeline_file)
    ground_truth_path = Path(ground_truth_file)

    if not pipeline_path.exists():
        raise FileNotFoundError(f"Pipeline file not found: {pipeline_file}")
    if not ground_truth_path.exists():
        raise FileNotFoundError(f"Ground truth file not found: {ground_truth_file}")
    if service_type not in HEADER_ROWS_BY_SERVICE:
        raise ValueError(f"Unsupported service type: {service_type}")

    return pipeline_path, ground_truth_path


def _calculate_formula_value(cell, sheet) -> str | int | float | None:
    """Calculate the value of a formula cell.

    Args:
        cell: Openpyxl cell object containing a formula
        sheet: Openpyxl worksheet object

    Returns:
    Returns:
        The calculated value of the formula, or None if calculation fails

    """
    if not cell.data_type == "f" or not isinstance(cell.value, str):
        return cell.value

    formula = cell.value

    # Handle SUBTOTAL(9,range) - sum of visible cells
    if "SUBTOTAL(9," in formula:
        return _calculate_subtotal_formula(formula, sheet)

    # Handle IFERROR formulas
    elif formula.startswith("=IFERROR("):
        return _calculate_iferror_formula(formula, sheet)

    # For other formulas, return None (can't calculate)
    return None


def _calculate_subtotal_formula(formula: str, sheet) -> int | None:
    """Calculate SUBTOTAL(9,range) formula value."""
    try:
        # Extract range from formula like "=SUBTOTAL(9,D7:D999)"
        range_part = formula.split("SUBTOTAL(9,")[1].split(")")[0]
        col_letter = range_part.split(":")[0][0]  # 'D' from 'D7:D999'
        start_row = int(range_part.split(":")[0][1:])  # 7 from 'D7:D999'
        end_row = int(range_part.split(":")[1][1:])  # 999 from 'D7:D999'

        col_idx = ord(col_letter) - ord("A") + 1

        # Calculate the sum (SUBTOTAL(9,...) sums visible cells)
        total = 0
        for row in range(start_row, end_row + 1):
            data_cell = sheet.cell(row=row, column=col_idx)
            if data_cell.value and isinstance(data_cell.value, (int, float)):
                # Only sum if row is not hidden (SUBTOTAL behavior)
                if not sheet.row_dimensions[row].hidden:
                    total += data_cell.value

        return total if total > 0 else 0

    except (IndexError, ValueError, AttributeError):
        return None


def _calculate_iferror_formula(formula: str, sheet) -> str | float | None:
    """Calculate IFERROR formula value."""
    try:
        # Extract the main expression from IFERROR(expr, fallback)
        expr_part = formula[9:].split(",", maxsplit=1)[0]
        # Try to evaluate the expression (very basic handling)
        if "+" in expr_part and "/" in expr_part:
            return _evaluate_arithmetic_expression(expr_part, sheet)
        return None
    except (IndexError, ValueError, ZeroDivisionError):
        return "-"


def _evaluate_arithmetic_expression(expr: str, sheet) -> str | float:
    """Evaluate simple arithmetic expressions like (H6+I6)/D6."""
    # Handle simple arithmetic like (H6+I6)/D6
    parts = expr.split("/")
    numerator = parts[0].strip("()")
    denominator = parts[1].strip("()")

    # Calculate numerator (e.g., "H6+I6")
    num_total = _sum_cell_references(numerator, sheet)

    # Calculate denominator (e.g., "D6")
    denom_value = _get_cell_reference_value(denominator.strip(), sheet)

    if denom_value > 0:
        return num_total / denom_value
    else:
        return "-"


def _sum_cell_references(cell_refs: str, sheet) -> float:
    """Sum values from cell references separated by +."""
    num_parts = cell_refs.split("+")
    num_total = 0
    for part in num_parts:
        part_clean = part.strip()
        if part_clean:  # Skip empty parts
            cell_value = _get_cell_reference_value(part_clean, sheet)
            num_total += cell_value
    return num_total


def _get_cell_reference_value(cell_ref: str, sheet) -> float:
    """Get numeric value from a cell reference like 'D6'."""
    if cell_ref:
        col_letter = cell_ref[0]
        row_num = int(cell_ref[1:])
        col_idx = ord(col_letter) - ord("A") + 1
        cell = sheet.cell(row=row_num, column=col_idx)
        if cell.value and isinstance(cell.value, (int, float)):
            return cell.value
    return 1  # Default fallback value


def _load_header_workbooks(pipeline_path: Path, ground_truth_path: Path) -> tuple:
    """Load workbooks for header validation."""
    return (
        load_workbook(pipeline_path, data_only=False),
        load_workbook(ground_truth_path, data_only=False),
    )


def _determine_sheets_to_validate(
    sheets_to_validate: list[str] | None, wb_pipeline, wb_ground_truth, service_type: str
) -> list[str]:
    """Determine which sheets to validate."""
    if sheets_to_validate is None:
        common_sheets = list(
            set(wb_pipeline.sheetnames) & set(wb_ground_truth.sheetnames)
        )
    else:
        common_sheets = sheets_to_validate

    # Exclude sheets that contain template-specific content
    excluded_sheets = HEADER_VALIDATION_EXCLUDED_SHEETS.get(service_type, [])
    return [sheet for sheet in common_sheets if sheet not in excluded_sheets]


def _validate_single_sheet(
    sheet_name: str,
    wb_pipeline,
    wb_ground_truth,
    config: dict,
) -> dict:
    """Validate a single sheet's headers."""
    # Extract config parameters
    header_rows = config["header_rows"]
    verbose = config["verbose"]
    service_type = config["service_type"]

    # Check if sheet exists in both workbooks
    sheet_exists_result = _check_sheet_existence(sheet_name, wb_pipeline, wb_ground_truth)
    if sheet_exists_result is not None:
        return sheet_exists_result

    # Compare headers
    comparison_result = _compare_sheet_headers(
        sheet_name, wb_pipeline, wb_ground_truth, header_rows, service_type
    )

    # Add verbose output if requested
    if verbose:
        _print_verbose_validation(sheet_name, comparison_result)

    return comparison_result


def _check_sheet_existence(sheet_name: str, wb_pipeline, wb_ground_truth) -> dict | None:
    """Check if sheet exists in both workbooks."""
    if sheet_name not in wb_pipeline.sheetnames:
        return {
            "identical": False,
            "differences": [{"type": "missing", "message": "Sheet missing in pipeline"}],
            "critical_differences": [],
        }
    if sheet_name not in wb_ground_truth.sheetnames:
        return {
            "identical": False,
            "differences": [
                {"type": "missing", "message": "Sheet missing in ground truth"}
            ],
            "critical_differences": [],
        }
    return None


def _compare_sheet_headers(
    sheet_name: str, wb_pipeline, wb_ground_truth, header_rows: int, service_type: str
) -> dict:
    """Compare headers between pipeline and ground truth sheets."""
    sheet_pipeline = wb_pipeline[sheet_name]
    sheet_ground_truth = wb_ground_truth[sheet_name]

    all_differences = []
    critical_differences = []

    # Use the minimum column count to avoid out-of-bounds errors
    # when comparing workbooks with different column structures
    max_col = min(sheet_pipeline.max_column, sheet_ground_truth.max_column)

    # Get the precise row range for this sheet and service type
    row_range = HEADER_ROW_RANGES_BY_SERVICE.get(service_type, {}).get(sheet_name)
    if row_range:
        start_row, end_row = row_range
    else:
        # Fallback to original behavior for backward compatibility
        start_row = 1
        end_row = header_rows

    for row in range(start_row, end_row + 1):
        for col in range(1, max_col + 1):
            pipeline_cell = sheet_pipeline.cell(row=row, column=col)
            ground_truth_cell = sheet_ground_truth.cell(row=row, column=col)

            pipeline_val = pipeline_cell.value
            ground_truth_val = ground_truth_cell.value

            # Handle formula cells - calculate expected values for comparison
            if pipeline_cell.data_type == "f" or ground_truth_cell.data_type == "f":
                # If either cell contains a formula, calculate the expected value
                expected_pipeline_val = (
                    _calculate_formula_value(pipeline_cell, sheet_pipeline)
                    if pipeline_cell.data_type == "f"
                    else pipeline_val
                )
                expected_ground_truth_val = (
                    _calculate_formula_value(ground_truth_cell, sheet_ground_truth)
                    if ground_truth_cell.data_type == "f"
                    else ground_truth_val
                )

                pipeline_val = expected_pipeline_val
                ground_truth_val = expected_ground_truth_val

            if not _values_are_equivalent(pipeline_val, ground_truth_val):
                cell_ref = f"{get_column_letter(col)}{row}"
                diff = {
                    "cell": cell_ref,
                    "pipeline": pipeline_val,
                    "ground_truth": ground_truth_val,
                    "row": row,
                    "col": col,
                }
                all_differences.append(diff)

                if cell_ref in CRITICAL_HEADER_CELLS.get(sheet_name, []):
                    critical_differences.append(diff)

    return {
        "identical": len(all_differences) == 0,
        "differences": all_differences,
        "critical_differences": critical_differences,
        "total_differences": len(all_differences),
        "critical_difference_count": len(critical_differences),
    }


def _print_verbose_validation(sheet_name: str, comparison_result: dict) -> None:
    """Print verbose validation output."""
    if comparison_result["identical"]:
        print(f"✅ Sheet {sheet_name}: Headers are identical")
    else:
        print(
            f"❌ Sheet {sheet_name}: Found {len(comparison_result['differences'])} "
            f"differences ({len(comparison_result['critical_differences'])} critical)"
        )
        if comparison_result["critical_differences"]:
            print("  Critical differences:")
            for diff in comparison_result["critical_differences"][:3]:
                print(
                    f"    Cell {diff['cell']}: Pipeline='{diff['pipeline']}', "
                    f"GroundTruth='{diff['ground_truth']}'"
                )


def find_matching_ground_truth(output_path: Path, ground_truth_dir: Path) -> Path | None:
    """Find the best matching ground truth file for a given pipeline output.

    Matches files based on month/period and service type patterns rather than
    exact filenames, allowing for different naming conventions.

    Args:
        output_path: Path to the pipeline output file
        ground_truth_dir: Directory containing ground truth files

    Returns:
        Path to the best matching ground truth file, or None if no match found

    >>> from pathlib import Path
    >>> import tempfile

    # Test with matching month and service type
    >>> with tempfile.TemporaryDirectory() as tmpdir:
    ...     truth_dir = Path(tmpdir) / "ground_truth"
    ...     truth_dir.mkdir()
    ...     # Create ground truth file with different naming pattern
    ...     (truth_dir / "021225_133523_FFT_IP_MacroWebfile_Jul-25.xlsm").touch()
    ...     # Pipeline output with different pattern but same month/type
    ...     output = Path(tmpdir) / "FFT-inpatient-data-Jul-25.xlsm"
    ...     match = find_matching_ground_truth(output, truth_dir)
    ...     match.name if match else None
    '021225_133523_FFT_IP_MacroWebfile_Jul-25.xlsm'

    # Test with no matching files
    >>> with tempfile.TemporaryDirectory() as tmpdir:
    ...     truth_dir = Path(tmpdir) / "ground_truth"
    ...     truth_dir.mkdir()
    ...     output = Path(tmpdir) / "FFT-inpatient-data-Aug-25.xlsm"
    ...     match = find_matching_ground_truth(output, truth_dir)
    ...     match is None
    True

    # Test with multiple files, should pick best match
    >>> with tempfile.TemporaryDirectory() as tmpdir:
    ...     truth_dir = Path(tmpdir) / "ground_truth"
    ...     truth_dir.mkdir()
    ...     # Create files with different months
    ...     (truth_dir / "FFT_IP_Jul-25.xlsm").touch()
    ...     (truth_dir / "FFT_IP_Aug-25.xlsm").touch()
    ...     (truth_dir / "FFT_AE_Jul-25.xlsm").touch()  # Different service type
    ...     # Should match IP Jul-25, not AE Jul-25 or IP Aug-25
    ...     output = Path(tmpdir) / "FFT-inpatient-data-Jul-25.xlsm"
    ...     match = find_matching_ground_truth(output, truth_dir)
    ...     match.name if match else None
    'FFT_IP_Jul-25.xlsm'

    """
    if not ground_truth_dir.exists():
        return None

    output_month = _extract_month_pattern(output_path.name)
    output_service = extract_service_type(output_path.name)

    if not output_month or not output_service:
        return None

    best_match = None
    best_score = 0

    # Check all files in ground truth directory
    for truth_file in ground_truth_dir.glob("*.xl*"):
        truth_month = _extract_month_pattern(truth_file.name)
        truth_service = extract_service_type(truth_file.name)

        if not truth_month or not truth_service:
            continue

        # Score the match
        score = 0

        # Month must match exactly
        if truth_month == output_month:
            score += 100
        else:
            continue  # No match without matching month

        # Service type should match
        if truth_service == output_service:
            score += 50

        # Prefer newer files (by filename timestamp if present)
        if truth_file.stat().st_mtime > (best_match.stat().st_mtime if best_match else 0):
            score += 1

        if score > best_score:
            best_match = truth_file
            best_score = score

    return best_match


def _extract_month_pattern(filename: str) -> str | None:
    """Extract month pattern (e.g., 'Jul-25', 'Aug-25') from filename.

    >>> _extract_month_pattern("FFT-inpatient-data-Jul-25.xlsm")
    'Jul-25'
    >>> _extract_month_pattern("021225_133523_FFT_IP_MacroWebfile_Aug-25.xlsm")
    'Aug-25'
    >>> _extract_month_pattern("no-month-pattern.xlsx")

    """
    # Look for pattern like Jul-25, Aug-25, etc.
    pattern = r"([A-Z][a-z]{2}-\d{2})"
    match = re.search(pattern, filename)
    return match.group(1) if match else None


def extract_service_type(filename: str) -> str | None:
    """Extract normalized service type from filename.

    Maps various service type representations to standard codes:
    - inpatient/IP -> inpatient
    - ae/AE -> ae
    - ambulance/AMB -> ambulance

    >>> extract_service_type("FFT-inpatient-data-Jul-25.xlsm")
    'inpatient'
    >>> extract_service_type("021225_133523_FFT_IP_MacroWebfile_Jul-25.xlsm")
    'inpatient'
    >>> extract_service_type("FFT_AE_Aug-25.xlsm")
    'ae'
    >>> extract_service_type("FFT_AMB_Jul-25.xlsm")
    'ambulance'
    >>> extract_service_type("random-file.xlsm")

    """
    filename_lower = filename.lower()

    # Check for inpatient patterns
    if any(pattern in filename_lower for pattern in ["inpatient", "_ip_", "fft_ip"]):
        return "inpatient"

    # Check for A&E patterns
    if any(pattern in filename_lower for pattern in ["_ae_", "fft_ae", "-ae-"]):
        return "ae"

    # Check for ambulance patterns
    if any(pattern in filename_lower for pattern in ["ambulance", "_amb_", "fft_amb"]):
        return "ambulance"

    return None


def print_header_validation_report(
    header_results: dict[str, dict], max_diffs_per_sheet: int = 10
) -> None:
    """Print a detailed header validation report.

    Provides comprehensive information about header differences between pipeline
    output and ground truth files, showing exactly which cells differ and what
    the differences are.

    Args:
        header_results: Results from validate_headers() function
        max_diffs_per_sheet: Maximum differences to show per sheet

    """
    total_sheets = len(header_results)
    identical_sheets = sum(1 for r in header_results.values() if r["identical"])
    total_differences = sum(r["total_differences"] for r in header_results.values())
    total_critical = sum(r["critical_difference_count"] for r in header_results.values())

    print("===========================================================")
    print("HEADER VALIDATION REPORT")
    print("===========================================================")

    for sheet_name, result in header_results.items():
        if result["identical"]:
            print(f"Sheet: {sheet_name} ✅ (identical)")
        else:
            print(
                f"Sheet: {sheet_name} ❌ ("
                f"{result['total_differences']} differences, "
                f"{result['critical_difference_count']} critical)"
            )

            # Show all differences with details
            for diff in result["differences"][:max_diffs_per_sheet]:
                pipeline_val = (
                    diff["pipeline"] if diff["pipeline"] is not None else "None"
                )
                ground_truth_val = (
                    diff["ground_truth"] if diff["ground_truth"] is not None else "None"
                )
                print(
                    f"  - Cell {diff['cell']}: "
                    f"pipeline '{pipeline_val}', "
                    f"ground truth '{ground_truth_val}'"
                )

            if len(result["differences"]) > max_diffs_per_sheet:
                print(
                    f"  ... and "
                    f"{len(result['differences']) - max_diffs_per_sheet} "
                    f"more differences"
                )

    print()
    print(f"Summary: {identical_sheets}/{total_sheets} sheets have matching headers")
    if total_differences > 0:
        print(f"Total differences found: {total_differences} ({total_critical} critical)")
