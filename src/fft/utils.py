"""Helper utilities."""

from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook import Workbook


def copy_sheet_content(
    source_wb_path: Path,
    target_wb_path: Path,
    sheet_name: str,
    dry_run: bool = False,
) -> dict:
    """Copy sheet content and formatting from source to target workbook.

    Args:
        source_wb_path: Path to source workbook
        target_wb_path: Path to target workbook
        sheet_name: Name of sheet to copy
        dry_run: If True, validate only without saving

    Returns:
        Dict with copy statistics (cells_copied, merged_ranges, etc.)

    Raises:
        FileNotFoundError: If source or target workbook doesn't exist
        KeyError: If sheet_name doesn't exist in either workbook

    >>> from pathlib import Path
    >>> from src.fft.utils import copy_sheet_content
    >>> # Dry run validation (no files modified)
    >>> from src.fft.config import TEMPLATES_DIR
    >>> template_path = TEMPLATES_DIR / 'FFT_IP_template.xlsm'
    >>> result = copy_sheet_content(template_path, template_path, 'BS', dry_run=True)
    >>> 'cells_copied' in result
    True
    >>> result['cells_copied'] > 0
    True

    # Edge case: Missing source file
    >>> copy_sheet_content(Path('nonexistent.xlsx'), template_path, 'BS', dry_run=True)
    Traceback (most recent call last):
        ...
    FileNotFoundError: Source workbook not found: nonexistent.xlsx

    # Edge case: Missing sheet
    >>> copy_sheet_content(template_path, template_path, 'NonExistent', dry_run=True)
    Traceback (most recent call last):
        ...
    KeyError: "Sheet 'NonExistent' not found in source workbook"
    """
    if not source_wb_path.exists():
        raise FileNotFoundError(f"Source workbook not found: {source_wb_path}")
    if not target_wb_path.exists():
        raise FileNotFoundError(f"Target workbook not found: {target_wb_path}")

    source_wb = load_workbook(source_wb_path)
    target_wb = load_workbook(target_wb_path, keep_vba=True)

    if sheet_name not in source_wb.sheetnames:
        raise KeyError(f"Sheet '{sheet_name}' not found in source workbook")
    if sheet_name not in target_wb.sheetnames:
        raise KeyError(f"Sheet '{sheet_name}' not found in target workbook")

    source_sheet = source_wb[sheet_name]
    target_sheet = target_wb[sheet_name]

    # Track statistics
    stats = {"cells_copied": 0, "merged_ranges": 0, "columns_sized": 0, "rows_sized": 0}

    # Copy cell values and styles
    for row in source_sheet.iter_rows():
        for cell in row:
            target_cell = target_sheet[cell.coordinate]
            target_cell.value = cell.value
            if cell.has_style:
                target_cell.font = copy(cell.font)
                target_cell.fill = copy(cell.fill)
                target_cell.border = copy(cell.border)
                target_cell.alignment = copy(cell.alignment)
                target_cell.number_format = cell.number_format
            stats["cells_copied"] += 1

    # Copy column widths
    for col_letter, col_dim in source_sheet.column_dimensions.items():
        target_sheet.column_dimensions[col_letter].width = col_dim.width
        stats["columns_sized"] += 1

    # Copy row heights
    for row_num, row_dim in source_sheet.row_dimensions.items():
        target_sheet.row_dimensions[row_num].height = row_dim.height
        stats["rows_sized"] += 1

    # Clear existing merged cells first, then copy new ones
    # Unmerge all existing merged cells (convert to list to avoid modification during iteration)
    for merged_range in list(target_sheet.merged_cells.ranges):
        target_sheet.unmerge_cells(str(merged_range))

    # Copy merged cells from source
    for merged_range in source_sheet.merged_cells.ranges:
        target_sheet.merge_cells(str(merged_range))
        stats["merged_ranges"] += 1

    if not dry_run:
        target_wb.save(target_wb_path)

    return stats


# %%
if __name__ == "__main__":
    import doctest
    import re
    from pathlib import Path
    from config import TEMPLATES_DIR, SUPPRESSION_FILES_DIR

    print("\nMatching templates with suppression files...")
    print(f"Templates directory: {TEMPLATES_DIR}")
    print(f"Suppression files directory: {SUPPRESSION_FILES_DIR}")

    # Get all template files
    template_files = list(TEMPLATES_DIR.glob("FFT_*_template.xlsm"))
    suppression_files = list(SUPPRESSION_FILES_DIR.glob("*.xlsm"))

    print(f"\nFound {len(template_files)} template files:")
    for template in template_files:
        print(f"  - {template.name}")

    print(f"\nFound {len(suppression_files)} suppression files:")
    for suppression in suppression_files:
        print(f"  - {suppression.name}")

    # Process each template
    matches_found = 0
    for template_path in template_files:
        # Extract service type from template filename: FFT_[SERVICE]_template.xlsm
        template_match = re.match(r"FFT_([^_]+)_template\.xlsm", template_path.name)
        if not template_match:
            print(f"  Warning: Could not extract service type from {template_path.name}")
            continue

        service_type = template_match.group(1)
        print(f"\nProcessing {service_type} service type...")

        # Find corresponding suppression file
        # Look for files containing "{SERVICE_TYPE}_Suppression"
        matching_suppressions = [
            f for f in suppression_files if f"{service_type}_Suppression" in f.name
        ]

        if not matching_suppressions:
            print(f"  No suppression file found for {service_type}")
            continue

        if len(matching_suppressions) > 1:
            print(
                f"  Warning: Multiple suppression files found for {service_type}: "
                f" {[f.name for f in matching_suppressions]}"
            )
            print(f"  Using: {matching_suppressions[0].name}")

        suppression_path = matching_suppressions[0]
        print(f"  Matched template: {template_path.name}")
        print(f"  Matched suppression: {suppression_path.name}")

        # Validate that both files have 'BS' sheet before copying
        try:
            # Check if suppression file has BS sheet
            from openpyxl import load_workbook
            suppression_wb = load_workbook(suppression_path)
            if "BS" not in suppression_wb.sheetnames:
                print(f"  ✗ Suppression file {suppression_path.name} does not contain 'BS' sheet")
                print(f"    Available sheets: {suppression_wb.sheetnames}")
                continue

            # Check if template file has BS sheet
            template_wb = load_workbook(template_path)
            if "BS" not in template_wb.sheetnames:
                print(f"  ✗ Template file {template_path.name} does not contain 'BS' sheet")
                print(f"    Available sheets: {template_wb.sheetnames}")
                continue

            print(f"  ✓ Both files contain 'BS' sheet")

            # Copy BS sheet content from suppression file to template
            print(f"  Copying BS sheet content...")
            stats = copy_sheet_content(
                source_wb_path=suppression_path,
                target_wb_path=template_path,
                sheet_name="BS",
                dry_run=False,
            )
            print(f"  ✓ Success: {stats}")
            matches_found += 1
        except Exception as e:
            print(f"  ✗ Error processing BS sheet: {e}")

    print(f"\n{matches_found} templates successfully updated with suppression data.")
