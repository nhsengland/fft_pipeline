import os  # operating system interaction
import glob  # find pathnames matching specified pattern
import pandas as pd
import numpy as np
import logging
from datetime import datetime
from pathlib import Path
from openpyxl import (
    Workbook,
    load_workbook,
)  # enable opening of and interaction with existing Macro enabled Excel files

# from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.styles import (
    PatternFill,
    Alignment,
    NamedStyle,
    Font,
)  # enable alterations to opened Excel file


def list_excel_files(folder_path, file_pattern, prefix_suffix_separator, date_format):
    """
    Function to list all files in the specified folder matching the given file pattern.
    Returns a list of matching file paths sorted by modification time in descending order.

    Parameters:
    folder_path: The path to the folder containing the Excel files.
    file_pattern: The pattern to match the Excel files (e.g., 'IPFFT-*.xlsx').
    prefix_suffix_separator: character used to split prefix and suffix (e.g. '-' or'_')
    date_format: The format the date being searched for is expected to be in (e.g. '%b%y')

    Returns:
    - list of matching file names sorted by date in descending order.

    Examples:
    >>> import tempfile
    >>> import os
    >>> from pathlib import Path
    >>> # Create a temporary directory with sample files
    >>> with tempfile.TemporaryDirectory() as temp_dir:
    ...     # Create test files with known modification times
    ...     file1 = Path(temp_dir) / "IPFFT-Jan24.xlsx"
    ...     file2 = Path(temp_dir) / "IPFFT-Feb24.xlsx"
    ...     file3 = Path(temp_dir) / "IPFFT-Mar24.xlsx"
    ...     # Touch the files to create them
    ...     file1.touch()
    ...     file2.touch()
    ...     file3.touch()
    ...     # Set modification times in ascending order
    ...     os.utime(file1, (1000, 1000))
    ...     os.utime(file2, (2000, 2000))
    ...     os.utime(file3, (3000, 3000))
    ...     # Call the function
    ...     result = list_excel_files(temp_dir, "IPFFT-*.xlsx", "-", "%b%y")
    ...     # Check that we got three files in the correct order (newest first)
    ...     len(result) == 3 and Path(result[0]).name == "IPFFT-Mar24.xlsx"
    True

    >>> # Test with non-existent files
    >>> try:
    ...     list_excel_files("/tmp/non-existent-dir", "*.xlsx", "-", "%b%y")
    ...     False  # Should not reach here
    ... except ValueError as e:
    ...     "No matching Excel files found in" in str(e)
    True
    """
    # Convert input to Path object if it's a string
    folder_path = Path(folder_path)

    # Use Path.glob to find matching files
    files = list(folder_path.glob(file_pattern))

    # Raise ValueError if no files matching pattern exists
    if not files:
        raise ValueError(
            f"No matching Excel files found in {folder_path} with pattern {file_pattern}"
        )

    # Sort files by modification time in descending order (newest first)
    sorted_files = sorted(files, key=lambda x: x.stat().st_mtime, reverse=True)

    # Convert Path objects to strings for compatibility with existing code
    sorted_files = [str(file) for file in sorted_files]

    return sorted_files


def load_excel_sheet(file_path, sheet_name):
    """
    Function to load the specified sheet from the given Excel file into a DataFrame.

    Parameters:
    file_path: The path to the Excel file.
    sheet_name: The name of the sheet to load. Can be either normalised (with underscores)
               or original (with spaces, &, and - characters).

    Returns:
    - df: containing content of the specified sheet.

    Examples:
    >>> import pandas as pd
    >>> from unittest.mock import patch
    >>> # Mock pd.read_excel to return a simple DataFrame
    >>> with patch('pandas.read_excel') as mock_read_excel:
    ...     mock_read_excel.return_value = pd.DataFrame({'Site Code': ['ABC01', 'DEF02'], '1 Very Good': [100, 50]})
    ...     df = load_excel_sheet('mock_file.xlsx', 'Parent_Self_Trusts_Site_Lev')
    ...     # Check that the function returns the mocked DataFrame
    ...     list(df.columns) == ['Site Code', '1 Very Good']
    True

    >>> # Test with sheet not found
    >>> with patch('pandas.read_excel') as mock_read_excel:
    ...     mock_read_excel.side_effect = ValueError("Sheet 'InvalidSheet' not found")
    ...     try:
    ...         load_excel_sheet('mock_file.xlsx', 'InvalidSheet')
    ...         False  # Should not reach here
    ...     except ValueError as e:
    ...         "Sheet 'InvalidSheet' not found in the file" in str(e)
    True

    >>> # Test with special sheet name conversion
    >>> with patch('pandas.read_excel') as mock_read_excel:
    ...     # First call raises ValueError, second call succeeds
    ...     mock_read_excel.side_effect = [
    ...         ValueError("Sheet not found"),
    ...         pd.DataFrame({'Column1': [1, 2]})
    ...     ]
    ...     df = load_excel_sheet('mock_file.xlsx', 'Parent_Self_Trusts_Site_Lev')
    ...     # Verify that second attempt with converted name was successful
    ...     isinstance(df, pd.DataFrame)
    True
    """
    # Helper function to process DataFrame with proper header handling
    def process_df(df):
        # Check if second row contains column names that match expected patterns
        # (like 'Yearnumber', 'Periodname', etc.)
        if df.shape[0] > 1 and isinstance(df.iloc[1, 0], str):
            # If second row appears to contain actual column headers
            # (checking if it has values like "Yearnumber")
            if "year" in str(df.iloc[1, 0]).lower() or "period" in str(df.iloc[1, 1]).lower():
                # Use second row as header
                new_headers = df.iloc[1].values
                df = df[2:].copy()  # Skip first two rows (titles and headers)
                df.columns = new_headers
        return df

    # Try to read using the provided sheet name
    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        return process_df(df)
    except ValueError:
        # If not found, try the alternative format
        try:
            # If the name has underscores, try replacing them with spaces/characters
            if "_" in sheet_name:
                # More specific conversions for known sheet name patterns
                if "Parent_Self_Trusts" in sheet_name:
                    if "Site_Lev" in sheet_name:
                        denormalised_name = "Parent & Self Trusts - Site Lev"
                    elif "Ward_Lev" in sheet_name:
                        denormalised_name = "Parent & Self Trusts - Ward Lev"
                    elif "Organisa" in sheet_name:
                        denormalised_name = "Parent & Self Trusts - Organisa"
                    elif "Collecti" in sheet_name:
                        denormalised_name = "Parent & Self Trusts - Collecti"
                    else:
                        # Generic conversion as fallback
                        denormalised_name = sheet_name.replace("_Self_", " & Self ").replace("_", " - ")
                else:
                    # Generic conversion as fallback
                    denormalised_name = sheet_name.replace("_Self_", " & Self ").replace("_", " - ")

                df = pd.read_excel(file_path, sheet_name=denormalised_name)
                return process_df(df)
            # If the name has spaces/&/-, try normalising it
            else:
                # Convert from original to normalised format
                normalised_name = sheet_name.replace(" & ", "_Self_").replace(" - ", "_").replace(" ", "_")
                df = pd.read_excel(file_path, sheet_name=normalised_name)
                return process_df(df)
        except ValueError as e:
            # Get all available sheet names for better error message
            try:
                available_sheets = pd.ExcelFile(file_path).sheet_names
                sheet_list = ", ".join(f"'{s}'" for s in available_sheets)
                raise ValueError(f"Sheet '{sheet_name}' not found in the file. Available sheets: {sheet_list}") from e
            except Exception:
                raise ValueError(f"Sheet '{sheet_name}' not found in the file.") from e


def validate_column_length(df, columns, lengths):
    """
    Function to validate that the length of each value in the specified column(s)
    matches one or more expected lengths.

    Parameters:
    df: DataFrame containing columns to check
    columns: Name of the column(s) to check (string or list of strings)
    lengths: int or list of ints, allowed lengths for the column values

    Returns:
    - df: The original DataFrame if validation passes

    Examples:
    >>> import pandas as pd
    >>> # Test single column with single length
    >>> df = pd.DataFrame({'Org Code': ['BKR', 'GRE', 'HER', 'DF3']})
    >>> validate_column_length(df, 'Org Code', 3) is df
    True

    >>> # Test single column with multiple valid lengths
    >>> df = pd.DataFrame({'Org Code': ['BKR', 'GR01E', 'HER', 'DF34F']})
    >>> validate_column_length(df, 'Org Code', [3, 5]) is df
    True

    >>> # Test multiple columns with single length
    >>> df = pd.DataFrame({
    ...     'Org Code': ['BKR', 'GR0', 'HER', 'DF3'],
    ...     'Site Code': ['BK0', 'GRE', 'RED', 'DFF']
    ... })
    >>> validate_column_length(df, ['Org Code', 'Site Code'], 3) is df
    True

    >>> # Test column not found
    >>> df = pd.DataFrame({'Org Code': ['BKR', 'GRE', 'HER', 'DF3']})
    >>> try:
    ...     validate_column_length(df, 'ICB Code', 3)
    ...     False  # Should not reach here
    ... except KeyError as e:
    ...     "Column 'ICB Code' not found in DataFrame" in str(e)
    True

    >>> # Test invalid length
    >>> df = pd.DataFrame({'Org Code': ['BKR', 'GR01E', 'HE', 'DF34F']})
    >>> try:
    ...     validate_column_length(df, 'Org Code', [3, 5])
    ...     False  # Should not reach here
    ... except ValueError as e:
    ...     "invalid length" in str(e)
    True
    """
    # Convert columns to list if a single string provided
    if isinstance(columns, str):
        columns = [columns]

    # Convert lengths to list if single integer provided
    if isinstance(lengths, int):
        lengths = [lengths]

    # Iterate over each specified column
    for column in columns:
        if column not in df.columns:
            raise KeyError(f"Column '{column}' not found in DataFrame")

        # Check if every value in the column has permitted length
        for idx, value in df[column].items():
            if len(str(value)) not in lengths:
                raise ValueError(
                    f"Row {idx} in column '{column}' contains a value with invalid length."
                )
    return df


def validate_numeric_columns(df, columns, expected_type):
    """
    Function to validate that all values in the specified column(s) are either integers or floats,
    depending on the expected type.

    Parameters:
    df: DataFrame containing data to validate
    columns: Name of the column(s) to check
    expected_type: Either 'int' or 'float', as the expected numeric type

    Returns:
    - df: The original DataFrame if validation passes

    Examples:
    >>> import pandas as pd
    >>> # Test valid single column (int)
    >>> df = pd.DataFrame({'1 Very Good': [255, 930, 459]})
    >>> validate_numeric_columns(df, '1 Very Good', 'int') is df
    True

    >>> # Test valid single column (float)
    >>> df = pd.DataFrame({'Prop_Pos': [98.5, 95.0, 96.7]})
    >>> validate_numeric_columns(df, 'Prop_Pos', 'float') is df
    True

    >>> # Test valid multiple columns
    >>> df = pd.DataFrame({'Prop_Pos': [98.5, 95.0, 96.7], 'Prop_Neg': [12.5, 15.0, 13.7]})
    >>> validate_numeric_columns(df, ['Prop_Pos', 'Prop_Neg'], 'float') is df
    True

    >>> # Test invalid column name
    >>> df = pd.DataFrame({'Prop_Pos': [98.5, 95.0, 96.7]})
    >>> try:
    ...     validate_numeric_columns(df, 'Prop_Eligible', 'float')
    ...     False  # Should not reach here
    ... except KeyError as e:
    ...     "Column 'Prop_Eligible' not found in DataFrame" in str(e)
    True

    >>> # Test invalid data type
    >>> df = pd.DataFrame({'1 Very Good': [255, 'nine hundred', 459]})
    >>> try:
    ...     validate_numeric_columns(df, '1 Very Good', 'int')
    ...     False  # Should not reach here
    ... except TypeError as e:
    ...     "non-integer value" in str(e)
    True

    >>> # Test invalid expected_type
    >>> df = pd.DataFrame({'Prop_Pos': [98.5, 95.0, 96.7]})
    >>> try:
    ...     validate_numeric_columns(df, 'Prop_Pos', 'string')
    ...     False  # Should not reach here
    ... except TypeError as e:
    ...     "Invalid expected_type 'string'" in str(e)
    True
    """

    # Ensure columns is a list
    if isinstance(columns, str):
        columns = [columns]

    # Check that the expected_type is either 'int' or 'float'
    if expected_type not in ["int", "float"]:
        raise TypeError(
            f"Invalid expected_type '{expected_type}'. Must be 'int' or 'float'."
        )

    # Iterate over all specified columns
    for column in columns:
        # Ensure the column exists in the DataFrame
        if column not in df.columns:
            raise KeyError(f"Column '{column}' not found in DataFrame.")

        # Check each value in the column
        for idx, value in df[column].items():
            if expected_type == "int" and not isinstance(value, int):
                raise TypeError(
                    f"Row {idx} in column '{column}' contains a non-integer value."
                )
            elif expected_type == "float" and not isinstance(
                value, (float, int)
            ):  # Allow int for float check
                raise TypeError(
                    f"Row {idx} in column '{column}' contains a non-float value."
                )

    return df


def get_cell_content_as_string(source_dataframe, source_row, source_col):
    """
    Function to combine a specified string with the contents of a DataFrame cell to enable dynamic update.

    Parameters:
    source_dataframe: The name of the DataFrame that contains the source text.
    source_row: Row index of the DataFrame.
    source_col: Column index of the DataFrame.

    Returns:
    - cell_content: the value within the cell specified cast as a string.
    """
    # Raise KeyError if column name does not exist in the DataFrame.
    if source_col not in source_dataframe.columns:
        raise KeyError(f"Column '{source_col}' not found in the DataFrame")

    # Raise IndexError if the row index is outside the DataFrame range.
    if source_row >= len(source_dataframe):
        raise IndexError(f"Row index {source_row} is out of range")

    # Extract cell content from the DataFrame and cast as a string
    cell_content = str(source_dataframe.at[source_row, source_col])

    return cell_content


def map_fft_period(periodname, yearnumber):
    """
    Function to generate the FFT_Period based on mapping 'Periodname' to the correct year based on 'Yearnumber'.
    Used for formatting healthcare data reporting periods in Friends and Family Test (FFT) datasets.

    Parameters:
    periodname: The period name (e.g., 'JANUARY', 'FEBRUARY').
    yearnumber: The year range (e.g., '2024-25').

    Returns:
    - A tuple containing:
      - fft_period_abbrev: The abbreviated FFT_Period (e.g., 'Jan-25')
      - fft_period_filename: The lowercase format for filenames (e.g., 'january-2025')
      - fft_period_datetime: A datetime object for the period (e.g., datetime(2025, 1, 1))

    Examples:
    >>> # Test month in first half of reporting year
    >>> result = map_fft_period('JULY', '2024-25')
    >>> result[0]  # Check abbreviated format
    'Jul-24'
    >>> result[1]  # Check filename format
    'july-2024'
    >>> from datetime import datetime
    >>> result[2].strftime('%Y-%m-%d')  # Check datetime object
    '2024-07-01'

    >>> # Test month in second half of reporting year
    >>> result = map_fft_period('JANUARY', '2024-25')
    >>> result[0]  # Check abbreviated format
    'Jan-25'
    >>> result[1]  # Check filename format
    'january-2025'
    >>> result[2].strftime('%Y-%m-%d')  # Check datetime object
    '2025-01-01'

    >>> # Test with underscore format in year range
    >>> result = map_fft_period('APRIL', '2024_25')
    >>> result[0]  # Check abbreviated format
    'Apr-24'

    >>> # Test invalid period name
    >>> try:
    ...     map_fft_period('INVALIDMONTH', '2024-25')
    ...     False  # Should not reach here
    ... except ValueError as e:
    ...     "Invalid period name 'INVALIDMONTH'" in str(e)
    True

    >>> # Test invalid year number
    >>> try:
    ...     map_fft_period('JANUARY', '2024-256')
    ...     False  # Should not reach here
    ... except ValueError as e:
    ...     "Yearnumber not in the correct format" in str(e)
    True
    """
    from datetime import datetime

    # Define a dictionary for month abbreviations
    month_abbrev = {
        "JANUARY": "Jan",
        "FEBRUARY": "Feb",
        "MARCH": "Mar",
        "APRIL": "Apr",
        "MAY": "May",
        "JUNE": "Jun",
        "JULY": "Jul",
        "AUGUST": "Aug",
        "SEPTEMBER": "Sep",
        "OCTOBER": "Oct",
        "NOVEMBER": "Nov",
        "DECEMBER": "Dec",
    }

    # Define a dictionary for month numbers
    month_num = {
        "JANUARY": 1,
        "FEBRUARY": 2,
        "MARCH": 3,
        "APRIL": 4,
        "MAY": 5,
        "JUNE": 6,
        "JULY": 7,
        "AUGUST": 8,
        "SEPTEMBER": 9,
        "OCTOBER": 10,
        "NOVEMBER": 11,
        "DECEMBER": 12,
    }

    # Raise ValueError if periodname is invalid
    if periodname not in month_abbrev:
        raise ValueError(f"Invalid period name '{periodname}'.")
    # Raise ValueError if yearnumber is invalid
    # Support both hyphen and underscore formats (2025-26 or 2025_26)
    if not (len(yearnumber) == 7 and (yearnumber[4] == "-" or yearnumber[4] == "_")):
        raise ValueError("Yearnumber not in the correct format.")

    # Extract the start and end years from the 'Yearnumber' (e.g., from '2024-25' or '2024_25')
    start_year = yearnumber[:4]  # First 4 characters represent the start year
    end_year = yearnumber[5:]  # Characters after the separator represent the end year

    # Determine which part of the year to use based on the period name
    if periodname in ["JANUARY", "FEBRUARY", "MARCH"]:
        # Use the end year for these months (e.g., Jan-25 for financial year 2024-25)
        year_to_use = "20" + end_year  # Convert to full year (e.g., 2025)
        fft_period_abbrev = f"{month_abbrev[periodname]}-{end_year}"
    else:
        # Use the start year for these months (e.g., Jul-24 for financial year 2024-25)
        year_to_use = start_year
        fft_period_abbrev = f"{month_abbrev[periodname]}-{start_year[2:]}"

    # Create a filename-friendly format (e.g., "january-2025")
    fft_period_filename = f"{periodname.lower()}-{year_to_use}"

    # Create a datetime object for the period
    month_number = month_num[periodname]
    fft_period_datetime = datetime(int(year_to_use), month_number, 1)

    # Return all formats as a tuple
    return (fft_period_abbrev, fft_period_filename, fft_period_datetime)


def remove_columns(df, columns_to_remove):
    """
    Function to remove columns not required from the DataFrame including helper columns added during suppression processing.

    Parameters:
    df: DataFrame form which to remove columns.
    columns_to_remove: single or list of columns to be removed (e.g. ['Period', 'Response Rate'] from Ward Level submissions).

    Returns:
    - df: with specified field(s) removed.
    """
    # Filter out columns that don't exist in the DataFrame
    columns_to_remove = [col for col in columns_to_remove if col in df.columns]
    # If no columns left to remove after filtering, return the DataFrame unchanged
    if not columns_to_remove:
        return df

    # drop defined columns
    return df.drop(columns=columns_to_remove)


def rename_columns(df, new_column_names):
    """
    Function to rename DataFrame columns to align with final product as specificed by stakeholders.

    Parameters:
    df: DataFrame requiring column name changes.
    new_column_names: dictionary of columns to be renamed (e.g. {'STP Code': 'ICB_Code', 'STP Name': 'ICB_Name'} from Ward Level submissions).

    Returns:
    - df: with specified fields renamed.
    """
    # Check and store any columns for renaming not present in the DataFrame and raise KeyError
    missing_columns = [col for col in new_column_names.keys() if col not in df.columns]
    if missing_columns:
        raise KeyError(
            f"The following columns to be renamed do not exist in the DataFrame: {missing_columns}"
        )

    # rename columns
    return df.rename(columns=new_column_names)


def replace_non_matching_values(df, column_name, target_value, replacement_value):
    """
    Function to replace values in a specified column of a DataFrame that do not match a target value with a replacement value.

    Parameters:
    df: The DataFrame being worked on.
    column_name: The name of the column to check and replace values.
    target_value: The value that should be retained in the column.
    replacement_value: The value to replace non-matching values with.

    Returns:
    - df: with non-matching values replaced in the specified column
    """
    # Raise a KeyError if the specified column name is not in the DataFrame
    if column_name not in df.columns:
        raise KeyError(f"Column '{column_name}' not found in the DataFrame")

    # Replace any value in the specified column that does not match the target_value (e.g. if not IS1 then NHS)
    df[column_name] = df[column_name].apply(
        lambda x: x if x == target_value else replacement_value
    )

    return df


def sum_grouped_response_fields(df, columns_to_group_by):
    """
    Function to aggregate data submitted at one level to data at a higher level e.g. aggregating Trust level data to ICB level data.
    This allows Total Response, Total Eligible and Breakdown of Responses fields to be aggregated to the necessary level before new
    calculations are carried out to generate Percentage Positive and Percentage Negative fields at the new level.

    Parameters:
    df: DataFrame requiring grouping and summing.
    columns_to_group_by: Column(s) to use for grouping (string or list of strings).

    Returns:
    - df: with selected content aggregated by specified fields and all numerical fields summed.

    Examples:
    >>> # This function is used to group data by one or more columns
    >>> # For example, it's used to aggregate Trust level data to ICB level
    >>> # or to combine NHS and Independent Provider data
    """
    # Check and store any columns specified not present in the DataFrame and raise KeyError
    missing_columns = [col for col in columns_to_group_by if col not in df.columns]
    if missing_columns:
        raise KeyError(
            f"The following columns are missing in the DataFrame: {missing_columns}"
        )

    # Identify numeric columns for summing
    numeric_columns = df.select_dtypes(include=['number']).columns.tolist()

    # Remove grouping columns from the numeric columns list
    numeric_columns = [col for col in numeric_columns if col not in columns_to_group_by]

    # For non-numeric columns (like 'Title'), we'll use first value instead of concatenating
    non_numeric_columns = [col for col in df.columns if col not in numeric_columns and col not in columns_to_group_by]

    # Create aggregation dictionary for different column types
    agg_dict = {col: 'sum' for col in numeric_columns}
    for col in non_numeric_columns:
        agg_dict[col] = 'first'  # Take the first value instead of trying to sum strings

    # Group by specified columns and apply appropriate aggregation for each column type
    return df.groupby(columns_to_group_by, as_index=False).agg(agg_dict)


def create_data_totals(df, current_fft_period, total_column_name, columns_to_sum):
    """
    Function to create monthly data totals row from existing DataFrame. Totals will have a specified column set to 'Total',
    'Period' will be the same as the current FFT Period, and all other specified columns will be the sum of the
    corresponding columns.

    Parameters:
    df: The DataFrame containing the original data.
    current_fft_period: period defined relating to the current source input file
    total_column_name: The name of the column that will have 'Total' as the value in the new row.
    columns_to_sum: A list of column names that should be summed.

    Returns:
    - df: containing single row of totals for specified columns.
    """
    # Check and store any columns specified not present in the DataFrame and raise KeyError
    missing_columns = [col for col in columns_to_sum if col not in df.columns]
    if missing_columns:
        raise KeyError(
            f"The following columns are missing in the DataFrame: {missing_columns}"
        )

    # Create a dictionary to hold the new total row data and set the specified column to 'Total'
    total_data = {total_column_name: "Total"}

    # Assume 'Period' is the same for all rows and take the value from the first row
    total_data["Period"] = current_fft_period

    # Iterate over the list of columns to sum and calculate the sum for each
    for col in columns_to_sum:
        total_data[col] = df[col].sum()

    # Create a DataFrame for the total row using the total_data dictionary
    total_df = pd.DataFrame([total_data])

    # Return the modified DataFrame with the total row added
    return total_df


def append_dataframes(df1, df2):
    """
    Function to append one DataFrame to another.

    Parameters:
    df1: DataFrame containing one set of data.
    df2: DataFrame containing another set of data with aligned columns.

    Returns:
    - df: with content of both source DataFrames appended into one DataFrame.
    """
    # Append second DataFrame to the first DataFrame
    return pd.concat([df1, df2], ignore_index=True)


def create_percentage_field(
    df, percentage_column, sum_column_one, sum_column_two, total_columm
):
    """
    Function to generate percentage positive and negative columns for survey results using percentage calculation
    sum_columns / total * 100. Used for calculating metrics like percentage positive responses in FFT data.

    Parameters:
    df: DataFrame with fields to calculate for addition of percentage columns.
    percentage_column: new column added containing results of percentage calculation.
    sum_column_one: first of two columns summed to establish percentage of total.
    sum_column_two: second of two columns summed to establish percentage of total.
    total_columm: total response field used to establish the percentage the two other columns represent.

    Returns:
    - df: with new percentage column added.

    Examples:
    >>> import pandas as pd
    >>> # Test normal case
    >>> df = pd.DataFrame({
    ...     'Very Good': [30, 50, 60],
    ...     'Good': [20, 30, 40],
    ...     'Total Responses': [100, 100, 120]
    ... })
    >>> result = create_percentage_field(df, 'Percentage Positive', 'Very Good', 'Good', 'Total Responses')
    >>> list(result.columns)
    ['Very Good', 'Good', 'Total Responses', 'Percentage Positive']
    >>> result['Percentage Positive'].tolist()
    [0.5, 0.8, 0.83]

    >>> # Test with zeros
    >>> df = pd.DataFrame({
    ...     'Very Good': [0, 10, 20],
    ...     'Good': [0, 0, 10],
    ...     'Total Responses': [0, 20, 50]
    ... })
    >>> result = create_percentage_field(df, 'Percentage Positive', 'Very Good', 'Good', 'Total Responses')
    >>> import math
    >>> result_values = result['Percentage Positive'].tolist()
    >>> len(result_values) == 3 and math.isnan(result_values[0]) and result_values[1:] == [0.5, 0.6]
    True

    >>> # Test missing columns
    >>> df = pd.DataFrame({
    ...     'Very Good': [50, 30, 70],
    ...     'Very Poor': [10, 20, 5]  # Missing 'Total Responses'
    ... })
    >>> try:
    ...     create_percentage_field(df, 'Percentage Positive', 'Very Good', 'Good', 'Total Responses')
    ...     False  # Should not reach here
    ... except KeyError as e:
    ...     "missing in the DataFrame" in str(e)
    True
    """
    # Check and store any columns specified not present in the DataFrame and raise KeyError
    missing_columns = [
        col
        for col in [sum_column_one, sum_column_two, total_columm]
        if col not in df.columns
    ]
    if missing_columns:
        raise KeyError(
            f"The following columns are missing in the DataFrame: {missing_columns}"
        )

    # Convert columns to numeric, coercing errors to NaN
    sum_col1 = pd.to_numeric(df[sum_column_one], errors='coerce')
    sum_col2 = pd.to_numeric(df[sum_column_two], errors='coerce')
    total_col = pd.to_numeric(df[total_columm], errors='coerce')

    # Calculate percentage, handling division by zero
    df[percentage_column] = round(
        ((sum_col1 + sum_col2) / total_col.replace(0, float('nan'))) * 1, 2
    )

    return df


def remove_rows_by_cell_content(df, column_name, cell_value):
    """
    Function to remove rows from DataFrame based on specified cell content in a given column.

    Parameters:
    df: DataFrame from which rows will be removed.
    column_name: The name of the column where the specific value is located.
    cell_value: The specific value that will determine row removal.

    Returns:
    - df: with specified rows removed.
    """
    # Raise KeyError if column_name for searching values not in DataFrame
    if column_name not in df.columns:
        raise KeyError(f"Column '{column_name}' not found in the DataFrame")

    # Display the initial number of rows in the DataFrame before removal
    logging.info(f"Initial number of rows: {len(df)}")

    # Filter DataFrame keeping rows where specified column content does NOT match specified cell value.
    df = df[df[column_name] != cell_value]

    # Display number of rows after filtering helping identify number removed
    logging.info(f"Number of rows after removal: {len(df)}")

    return df


def reorder_columns(df, column_order):
    """
    Function to reorder the DataFrame columns based on a specified order.

    Parameters:
    df: The DataFrame to reorder.
    column_order: List of desired column order.

    Returns:
    - DataFrame with columns reordered according to provided list.
    """
    # Filter column_order to only include columns that exist in the DataFrame
    available_columns = [col for col in column_order if col in df.columns]

    # Get columns in the DataFrame that aren't in column_order
    remaining_columns = [col for col in df.columns if col not in available_columns]

    # Combine the available ordered columns with any remaining columns
    final_column_order = available_columns + remaining_columns

    # Log the discrepancy if there is one
    if set(final_column_order) != set(column_order) or set(final_column_order) != set(df.columns):
        logging.warning("Some columns in the order list were not found in the DataFrame or vice versa. Using available columns in the specified order, followed by remaining columns.")
        logging.warning(f"Columns in DataFrame but not in order list: {[col for col in df.columns if col not in column_order]}")
        logging.warning(f"Columns in order list but not in DataFrame: {[col for col in column_order if col not in df.columns]}")

    # Reorder the columns of the DataFrame
    df = df[final_column_order]

    return df


def convert_fields_to_object_type(df, fields_to_convert):
    """
    Function converting specified columns to object type (e.g. like those undergoing suppression process). When saved into Excel this ensures number values
    remain as numbers, while avoiding raising python errors when suppressing values with non-numeric values. Object data type stores any
    data type. - https://pandas.pydata.org/pandas-docs/stable/user_guide/basics.html#basics-dtypes

    Parameters:
    df: The input DataFrame.
    fields_to_convert: List of fields that require converting e.g. for number (int) fields where suppression applied to string (str)


    Returns:
    - df: specified columns converted to object type.
    """

    # Filter fields to only include columns that exist in the DataFrame
    available_fields = [col for col in fields_to_convert if col in df.columns]

    # Log warning for missing columns
    missing_fields = [col for col in fields_to_convert if col not in df.columns]
    if missing_fields:
        logging.warning(f"The following fields for conversion do not exist in the DataFrame: {missing_fields}")

    # Only convert columns that actually exist
    if available_fields:
        df[available_fields] = df[available_fields].astype(object)

    return df


def replace_missing_values(df, replacement_value):
    """
    Function to replace null (NaN) values in a DataFrame with specified value.
    Used to clean data before display or analysis.

    Parameters:
    df: DataFrame to be cleaned.
    replacement_value: Value to replace NaN values with (e.g., "NA", 0, etc.)

    Returns:
    - df: Cleaned DataFrame with NaN values replaced.

    Examples:
    >>> # Example: Create a dataframe with NaN values
    >>> import pandas as pd
    >>> import numpy as np
    >>> df = pd.DataFrame({'A': [1, np.nan]})
    >>> # Replace NaNs with 0
    >>> result = replace_missing_values(df, 0)
    >>> # Get value from first and second rows
    >>> list(result['A'])
    [1.0, 0.0]
    """
    df_cleaned = df.copy()
    # Replace nan with specified value
    with pd.option_context("future.no_silent_downcasting", True):
        df_cleaned.fillna(replacement_value, inplace=True)

    return df_cleaned


def count_nhs_is1_totals(df, code_column_name, is1_count_var_name, nhs_count_var_name):
    """
    Function counting number of rows with 'IS1' and 'NHS' in a specified column, storing counts in variables with
    specified names for use as summary totals.

    Parameters:
    df: containing the data.
    code_column_name: The name of the column to check for 'IS1' and 'NHS' values.
    is1_count_var_name: The name of the variable to store the count of 'IS1'.
    nhs_count_var_name: The name of the variable to store the count of 'NHS'.

    Returns:
    - Dictionary with counts stored against specified variable names.
    """
    # Raise KeyError if the code_column_name does not exist in the DataFrame
    if code_column_name not in df.columns:
        raise KeyError(f"Column '{code_column_name}' not found in the DataFrame.")

    # Count number of rows containing 'IS1'
    count_of_IS1 = df[df[code_column_name] == "IS1"].shape[0]

    # Count number of rows containing 'NHS'
    count_of_NHS = df[df[code_column_name] == "NHS"].shape[0]

    # Return counts in a dictionary with specified variable names
    return {is1_count_var_name: count_of_IS1, nhs_count_var_name: count_of_NHS}


def add_dataframe_column(df, column_name, column_value):
    """
    Function to create a new DataFrame column, specifying the DataFrame, column name and value it has.

    Parameters:
    df: DataFrame to which the new column will be added.
    column_name: The name of the new column.
    column_value: The value with which to populate or instantiate the new column.

    Returns:
    - df: with the new populated column added.
    """
    # Rasie TypeError if new column name not specified as a string or list
    if isinstance(column_name, str):
        column_name = [column_name]

    if not isinstance(column_name, list):
        raise TypeError("New column name should be a list or a string.")

    # Add the new column to the DataFrame with the specified initial value
    df[column_name] = column_value

    return df


def add_submission_counts_to_df(
    df, code_column_name, is1_count, nhs_count, target_column
):
    """
    Function to add submission counts as specified column to the specified DataFrame.

    Parameters:
    df: The DataFrame to which the counts should be added.
    code_column_name: The name of the column containing 'IS1' and 'NHS'.
    is1_count: The count of rows with 'IS1' taken from a previous aggregation.
    nhs_count: The count of rows with 'NHS' taken from a previous aggregation.
    target_column: The name of the column to add counts to.

    Returns:
    - df: updated with tatget column populated with counts.
    """
    # Raise KeyError if `code_column_name` or `target_column` does not exist in the DataFrame.
    if code_column_name not in df.columns or target_column not in df.columns:
        raise KeyError(
            f"One or both of columns '{code_column_name}' or '{target_column}' are missing from the DataFrame."
        )

    # Add the count to the row where ICB Code is 'IS1'
    df.loc[df[code_column_name] == "IS1", target_column] = is1_count

    # Add the count to the row where ICB Code is 'NHS'
    df.loc[df[code_column_name] == "NHS", target_column] = nhs_count

    return df


def update_monthly_rolling_totals(df1, df2, current_fft_period):
    """
    Function to transfer data from df1 to df2 in the next blank row or overwrite an existing row if the period exists.
    Used for maintaining monthly rolling totals of FFT survey data.

    Parameters:
    df1: Source DataFrame containing data to be transferred (must have specific columns and format).
    df2: Destination DataFrame where the data will be transferred (must have specific columns and format).
    current_fft_period: value generated for the current fft period.

    Returns:
    - df2: updated with data from df1 transferred to the next available row if the period doesn't already exist,
    or overwritten if it does.

    Examples:
    >>> import pandas as pd
    >>> # Normal case - Add new period data
    >>> # Source data with responses by submitter type
    >>> source_df = pd.DataFrame({
    ...     'Submitter Type': ['Total', 'NHS', 'IS1'],
    ...     'Number of organisations submitting': [10, 7, 3],
    ...     'Total Responses': [1000, 700, 300],
    ...     'Percentage Positive': [80, 75, 85],
    ...     'Percentage Negative': [10, 15, 5]
    ... })
    >>> # Rolling totals with previous periods
    >>> rolling_df = pd.DataFrame({
    ...     'FFT Period': ['2024-04', '2024-05'],
    ...     'Total submitters': [12, 15],
    ...     'Number of NHS submitters': [8, 9],
    ...     'Number of Independent submitters': [4, 6],
    ...     'Total responses to date': [1200, 1500],
    ...     'Total NHS responses to date': [800, 900],
    ...     'Total independent responses to date': [400, 600],
    ...     'Monthly total responses': [1000, 300],
    ...     'Monthly NHS responses': [700, 100],
    ...     'Monthly independent responses': [300, 200],
    ...     'Monthly total percentage positive': [80, 85],
    ...     'Monthly NHS percentage positive': [75, 80],
    ...     'Monthly independent percentage positive': [85, 90],
    ...     'Monthly total percentage negative': [10, 12],
    ...     'Monthly NHS percentage negative': [15, 10],
    ...     'Monthly independent percentage negative': [5, 4],
    ... })
    >>> result = update_monthly_rolling_totals(source_df, rolling_df, '2024-06')
    >>> len(result)  # One row should be added
    3
    >>> result['FFT Period'].tolist()[-1]  # Should contain the new period
    '2024-06'
    >>> result['Monthly total responses'].tolist()[-1]  # Should match source data
    1000

    >>> # Case: Updating existing period
    >>> result = update_monthly_rolling_totals(source_df, result, '2024-06')
    >>> len(result)  # Length should stay the same (update, not add)
    3

    >>> # Edge case: Missing required columns
    >>> incomplete_df = pd.DataFrame({'Submitter Type': ['Total', 'NHS']})
    >>> try:
    ...     update_monthly_rolling_totals(incomplete_df, rolling_df, '2024-07')
    ...     False  # Should not reach here
    ... except KeyError as e:
    ...     'Missing columns in df1' in str(e)
    True

    >>> # Edge case: Missing FFT Period column
    >>> invalid_rolling_df = pd.DataFrame({'Period': ['2024-04', '2024-05']})
    >>> try:
    ...     update_monthly_rolling_totals(source_df, invalid_rolling_df, '2024-07')
    ...     False  # Should not reach here
    ... except KeyError as e:
    ...     "'FFT Period' column is missing" in str(e)
    True
    """
    # Raise KeyError if necessary columns are missing from df1.
    required_columns_df1 = [
        "Submitter Type",
        "Number of organisations submitting",
        "Total Responses",
        "Percentage Positive",
        "Percentage Negative",
    ]
    missing_columns_df1 = [col for col in required_columns_df1 if col not in df1.columns]
    if missing_columns_df1:
        raise KeyError(f"Missing columns in df1: {missing_columns_df1}")

    # Raise KeyError if FFT Period column is missing from df2.
    if "FFT Period" not in df2.columns:
        raise KeyError("'FFT Period' column is missing in df2.")

    # Create a new row for df2 based on df1 data
    new_row = {
        "FFT Period": current_fft_period,
        "Total submitters": df1.loc[
            df1["Submitter Type"] == "Total", "Number of organisations submitting"
        ].values[0] if len(df1[df1["Submitter Type"] == "Total"]) > 0 else 0,
        "Number of NHS submitters": df1.loc[
            df1["Submitter Type"] == "NHS", "Number of organisations submitting"
        ].values[0] if len(df1[df1["Submitter Type"] == "NHS"]) > 0 else 0,
        "Number of Independent submitters": df1.loc[
            df1["Submitter Type"] == "IS1", "Number of organisations submitting"
        ].values[0] if len(df1[df1["Submitter Type"] == "IS1"]) > 0 else 0,
        "Total responses to date": 0,  # Initiating with 0
        "Total NHS responses to date": 0,  # Initiating with 0
        "Total independent responses to date": 0,  # Initiating with 0
        "Monthly total responses": df1.loc[
            df1["Submitter Type"] == "Total", "Total Responses"
        ].values[0] if len(df1[df1["Submitter Type"] == "Total"]) > 0 else 0,
        "Monthly NHS responses": df1.loc[
            df1["Submitter Type"] == "NHS", "Total Responses"
        ].values[0] if len(df1[df1["Submitter Type"] == "NHS"]) > 0 else 0,
        "Monthly independent responses": df1.loc[
            df1["Submitter Type"] == "IS1", "Total Responses"
        ].values[0] if len(df1[df1["Submitter Type"] == "IS1"]) > 0 else 0,
        "Monthly total percentage positive": df1.loc[
            df1["Submitter Type"] == "Total", "Percentage Positive"
        ].values[0],
        "Monthly NHS percentage positive": df1.loc[
            df1["Submitter Type"] == "NHS", "Percentage Positive"
        ].values[0],
        "Monthly independent percentage positive": df1.loc[
            df1["Submitter Type"] == "IS1", "Percentage Positive"
        ].values[0] if len(df1[df1["Submitter Type"] == "IS1"]) > 0 else 0,
        "Monthly total percentage negative": df1.loc[
            df1["Submitter Type"] == "Total", "Percentage Negative"
        ].values[0],
        "Monthly NHS percentage negative": df1.loc[
            df1["Submitter Type"] == "NHS", "Percentage Negative"
        ].values[0],
        "Monthly independent percentage negative": df1.loc[
            df1["Submitter Type"] == "IS1", "Percentage Negative"
        ].values[0] if len(df1[df1["Submitter Type"] == "IS1"]) > 0 else 0,
    }

    # Check if 'current_fft_period' already exists in 'FFT Period' column of df2
    if current_fft_period in df2["FFT Period"].values:
        # Find the index of the existing row
        row_index = df2[df2["FFT Period"] == current_fft_period].index[0]

        # Overwrite the existing row with the new values
        for key, value in new_row.items():
            df2.at[row_index, key] = value

        logging.info(
            f"Figures for the current month ({current_fft_period}) have been overwritten."
        )
    else:
        # Convert the dictionary to a DataFrame and add it to bottom of df2
        new_row_df = pd.DataFrame([new_row])
        df2 = pd.concat([df2, new_row_df], ignore_index=True)

        logging.info(
            f"New figures for the current month ({current_fft_period}) have been added to the DataFrame."
        )

    return df2


def update_cumulative_value(df, first_column, second_column):
    """
    Function takes the value from last row of a specified source column, adds it to the value from the second-to-last row of
    another specified source column, and updates the target column with the result in the last row.
    Used to update cumulative totals in the monthly rolling totals report.

    Parameters:
    df: The DataFrame containing the data.
    first_column: The column from which to take the value from the last row (e.g., 'Monthly total responses').
    second_column: The column to update with the cumulative value in the last row (e.g., 'Total responses to date').

    Returns:
    - df: updated with the cumulative value in the last row.

    Examples:
    >>> import pandas as pd
    >>> # Normal case - Update cumulative values
    >>> df = pd.DataFrame({
    ...     'Period': ['2024-04', '2024-05', '2024-06'],
    ...     'Monthly responses': [100, 150, 200],
    ...     'Total responses to date': [100, 250, 0]  # Last value needs updating
    ... })
    >>> result = update_cumulative_value(df, 'Monthly responses', 'Total responses to date')
    >>> result['Total responses to date'].tolist()  # Last value should be 250 (previous) + 200 (current)
    [100, 250, 450]

    >>> # Edge case - Fewer than two rows
    >>> df_single = pd.DataFrame({
    ...     'Monthly responses': [100],
    ...     'Total responses to date': [0]
    ... })
    >>> try:
    ...     update_cumulative_value(df_single, 'Monthly responses', 'Total responses to date')
    ...     False  # Should not reach here
    ... except ValueError as e:
    ...     "must have at least two rows" in str(e)
    True

    >>> # Edge case - Missing columns
    >>> df_missing = pd.DataFrame({
    ...     'Period': ['2024-04', '2024-05'],
    ...     'Wrong Column': [100, 150]
    ... })
    >>> try:
    ...     update_cumulative_value(df_missing, 'Monthly responses', 'Total responses to date')
    ...     False  # Should not reach here
    ... except KeyError as e:
    ...     "do not exist in the DataFrame" in str(e)
    True
    """
    # Raise KeyError if either column does not exist in the DataFrame
    if first_column not in df.columns or second_column not in df.columns:
        raise KeyError(
            f"One or both columns '{first_column}' and '{second_column}' do not exist in the DataFrame."
        )

    # Raise ValueError if there are at less than two rows
    if len(df) < 2:
        raise ValueError(
            "The DataFrame must have at least two rows to update cumulative values."
        )

    # Identify the index of the last and second-to-last rows
    last_idx = df.index[-1]  # Index of the last row
    second_last_idx = df.index[-2]  # Index of the second-to-last row

    # Extract the values from the specified columns
    first_value = df.at[last_idx, first_column]
    second_value = df.at[second_last_idx, second_column]

    # Addition values together
    cumulative_value = first_value + second_value

    # Update the target column in the second-to-bottom row with the new value
    df.at[last_idx, second_column] = cumulative_value

    # Return the updated DataFrame
    return df


def update_existing_excel_sheet(file_path, sheet_name, updated_df):
    """
    Function to update a specific sheet in an existing Excel file with new data from a DataFrame.

    Parameters:
    file_path: Path to the existing Excel file.
    sheet_name: Name of the sheet to update.
    updated_df: The DataFrame containing the updated data to be written back to the sheet.

    Returns:
    None - Specified Excel sheet updated.
    """
    # Convert to Path object if it's a string
    path = Path(file_path)

    # Raise FileNotFoundError if the file does not exist
    if not path.exists():
        raise FileNotFoundError(f"The file '{path}' does not exist.")

    # Load the existing Excel file using openpyxl engine
    with pd.ExcelWriter(
        str(path), engine="openpyxl", mode="a", if_sheet_exists="replace"
    ) as writer:
        # Write the updated DataFrame to the specified sheet
        updated_df.to_excel(writer, sheet_name=sheet_name, index=False)

    logging.info(f"Sheet '{sheet_name}' in '{file_path}' has been updated successfully.")


def copy_value_between_dataframes(
    df_source, df_target, source_column, source_row, target_column, target_row
):
    """
    Function to copy a value from a specific row and column in one DataFrame to a specific row and column in another DataFrame.

    Parameters:
    df_source: The source DataFrame from which the value will be copied.
    df_target: The target DataFrame where the value will be pasted.
    source_column: The column in the source DataFrame from which to copy the value.
    source_row: The row index in the source DataFrame from which to copy the value.
    target_column: The column in the target DataFrame where the value will be pasted.
    target_row: The row index in the target DataFrame where the value will be pasted.

    Returns:
    - df: with the value pasted into the specified row and column.
    """
    # Raise KeyError if either source or target columns do not exist in respective DataFrames
    if source_column not in df_source.columns:
        raise KeyError(f"Source column '{source_column}' does not exist in df_source.")
    if target_column not in df_target.columns:
        raise KeyError(f"Target column '{target_column}' does not exist in df_target.")

    # Check if source row exists in the source DataFrame
    if source_row not in df_source.index:
        # Log a warning but don't raise an error
        logging.warning(f"Source row '{source_row}' does not exist in df_source. No value copied.")
        return df_target

    # Check if target row exists in the target DataFrame
    if target_row not in df_target.index:
        # Log a warning but don't raise an error
        logging.warning(f"Target row '{target_row}' does not exist in df_target. No value copied.")
        return df_target

    # Copy the value from the source DataFrame
    value_to_copy = df_source.at[source_row, source_column]

    # Paste the value into the target DataFrame
    df_target.at[target_row, target_column] = value_to_copy

    return df_target


def new_column_name_with_period_prefix(period_prefix, new_column_suffix):
    """
    Function to create a new column name for use in creating a new column, with the prefix of a specified data period.

    Parameters:
    period_prefix: month-year variable e.g., current_fft_month, previous_fft_month required for new column name.
    new_column_suffix: suffix required of the new column name.

    Returns:
    - new column name
    """
    # Raise TypeError if either period_prefix or new_column_suffix is not a string.
    if not isinstance(period_prefix, str):
        raise TypeError("period_prefix must be a string")

    if not isinstance(new_column_suffix, str):
        raise TypeError("new_column_suffix must be a string")

    # Join new_column_suffix wih period_prefix to form full column name
    new_column_name = period_prefix + "_" + new_column_suffix

    return new_column_name


def sort_dataframe(df, df_fields, directions):
    """
    Function to sort the selected DataFrame by specified field(s).

    Parameters:
    df: DataFrame to sort by.
    df_fields: field(s) to sort DataFrame by. Can be single field (string) or list of fields.
    directions: sort direction(s). (True for A to Z, False for Z to A). Can be single boolean or a list for each field.

    Returns:
    - df: with content ordered by the specified field(s).
    (ignore_index (True) rests the index to 0-based sequence rather than retaining the excisting indexing)
    """
    # Ensure df_fields is a list
    if isinstance(df_fields, str):
        df_fields = [df_fields]

    # Filter df_fields to only include columns that exist in the DataFrame
    available_fields = [field for field in df_fields if field in df.columns]

    # If no fields are available for sorting, return the DataFrame unchanged
    if not available_fields:
        logging.warning("None of the specified sort fields exist in the DataFrame. Returning DataFrame unchanged.")
        return df

    # Log a warning if some fields are missing
    if len(available_fields) < len(df_fields):
        missing_fields = [field for field in df_fields if field not in df.columns]
        logging.warning(f"The following sort fields are not in the DataFrame: {missing_fields}")

    # Ensure directions is appropriate for the available fields
    if isinstance(directions, bool):
        # If directions is a single boolean, use it for all fields
        adjusted_directions = [directions] * len(available_fields)
    else:
        # If directions is a list, filter it to match available fields
        if len(directions) == len(df_fields):
            adjusted_directions = [directions[i] for i, field in enumerate(df_fields) if field in available_fields]
        else:
            # If directions length doesn't match df_fields, use True for all fields
            adjusted_directions = [True] * len(available_fields)
            logging.warning("Sort directions length does not match fields length. Using ascending (True) for all fields.")

    # Sort the DataFrame using the available fields and adjusted directions
    df = df.sort_values(by=available_fields, ascending=adjusted_directions, ignore_index=True)
    return df


def create_first_level_suppression(df, first_level_suppression, responses_field):
    """
    Function to add first level suppression field to ensure any row of data reporting less than 5 total responses is marked for suppression.
    This also applies to first level suppression of next level aggregation e.g. at Trust level this applies to ICB,
    at Site Level it applies to Trust, and at Ward Level it applies to Site.

    Parameters:
    df: The input DataFrame.
    responses_field: the field containing the survey response totals e.g. Responses.
    first_level_suppression: the name of the new field to add containing first level suppression.

    Returns:
    - df: with additional column distinguishing columns requiring direct (first level) suppression (1) versus not (0).

    Examples:
    >>> import pandas as pd
    >>> # Normal case with small and large counts
    >>> df = pd.DataFrame({
    ...     'Trust Code': ['ABC', 'DEF', 'GHI', 'JKL', 'MNO'],
    ...     'Total Responses': [0, 3, 4, 5, 20]
    ... })
    >>> result = create_first_level_suppression(df, 'Suppress', 'Total Responses')
    >>> result['Suppress'].tolist()  # 0 responses → 0, 1-4 responses → 1, 5+ responses → 0
    [0, 1, 1, 0, 0]

    >>> # Edge case - Exact threshold (5)
    >>> df = pd.DataFrame({'Total Responses': [4, 5, 6]})
    >>> result = create_first_level_suppression(df, 'Suppress', 'Total Responses')
    >>> result['Suppress'].tolist()
    [1, 0, 0]

    >>> # Edge case - Missing response column
    >>> df = pd.DataFrame({'Wrong Column': [1, 2, 3]})
    >>> try:
    ...     create_first_level_suppression(df, 'Suppress', 'Total Responses')
    ...     False  # Should not reach here
    ... except KeyError as e:
    ...     "'Total Responses' does not exist" in str(e)
    True
    """
    # Raise KeyError if responses_field does not exist
    if responses_field not in df.columns:
        raise KeyError(f"'{responses_field}' does not exist in the DataFrame.")

    # Create the direct suppression column based on condition applied to the responses column (value is 1 if less than 5, otherwise value is 0)
    df[first_level_suppression] = df[responses_field].apply(
        lambda x: 1 if 0 < x < 5 else 0
    )
    return df


def create_icb_second_level_suppression(
    df, first_level_suppression, second_level_suppression
):
    """
    Function to add second level suppression column for icb level only to ensure for any icb with first level suppression, an additional icb is suppressed
    maximising security against submissions being patient identifiable.

    Parameters:
    df: The input DataFrame.
    first_level_suppression: the field showing rows where first level suppression has been added (as 1).
    second_level_suppression: the new field to be added to highlight rows where second level suppression is required (1) or not (0).

    Returns:
    - df: with additional column distinguishing columns requiring second level suppression (1) versus not (0).
    """
    # Create the second level suppression field with default value of 0
    df[second_level_suppression] = 0

    # Iterate through the DataFrame starting from the 2nd row (index 1) as the first row can't have a preceding row being suppressed
    for i in range(1, len(df)):
        # Check if the previous rows 'first_level_suppression' is 1
        if df.at[i - 1, first_level_suppression] == 1:
            # Set second_level_suppression to 1 for the current row
            df.at[i, second_level_suppression] = 1
    return df


def confirm_row_level_suppression(df, suppression_field, *suppression_columns):
    """
    Function to add a new suppression field which shows 1 if the row needs suppressing based on need for suppression in any of the other fields checking
    need for suppression, and 0 if it doesn't need suppressing.

    Parameters:
    df: The input DataFrame.
    suppression_field: new field added to show where rows will need suppressing.
    *suppression_columns: fields to check for suppression.

    Returns:
    - df: with a additional field confirming which rows need suppressing based on all other suppression fields.
    """
    # Check and store any columns specified not present in the DataFrame and raise KeyError
    missing_columns = [col for col in suppression_columns if col not in df.columns]
    if missing_columns:
        raise KeyError(
            f"The following suppression columns are missing from the DataFrame: {', '.join(missing_columns)}"
        )

    # Initialize the overall suppression_field column with 0s
    df[suppression_field] = 0

    # Iterate over all specified suppression columns
    for column in suppression_columns:
        # Update the 'suppression_field' column if any specified suppression column has a 1
        df[suppression_field] = df[suppression_field] | (df[column] == 1)

    # Convert 'suppress_field' to integer (whole number) type
    df[suppression_field] = df[suppression_field].astype(int)

    return df


def suppress_data(df, overall_suppression_field, first_level_suppression_field):
    """
    Function replaces specified column values with '*' based on conditions of 'first_level_suppression' and 'overall_suppression_field'.
    Used to mask sensitive healthcare data with small counts to prevent identification.

    Parameters:
    df: The input DataFrame
    overall_suppression_field: field highlighting rows in the DataFrame that require some level of suppression.
    first_level_suppression_field: field highlighting rows in the DataFrame that require percentage fields suppressing as well.

    Returns:
    - df: modified with suppressed values.

    Examples:
    >>> import pandas as pd
    >>> # Normal case with both levels of suppression
    >>> df = pd.DataFrame({
    ...     'Very Good': [3, 10, 30],
    ...     'Good': [1, 5, 20],
    ...     'Neither Good nor Poor': [0, 3, 10],
    ...     'Poor': [0, 2, 5],
    ...     'Very Poor': [0, 0, 3],
    ...     'Dont Know': [0, 0, 2],
    ...     'Percentage Positive': [0.8, 0.75, 0.7],
    ...     'Percentage Negative': [0.0, 0.1, 0.11],
    ...     'overall_suppression': [1, 1, 0],
    ...     'first_level_suppression': [1, 0, 0]
    ... }).astype(object)
    >>> result = suppress_data(df, 'overall_suppression', 'first_level_suppression')
    >>> result['Very Good'].tolist()  # All response columns for overall suppression
    ['*', '*', 30]
    >>> result['Percentage Positive'].tolist()  # Only first row has percentages suppressed
    ['*', 0.75, 0.7]

    >>> # Edge case - Empty DataFrame
    >>> df_empty = pd.DataFrame(columns=['Very Good', 'overall_suppression', 'first_level_suppression'])
    >>> result = suppress_data(df_empty, 'overall_suppression', 'first_level_suppression')
    >>> result.empty
    True

    >>> # Edge case - Missing columns handled gracefully
    >>> df = pd.DataFrame({
    ...     'Very Good': [10, 20],  # Only one response column
    ...     'overall_suppression': [1, 0],
    ...     'first_level_suppression': [1, 0]
    ... }).astype(object)
    >>> result = suppress_data(df, 'overall_suppression', 'first_level_suppression')
    >>> result['Very Good'].tolist()  # Suppression works with available columns
    ['*', 20]
    """
    # List of columns to replace with '*' if 'overall suppression required
    response_column_suppression = [
        "Very Good",
        "Good",
        "Neither Good nor Poor",
        "Poor",
        "Very Poor",
        "Dont Know",
    ]

    # List of additional columns to replace with '*' if 'first_level_suppression' is also 1
    percentage_column_suppression = ["Percentage Positive", "Percentage Negative"]

    # Check and store any specified columns not present in the DataFrame and log warning
    all_suppression_columns = response_column_suppression + percentage_column_suppression
    available_response_columns = [col for col in response_column_suppression if col in df.columns]
    available_percentage_columns = [col for col in percentage_column_suppression if col in df.columns]
    missing_columns = [col for col in all_suppression_columns if col not in df.columns]

    if missing_columns:
        logging.warning(f"The following columns are missing in the DataFrame for suppression: {', '.join(missing_columns)}")

    # Iterate over each row in the DataFrame
    for index, row in df.iterrows():
        if row[overall_suppression_field] == 1:
            # Replace specified columns with '*' for 'overall level suppression columns'
            if available_response_columns:
                df.loc[index, available_response_columns] = "*"
            if row[first_level_suppression_field] == 1:
                # Replace percentage columns with '*' where 'first_level_suppression' required
                if available_percentage_columns:
                    df.loc[index, available_percentage_columns] = "*"
    return df


def move_independent_provider_rows_to_bottom(df):
    """
    Function to move all indpendent provider (IS1) rows to the bottom of the DataFrame.

    Parameters:
    df: The input DataFrame.

    Returns:
    - df: modified with suppressed values.
    """
    # Create a boolean mask identifying rows where 'ICB Code' contains independent provider marker 'IS1'
    is_ip = df["ICB Code"].str.contains("IS1")

    # Separate the DataFrame into two parts: rows with 'IS1' and rows without 'IS1'
    ip_rows = df[is_ip]  # Rows where 'ICB Code' contains 'IS1'
    other_rows = df[~is_ip]  # Rows where 'ICB Code' does not contain 'IS1'

    # Concatenate the DataFrame with 'other_rows' first and 'ip_rows' second
    result_df = pd.concat([other_rows, ip_rows], ignore_index=True)

    return result_df


def adjust_percentage_field(df, percentage_column):
    """
    Function to convert existing percentage positive and negative columns to ensure they present correctly in final outputs
    (percentage_column / 100)

    Parameters:
    df: DataFrame with fields to calculate for adjusting of percentage columns.
    percentage_column: existing column containing percentage in incorrect format.

    Returns:
    - df: with percentage column adjusted.
    """
    # Raise KeyError if the specified percentage column does not exist in the DataFrame.
    if percentage_column not in df.columns:
        raise KeyError(f"Column '{percentage_column}' not found in the DataFrame.")

    # Convert column to numeric values first, coercing errors to NaN
    df[percentage_column] = pd.to_numeric(df[percentage_column], errors='coerce')

    # Now perform the division safely
    df[percentage_column] = round(df[percentage_column] / 100, 2)

    return df


def rank_organisation_results(df, org_field, responses_field, rank_field):
    """
    Function to add Rank column to the DataFrame, ranking Responses within each e.g. Site Code group.

    Parameters:
    df: The input DataFrame
    org_field: the field containing the organisation level code to sort by e.g. Site_Code.
    responses_field: the field containing the survey response totals e.g. Responses.
    rank_field: the name of the new field to add containing ranking e.g. Rank.

    Returns:
    - df: with an additional column containing rankings by organisation code. Any Responses value of 0 will be ranked 0.
    All remaining responses by organisation code will be ranked from 1 for lowest non 0 response level, upwards.
    """
    # Raise KeyError if either org_field or responses_field do not exist
    if org_field not in df.columns or responses_field not in df.columns:
        raise KeyError(
            f"'{org_field}' or '{responses_field}' does not exist in the DataFrame."
        )

    # Create a mask for non-zero responses
    non_zero_mask = df[responses_field] != 0

    # Create the ranking column with default of 0 for all values
    df[rank_field] = 0

    # Apply ranking to non-zero responses within each org_field grouping excluding rows where responses are 0
    df.loc[non_zero_mask, rank_field] = (
        df[non_zero_mask]
        .groupby(org_field)[responses_field]
        .rank(method="dense")
        .astype(int)
    )
    return df


def create_second_level_suppression(
    df, first_level_suppression, rank_field, second_level_suppression
):
    """
    Function to add second level suppression column to ensure where an organisation has more than one submission row,
    where the first submission is greater than 0 but less than 5 and requires suppression,
    the second submission is also suppressed maximising security against submissions being patient identifiable.

    Parameters:
    df: The input DataFrame.
    rank_field: the field showing ranked order for an organisations submissions.
    first_level_suppression: the field showing rows where first level suppression has been added (as 1).
    second_level_suppression: the new field to be added to highlight rows where second level suppression is required (1) or not (0).

    Returns:
    - df: with additional column distinguishing columns requiring second level suppression (1) versus not (0).
    """
    # Raise KeyError if either first_level_suppression or rank_field do not exist
    if first_level_suppression not in df.columns or rank_field not in df.columns:
        raise KeyError(
            f"'{first_level_suppression}' or '{rank_field}' does not exist in the DataFrame."
        )

    # Create the second level suppression field with default value of 0
    df[second_level_suppression] = 0

    # Iterate through the DataFrame starting from the 2nd row (index 1) as first row can't contain a 2 in rank_field
    for i in range(1, len(df)):
        # Check if 'Rank' is 2 and the previous rows 'first_level_suppression' is 1
        if df.at[i, rank_field] == 2 and df.at[i - 1, first_level_suppression] == 1:
            # Set second_level_suppression to 1 for the current row
            df.at[i, second_level_suppression] = 1
    return df


def add_suppression_required_from_upper_level_column(
    upper_level_df,
    lower_level_df,
    upper_level_suppression_column,
    code_lookup_field,
    suppression_lookup_field,
):
    """
    Function adding a new column to DataFrame with value of 1 for rows where upper level output shows suppression is required against
    lower level orgnaisations/sites/wards, resulting in suppression where 'Rank' column in the lower level DataFrame is 1
    (if it features for the Business Level Code) or 1 and 2 (if they both feature for the Business Level Code).
    Example - if an ICB Code is flagged as requiring suppression in the ICB level output, in the Trust level output, with all Trusts sorted in Rank order
    by ICB, the first Trust reporting more than 0 responses (the Rank 1 Trust for the ICB), and the Rank 2 Trust (if there is one) will be marked for suppresison
    at Trust level.

    Parameters:
    upper_level_df: DataFrame containing data aggregated to upper level where identifier codes are unique.
    lower_level_df: DataFrame containing data aggregated to lower level where identifier codes present in
                    upper level are not unique, where the new column will be added for suppression.
    upper_level_suppression_column: new field added to lower_level_df showing need for suppression.
    code_lookup_field: field from upper_level_df containing the organisation site/ward codes for dictionary key.
    suppression_lookup_field: field from upper_level_df containing suppression status for dictionary value.

    Returns:
     - lower_level_df: with the new suppression column added aligned with need to suppress from upper level DataFrame.
    """
    # Check if required columns exist in upper_level_df
    if (
        code_lookup_field not in upper_level_df.columns
        or suppression_lookup_field not in upper_level_df.columns
    ):
        logging.warning(
            f"Missing required columns '{code_lookup_field}' or '{suppression_lookup_field}' in upper_level_df. "
            f"No suppression will be applied from upper level."
        )
        # Initialise the new column in 'lower_level_df' with 0s and return
        lower_level_df[upper_level_suppression_column] = 0
        return lower_level_df

    # Check if required columns exist in lower_level_df
    if (
        code_lookup_field not in lower_level_df.columns
        or "Rank" not in lower_level_df.columns
    ):
        logging.warning(
            f"Missing required columns '{code_lookup_field}' or 'Rank' in lower_level_df. "
            f"No suppression will be applied from upper level."
        )
        # Initialise the new column in 'lower_level_df' with 0s and return
        lower_level_df[upper_level_suppression_column] = 0
        return lower_level_df

    # Create a dictionary from 'upper_level_df' with code_lookup_field as the key and suppression_lookup_field as the value
    suppression_dict = upper_level_df.set_index(code_lookup_field)[
        suppression_lookup_field
    ].to_dict()

    # Initialise the new column in 'lower_level_df' with 0s
    lower_level_df[upper_level_suppression_column] = 0

    # Iterate over each Business Level Code in lower_level_df
    for business_level_code in lower_level_df[code_lookup_field].unique():
        # Check if suppression is required for this Business Level Code based on upper_level_df
        if suppression_dict.get(business_level_code, 0) == 1:
            # Filter the rows with the current Business Level Code
            icb_rows = lower_level_df[
                lower_level_df[code_lookup_field] == business_level_code
            ]

            # Assign suppression for Rank 1 if it exists for the business level code
            if any(icb_rows["Rank"] == 1):
                lower_level_df.loc[
                    (lower_level_df[code_lookup_field] == business_level_code)
                    & (lower_level_df["Rank"] == 1),
                    upper_level_suppression_column,
                ] = 1

            # Assign suppression for Rank 2 if it exists for the business level code
            if any(icb_rows["Rank"] == 2):
                lower_level_df.loc[
                    (lower_level_df[code_lookup_field] == business_level_code)
                    & (lower_level_df["Rank"] == 2),
                    upper_level_suppression_column,
                ] = 1

    return lower_level_df


def join_dataframes(df1, df2, on="column_to_join_on", how="left", validate="one_to_one"):
    """
    Function to join one DataFrame with another (side by side).

    Parameters:
    df1: DataFrame containing one set of data.
    df2: DataFrame containing another set of data with aligned columns.

    Return:
    - df: with content of both source DataFrames joined into one DataFrame
    """
    # Raise KeyError if the join `on` column is missing from either DataFrame.
    if on not in df1.columns or on not in df2.columns:
        raise KeyError(f"Join column '{on}' not found in one of the DataFrames.")

    # Raise ValueError if an invalid join 'how' type is specified
    valid_how = ["left", "right", "inner", "outer"]
    if how not in valid_how:
        raise ValueError(f"Invalid join type '{how}'.")

    df = df1.join(df2.set_index(on), on, how, validate)

    return df


def replace_character_in_columns(df, columns, target_chars, replacement_char):
    """
    Function to replace specific character(s) with another character in the specified DataFrame column(s).

    Parameters:
    df: DataFrame where replacement is required.
    columns: Single column or list of columns where replacement might be needed.
    target_chars: The character(s) to search for in the specified column(s).
    replacement_char: The character to replace the target character.

    Returns:
    - df: modified with the specified character(s) replaced.
    """

    # If 'columns' is a single string (i.e., one column), convert it to a list for uniform handling
    if isinstance(columns, str):
        columns = [columns]

    # Filter columns to only include those that exist in the DataFrame
    available_columns = [column for column in columns if column in df.columns]

    # Log warning for missing columns
    missing_columns = [column for column in columns if column not in df.columns]
    if missing_columns:
        logging.warning(f"The following columns for character replacement do not exist in the DataFrame: {missing_columns}")

    # If no columns to process, return DataFrame unchanged
    if not available_columns:
        logging.warning("No columns available for character replacement. Returning DataFrame unchanged.")
        return df

    # Loop through each available column
    for column in available_columns:
        # Loop through each target character in the target_chars list
        for char in target_chars:
            # Replace the target character with the replacement character in the specified column
            df[column] = (
                df[column]
                .astype(str)
                .str.replace(char, replacement_char, regex=False)
            )

    return df


def remove_duplicate_rows(df):
    """
    Function remove duplicate rows in a DataFrame to retain only rows of unique values.

    Parameters:
    df: DataFrame where duplicate rows require removing.

    Returns:
    - df: with any duplicate rows removed.
    """

    return df.drop_duplicates()


def standardise_fft_column_names(df):
    """
    Function to standardise column names in FFT data files.
    Handles two common mappings:
    1. Maps "Parent org code"/"Parent name" to "STP Code"/"STP Name"
    2. Removes "SUM" suffix from response column names

    Parameters:
    df: DataFrame with raw column names from Excel import

    Returns:
    - df: DataFrame with standardised column names
    """
    # Define mappings
    parent_to_stp_mapping = {
        "Parent org code": "STP Code",
        "Parent name": "STP Name"
    }

    response_column_mapping = {
        "1 Very Good SUM": "1 Very Good",
        "2 Good SUM": "2 Good",
        "3 Neither Good nor Poor SUM": "3 Neither good nor poor",
        "4 Poor SUM": "4 Poor",
        "5 Very Poor SUM": "5 Very poor",
        "6 Dont Know SUM": "6 Dont Know",
        "Total Eligible SUM": "Total Eligible",
        "Prop_Pos": "Prop Pos",
        "Prop_Neg": "Prop Neg"
    }

    # Filter mappings to only include columns that exist in the DataFrame
    parent_columns = {k: v for k, v in parent_to_stp_mapping.items() if k in df.columns}
    response_columns = {k: v for k, v in response_column_mapping.items() if k in df.columns}

    # Combine mappings
    all_mappings = {**parent_columns, **response_columns}

    # Apply mappings if any exist
    if all_mappings:
        df = df.rename(columns=all_mappings)

    return df


def limit_retained_columns(existing_df, columns):
    """
    Function to retain column/columns of an existing DataFrame.
    Useful where the columns to retain are fewer than the columns to remove.

    Parameters:
    existing_df: DataFrame form which to filter columns to retain.
    columns: single or list of columns to be retained (e.g. ['ICB Code', 'ICB Name'] from Ward Level submissions).

    Returns:
    - df: with just the specified field(s) retained.
    """
    # If 'columns' is a single string (i.e., one column), convert it to a list for uniform handling
    if isinstance(columns, str):
        columns = [columns]

    # Raise TypeError if columns is not a string or list
    if not isinstance(columns, list):
        raise TypeError("columns should be a string or list of strings")

    # Check for missing columns and log a warning
    available_columns = [col for col in columns if col in existing_df.columns]
    missing_columns = [col for col in columns if col not in existing_df.columns]

    if missing_columns:
        logging.warning(f"The following columns to retain do not exist in the DataFrame: {missing_columns}")

    # If no columns are available, return an empty DataFrame with the same structure
    if not available_columns:
        logging.warning("None of the specified columns to retain exist in the DataFrame. Returning empty DataFrame.")
        return pd.DataFrame(columns=columns)

    # Return DataFrame with only the available columns
    return existing_df.filter(available_columns, axis=1)


def open_macro_excel_file(source_file_path):
    """
    Function to open existing Macro-Enabled Excel file where one exists to be used as a template to generate Output file from.

    Parameters:
    source_file_path: Path to the existing macro-enabled Excel file.

    Returns:
    - workbook: Loaded as object.
    """
    # Convert to Path object if it's a string
    file_path = Path(source_file_path)

    # Load the existing macro-enabled workbook retaining vba
    workbook = load_workbook(str(file_path), keep_vba=True)

    return workbook


def write_dataframes_to_sheets(workbook, dfs_info):
    """
    Function to write DataFrames to specified sheets and cells in a workbook already opened as an object.

    Parameters:
    workbook: The workbook object where DataFrames will be written.
    dfs_info: Dataframes information written in tuples. For each sheet the tuple will contain
    (DataFrame, Excel sheet name, start row, start column).

    Returns:
    - None: Workbook object is updated with content from Dataframes based on Tuples.
    """
    # Find the template path to use for column validation
    template_path = None
    for sheet_name in workbook.sheetnames:
        if "Note" in sheet_name:  # The Note sheet indicates this is an FFT template
            template_path = str(Path("inputs") / "template_files" / "FFT-inpatient-data-template.xlsm")
            logging.info(f"Using template file: {template_path}")
            break

    # Cache for template headers
    template_headers = {}

    # Iterate over the DataFrames and corresponding sheet and cell positions
    for df, sheet_name, start_row, start_col in dfs_info:
        # Select the sheet
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            raise ValueError(f"Sheet {sheet_name} does not exist in the workbook.")

        # Create a copy of the DataFrame to avoid modifying the original
        df_to_write = df.copy()

        # Remove the Title column if it exists to prevent concatenation issues
        if 'Title' in df_to_write.columns:
            df_to_write = df_to_write.drop(columns=['Title'])
            logging.info(f"Removed 'Title' column from DataFrame before writing to {sheet_name} sheet")

        # Load template headers for this sheet if available
        if template_path and sheet_name not in template_headers and sheet_name in ["Summary", "ICB", "Sites"]:
            try:
                # Read the template file to get headers
                import openpyxl
                template_wb = openpyxl.load_workbook(template_path, read_only=True, keep_vba=True)
                if sheet_name in template_wb.sheetnames:
                    template_sheet = template_wb[sheet_name]
                    headers = []

                    if sheet_name == "Summary":
                        # Check row 6 for Summary sheet headers
                        header_row = 6
                        for col in range(1, 15):
                            val = template_sheet.cell(row=header_row, column=col).value
                            if val:
                                headers.append((col, str(val)))
                    elif sheet_name in ["ICB", "Sites"]:
                        # For ICB and Sites sheets, we need to combine headers from rows 10 and 11
                        header_rows = [10, 11]
                        max_col = 20  # Increased to capture Don't Know column
                        for header_row in header_rows:
                            for col in range(1, max_col):
                                val = template_sheet.cell(row=header_row, column=col).value
                                if val and (col, str(val)) not in headers:
                                    headers.append((col, str(val)))

                    template_headers[sheet_name] = headers
                    logging.info(f"Found headers in template {sheet_name} sheet: {headers}")
            except Exception as e:
                logging.warning(f"Could not read template headers: {str(e)}")

        # Filter columns for sheets with template headers
        if sheet_name in ["Summary", "ICB", "Sites"] and sheet_name in template_headers:
            template_col_names = [h[1] for h in template_headers[sheet_name]]
            max_col_position = max([h[0] for h in template_headers[sheet_name]], default=0)
            logging.info(f"Maximum column position in {sheet_name} template: {max_col_position}")

            # These are the base column names without date prefixes
            columns_to_keep = []

            # For Summary sheet with date prefixes
            if sheet_name == "Summary":
                # First, handle exact matches (like 'FFT')
                for col in df_to_write.columns:
                    if col in template_col_names:
                        columns_to_keep.append(col)

                # Then handle columns with date prefixes (like 'Aug-25_Responses')
                for template_col in template_col_names:
                    if template_col != 'FFT':  # Skip FFT as it doesn't have date prefixes
                        # Find columns that end with the template column name
                        matching_cols = [c for c in df_to_write.columns
                                         if '_' in c and c.split('_', 1)[1] == template_col]
                        columns_to_keep.extend(matching_cols)

            # For ICB and Sites sheets (standard columns)
            elif sheet_name in ["ICB", "Sites"]:
                # Only remove problematic columns like Response Rate,
                # but keep all needed columns including Dont Know
                if 'Response Rate' in df_to_write.columns:
                    df_to_write = df_to_write.drop(columns=['Response Rate'])
                    logging.info(f"Removed 'Response Rate' column from {sheet_name} sheet")

                # Important: Don't perform any template-based column filtering
                # Just keep all columns as is after removing Response Rate
                columns_to_keep = list(df_to_write.columns)

            # Keep only the columns that match the template
            columns_to_drop = [col for col in df_to_write.columns if col not in columns_to_keep]

            if columns_to_drop:
                df_to_write = df_to_write.drop(columns=columns_to_drop)
                logging.info(f"Removed columns from {sheet_name} sheet to match template: {columns_to_drop}")

        # For other sheets, continue with the existing logic for Response Rate
        elif 'Response Rate' in df_to_write.columns:
            # Calculate where the Response Rate column would appear in the output
            response_rate_position = start_col + list(df_to_write.columns).index('Response Rate')

            # Identify problematic positions by sheet
            problematic_position = False
            if (sheet_name == 'Trusts' and response_rate_position >= 22) or \
               (sheet_name == 'Sites' and response_rate_position >= 16) or \
               (sheet_name == 'ICB' and response_rate_position >= 13) or \
               (sheet_name == 'Wards' and response_rate_position >= 17):
                problematic_position = True

            # Only remove if it would appear in a problematic position
            if problematic_position:
                df_to_write = df_to_write.drop(columns=['Response Rate'])
                logging.info(f"Removed 'Response Rate' column from DataFrame before writing to {sheet_name} sheet to prevent values outside table boundaries")

        # Iterate over DataFrame rows (i) and columns (j) to write data to Excel without headers
        for i, row in enumerate(df_to_write.itertuples(index=False, name=None), start=start_row):
            for j, value in enumerate(row, start=start_col):
                sheet.cell(row=i, column=j).value = value


def update_cell_with_formatting(
    workbook,
    sheet_name,
    start_row,
    start_col,
    data,
    font_size=10,
    bg_color="FFFFFF",
    bold=True,
    font_name="Verdana",
    align_horizontal="center",
    align_vertical="center",
):
    """
    Function to update a specific cell or range of cells with data and apply formatting.

    Parameters:
    workbook: The workbook object where the sheet resides.
    sheet_name: Name of the sheet to update.
    start_row: The starting row index (number/integer) for updating.
    start_col: The starting column index (number/integer) for the updating.
    data: Data to paste into the cell or range of cells.
    font_size: Font size (number/integer) for the cell(s).
    bg_color: Background color for the cell(s) in hex format (e.g., 'FFFFFF' for white; 'BFBFBF' for grey).
    bold:
    font_name: Font name for the cell(s) (e.g., "Calibri").
    align_horizontal: Horizontal alignment (e.g., "center", "left", "right").
    align_vertical: Vertical alignment (e.g., "center", "top", "bottom").

    Returns:
    - None: Workbook object is updated with specified content.
    """
    # Select the sheet
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        raise ValueError(f"Sheet {sheet_name} does not exist in the workbook.")

    # Check if the data is a list of lists (indicating multiple rows/columns)
    if isinstance(data, list):
        for i, row_data in enumerate(data):
            for j, value in enumerate(row_data):
                cell = sheet.cell(row=start_row + i, column=start_col + j)
                cell.value = value
                # Apply formatting
                cell.font = Font(size=font_size, bold=bold, name=font_name)
                cell.fill = PatternFill(
                    start_color=bg_color, end_color=bg_color, fill_type="solid"
                )
                cell.alignment = Alignment(
                    horizontal=align_horizontal, vertical=align_vertical
                )
    else:
        # Single cell update
        cell = sheet.cell(row=start_row, column=start_col)
        cell.value = data
        # Apply formatting
        cell.font = Font(size=font_size, bold=bold, name=font_name)
        cell.fill = PatternFill(
            start_color=bg_color, end_color=bg_color, fill_type="solid"
        )
        cell.alignment = Alignment(horizontal=align_horizontal, vertical=align_vertical)


def create_percentage_style(workbook):
    """
    Function to create a percentage style with 0 decimal places and add it to the workbook. NameStyles can
    only be defined and registered to a workbook once, so definition needs to be distinct from application.

    Parameters:
    workbook: The openpyxl workbook object.

    Returns:
    - percentage_style: The created NamedStyle.
    """
    # Raise TypeError if the the workbook is invalid object
    if not isinstance(workbook, Workbook):
        raise TypeError("The 'workbook' must be an openpyxl Workbook object.")

    # Check if the percentage style already exists to avoid re-creating it
    for style in workbook.named_styles:
        # Ensure we are dealing with NamedStyle objects and not strings
        if isinstance(style, NamedStyle) and style.name == "percentage_style":
            return style  # If style exists, return it

    # Create the percentage style if it does not exist
    percentage_style = NamedStyle(name="percentage_style")
    percentage_style.number_format = (
        "0%"  # Set number format for percentage with 0 decimal places
    )

    # Register the new style to the workbook
    workbook.add_named_style(percentage_style)

    return percentage_style


def format_column_as_percentage(
    workbook, sheet_name, start_row, start_cols, percentage_style
):
    """
    Function to format specified Excel sheet column(s) as percentages with 0 decimal places.

    Parameters:
    workbook: The workbook object where the sheet resides.
    sheet_name: The name of the sheet to update.
    start_row: The starting row index (number/integer) for foramtting.
    start_col: List of starting column indecies (number/integer) for formatting.

    Returns:
    - None: Workbook object is updated with specified formatting.
    """

    # Select the sheet
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        raise ValueError(f"Sheet {sheet_name} does not exist in this workbook.")

    # Apply formatting (percentage format with 0 decimal places) to specified columns starting from the specified row
    for start_col in start_cols:
        # Loop from the starting row to the last row
        for row in range(start_row, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=start_col)
            # Apply the percentage format with 0 decimal places
            cell.style = percentage_style


def combine_text_and_dataframe_cells(input_text_or_cell_1, input_text_or_cell_2):
    """
    Function to combine a specified string with the contents of a DataFrame cell to enable dynamic update.

    Parameters:
    input_text_or_cell_1: First text string or variable with cell content.
    input_text_or_cell_2: Second text string or variable with cell content.

    Returns:
    - combined_string: the text string and cell content combined into a string for further use.
    """

    # Combine the input text or cell content
    combined_string = f"{input_text_or_cell_1} {input_text_or_cell_2}"

    return combined_string


def save_macro_excel_file(
    workbook,
    source_file_path,
    new_folder_path,
    new=False,
    prefix=None,
    fft_period_suffix=None,
):
    """
    Save the workbook, either replacing the existing file or saving it as a new file with a specified name.

    Parameters:
    workbook: The workbook object to be saved.
    source_file_path: The path of the existing macro-enabled Excel file. Retain name used for loading in file.
    new_folder_path: The folder path assigned for saving a new macro-enabled Excel file.
    new: Boolean - If True, saves the file as a new file with a name generated by prefix and suffix.
         If False, it saves to the original source file path.
    prefix: Prefix for the name of the Excel file.
    fft_period_suffix = the fft period for which the output is being saved e.g., current_fft_period

    Returns:
    - None: Excel file saved to specified file path.
    """
    if new:
        # Construct the new file name
        new_file_name = f"{prefix}-{fft_period_suffix}.xlsm"
        # Convert to Path object and create the full path
        output_path = Path(new_folder_path) / new_file_name
        # Save as a new file
        workbook.save(str(output_path))
        logging.info(f"Workbook saved as {output_path}")
    else:
        # Replace the existing file
        workbook.save(source_file_path)
        logging.info(f"Workbook saved as {source_file_path}")

    # def check_current_month_in_rolling_totals(current_fft_period, df2):
    """
    Function to check if the current period already exists in df2.

    Parameters:
    current_fft_period: value generated for the current fft period.
    df2: Destination DataFrame where the data would be checked against (must have specific columns and format).

    Returns:
    - message stating if the period does not exist in df2.
    """

    # Check if the period from df1 already exists in df2


#    period = current_fft_period
#    if period in df2['FFT Period'].values:
#        raise ValueError(f"The period '{period}' already exists in df2.")
#    else:
#        return 'Current FFT Period not yet added to Monthly Rolling Totals.'
